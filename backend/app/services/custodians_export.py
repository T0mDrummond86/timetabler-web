"""Class custodians as a workbook, for circulating outside the app."""
from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from .global_sessions import aggregated_class_custodians

HEADERS = [
    "Class",
    "Units",
    "Used in sessions",
    "Linked qualifications",
    "Lecturers (deliveries)",
    "Custodian",
]

# Roughly the width each column needs; the lecturers column carries a list.
COLUMN_WIDTHS = [38, 30, 26, 34, 46, 24]


def write_custodians_xlsx(path: str | Path, rows: list[dict], *, title: str) -> Path:
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    # Excel refuses sheet names over 31 characters or containing []:*?/\
    ws.title = "Class custodians"

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(start_color="FF374151", end_color="FF374151", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="top", wrap_text=True)

    for col, heading in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=heading)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for r, row in enumerate(rows, start=2):
        sessions = row.get("session_names") or []
        values = [
            row.get("unit_name") or "",
            row.get("units") or "",
            ", ".join(sessions) if isinstance(sessions, list) else str(sessions),
            row.get("qualifications") or "",
            row.get("lecturers") or "",
            row.get("custodian") or "",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=r, column=col, value=value).alignment = body_align

    for col, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    # The workspace name belongs somewhere findable, but not in a data column.
    ws.oddHeader.left.text = title

    wb.save(path)
    return path


def _first_qualification(row: dict) -> str:
    raw = (row.get("qualifications") or "").strip()
    if not raw or raw == "—":
        return ""
    return raw.split(",")[0].strip()


def order_rows(rows: list[dict], order_by: str) -> list[dict]:
    """Match the on-screen ordering, so the download is what was being read.

    Sorts on the first linked qualification with the class name breaking ties,
    and drops classes linked to no qualification to the bottom rather than
    sorting them under an em dash.
    """
    if order_by != "qualification":
        return rows
    return sorted(
        rows,
        key=lambda r: (
            _first_qualification(r) == "",
            _first_qualification(r).casefold(),
            (r.get("unit_name") or "").casefold(),
        ),
    )


def export_class_custodians_xlsx(
    db: Session, *, global_session_id: int, title: str, order_by: str = "class"
) -> Path:
    report = aggregated_class_custodians(db, global_session_id)
    rows = order_rows(report.get("rows", []), order_by)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    return write_custodians_xlsx(tmp.name, rows, title=title)
