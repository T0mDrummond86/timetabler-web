"""Export selected qualifications' clusters (classes) and their units to xlsx.

One worksheet per qualification. In each sheet, column A lists the cluster
(class) names and column B the associated study-unit codes for that cluster.
"""
from __future__ import annotations

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from timetable.core.models import Qualification, Unit, UnitQualification

from .export_filenames import session_export_filename, timetable_session_name

_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Excel sheet names: <=31 chars, no []:*?/\\, unique within the book."""
    base = _INVALID_SHEET_CHARS.sub(" ", (name or "Qualification").strip()) or "Qualification"
    base = base[:31]
    candidate = base
    n = 2
    while candidate.lower() in used:
        suffix = f" ({n})"
        candidate = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def export_qualification_clusters(
    db: Session,
    *,
    timetable_session_id: int,
    qualification_ids: list[int],
) -> tuple[bytes, str]:
    quals = (
        db.query(Qualification)
        .filter(
            Qualification.timetable_session_id == timetable_session_id,
            Qualification.id.in_(qualification_ids or [-1]),
        )
        .order_by(Qualification.name)
        .all()
    )
    if not quals:
        raise ValueError("Select at least one qualification to export")

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    header_font = Font(bold=True)
    used_names: set[str] = set()

    for qual in quals:
        ws = wb.create_sheet(_safe_sheet_name(qual.name, used_names))
        ws["A1"] = "Cluster"
        ws["B1"] = "Units"
        ws["C1"] = "Hours"
        for col in ("A1", "B1", "C1"):
            ws[col].font = header_font

        units = (
            db.query(Unit)
            .join(UnitQualification, UnitQualification.unit_id == Unit.id)
            .filter(
                UnitQualification.qualification_id == qual.id,
                Unit.timetable_session_id == timetable_session_id,
            )
            .order_by(Unit.name)
            .all()
        )
        row = 2
        for unit in units:
            codes = (unit.component_codes or "").strip()
            # Normalise separators/spacing to a clean comma-joined list.
            codes = ", ".join(c.strip() for c in re.split(r"[,;/]", codes) if c.strip())
            # Each timetable slot is 30 minutes; length_slots covers both parts
            # of a double session, so total class hours = length_slots / 2.
            hours = unit.length_slots / 2 if unit.length_slots else None
            ws.cell(row=row, column=1, value=unit.name)
            ws.cell(row=row, column=2, value=codes)
            ws.cell(row=row, column=3, value=hours)
            row += 1

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 10

    buffer = io.BytesIO()
    wb.save(buffer)

    session_name = timetable_session_name(db, timetable_session_id)
    filename = session_export_filename(session_name, ".xlsx", label="qualification clusters")
    return buffer.getvalue(), filename
