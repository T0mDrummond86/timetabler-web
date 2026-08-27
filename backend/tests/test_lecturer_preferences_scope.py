"""The preferences workbook belongs to one session, not to the database.

On the desktop this could be taken for granted: one *.db file holds one
session, so an unscoped query was the session. The web app keeps every session
in one Postgres database, where the same query returns the whole organisation.

The export leaked: a template for one campus listed every lecturer, every
qualification and every class in the database. The import was worse -- a sheet
whose lecturer name happened to exist in another session matched it and wrote
the preference there.
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

import pytest

BACKEND = Path(__file__).resolve().parents[1]
DOMAIN = BACKEND.parent / "packages" / "domain"
sys.path.insert(0, str(BACKEND))
sys.path.insert(0, str(DOMAIN))

os.environ.setdefault("DATABASE_URL", "sqlite+pysqlite:///:memory:")
os.environ.setdefault("AUTO_CREATE_TABLES", "false")
os.environ.setdefault("JWT_SECRET", "test-secret")

from openpyxl import load_workbook  # noqa: E402
from sqlalchemy import create_engine  # noqa: E402
from sqlalchemy.orm import sessionmaker  # noqa: E402
from sqlalchemy.pool import StaticPool  # noqa: E402

from timetable.core.models import (  # noqa: E402
    Base,
    Qualification,
    Staff,
    StaffPreference,
    Unit,
    UnitQualification,
)
from timetable.core.tenancy_models import Organization, TimetableSession  # noqa: E402
from timetable.io.lecturer_preferences_import import import_lecturer_preferences  # noqa: E402
from timetable.io.lecturer_preferences_template import (  # noqa: E402
    CLASS_LIST_SHEET,
    write_lecturer_preferences_template,
)

HERE = 1
ELSEWHERE = 2


@pytest.fixture()
def db():
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    s = sessionmaker(bind=engine, expire_on_commit=False)()
    s.add(Organization(id=1, name="T", slug="t"))
    s.flush()
    s.add(TimetableSession(id=HERE, organization_id=1, name="Joondalup"))
    s.add(TimetableSession(id=ELSEWHERE, organization_id=1, name="Midland"))
    s.commit()
    try:
        yield s
    finally:
        s.close()


def _populate(db, sid: int, *, staff: str, qual: str, unit: str):
    st = Staff(timetable_session_id=sid, name=staff, fte=1.0)
    q = Qualification(timetable_session_id=sid, name=qual)
    u = Unit(timetable_session_id=sid, name=unit, length_slots=4)
    db.add_all([st, q, u])
    db.flush()
    db.add(UnitQualification(unit_id=u.id, qualification_id=q.id))
    db.commit()
    return st, q, u


def _write(db, tmp_path, sid: int | None) -> Path:
    out = tmp_path / "prefs.xlsx"
    write_lecturer_preferences_template(db, out, timetable_session_id=sid)
    return out


class TestTheTemplateOnlyShowsThisSession:
    def test_other_sessions_lecturers_get_no_tab(self, db, tmp_path):
        _populate(db, HERE, staff="A. Rivers", qual="Cert IV Cyber", unit="Threat data")
        _populate(db, ELSEWHERE, staff="B. Nakamura", qual="Dip Networking", unit="Routing")

        wb = load_workbook(_write(db, tmp_path, HERE))

        tabs = [t for t in wb.sheetnames if not t.startswith("_")]
        assert tabs == ["A. Rivers"]

    def test_other_sessions_qualifications_are_not_in_the_dropdown(self, db, tmp_path):
        _populate(db, HERE, staff="A. Rivers", qual="Cert IV Cyber", unit="Threat data")
        _populate(db, ELSEWHERE, staff="B. Nakamura", qual="Dip Networking", unit="Routing")

        ws = load_workbook(_write(db, tmp_path, HERE))[CLASS_LIST_SHEET]

        quals = [c.value for c in ws["A"][1:] if c.value]
        assert quals == ["Cert IV Cyber"]

    def test_other_sessions_classes_are_not_in_the_dropdown(self, db, tmp_path):
        _populate(db, HERE, staff="A. Rivers", qual="Cert IV Cyber", unit="Threat data")
        _populate(db, ELSEWHERE, staff="B. Nakamura", qual="Dip Networking", unit="Routing")

        ws = load_workbook(_write(db, tmp_path, HERE))[CLASS_LIST_SHEET]

        pairs = [
            (r[0], r[1])
            for r in ws.iter_rows(min_row=2, min_col=3, max_col=4, values_only=True)
            if r[0]
        ]
        assert pairs == [("Cert IV Cyber", "Threat data")]

    def test_a_session_with_no_staff_still_produces_a_workbook(self, db, tmp_path):
        _populate(db, ELSEWHERE, staff="B. Nakamura", qual="Dip Networking", unit="Routing")

        wb = load_workbook(_write(db, tmp_path, HERE))

        assert "(no staff)" in wb.sheetnames

    def test_without_a_session_id_the_whole_database_is_used(self, db, tmp_path):
        # The desktop's behaviour, where one file is one session. Kept working
        # so the desktop caller and its tests are unaffected.
        _populate(db, HERE, staff="A. Rivers", qual="Cert IV Cyber", unit="Threat data")
        _populate(db, ELSEWHERE, staff="B. Nakamura", qual="Dip Networking", unit="Routing")

        wb = load_workbook(_write(db, tmp_path, None))

        assert sorted(t for t in wb.sheetnames if not t.startswith("_")) == [
            "A. Rivers",
            "B. Nakamura",
        ]


class TestTheImportOnlyWritesToThisSession:
    def _filled(self, db, tmp_path, *, sid, staff, qual, unit) -> Path:
        """A template for `sid`, with one first preference filled in."""
        path = _write(db, tmp_path, sid)
        from openpyxl import load_workbook as lw

        wb = lw(path)
        ws = wb[staff]
        ws.cell(row=6, column=2, value=qual)
        ws.cell(row=6, column=3, value=unit)
        wb.save(path)
        return path

    def test_a_lecturer_of_the_same_name_elsewhere_is_not_matched(self, db, tmp_path):
        here = _populate(db, HERE, staff="A. Rivers", qual="Cert IV Cyber", unit="Threat data")
        there = _populate(
            db, ELSEWHERE, staff="A. Rivers", qual="Cert IV Cyber", unit="Threat data"
        )
        path = self._filled(
            db, tmp_path, sid=HERE, staff="A. Rivers", qual="Cert IV Cyber", unit="Threat data"
        )

        import_lecturer_preferences(db, path, timetable_session_id=HERE)
        db.commit()

        prefs = db.query(StaffPreference).all()
        assert len(prefs) == 1
        # The whole point: it landed on this session's lecturer, not the
        # identically named one in another session.
        assert prefs[0].staff_id == here[0].id
        assert prefs[0].staff_id != there[0].id
        assert prefs[0].unit_id == here[2].id

    def test_a_lecturer_only_in_another_session_is_reported_not_written(self, db, tmp_path):
        _populate(db, ELSEWHERE, staff="B. Nakamura", qual="Dip Networking", unit="Routing")
        path = self._filled(
            db, tmp_path, sid=ELSEWHERE, staff="B. Nakamura",
            qual="Dip Networking", unit="Routing",
        )

        rep = import_lecturer_preferences(db, path, timetable_session_id=HERE)
        db.commit()

        assert db.query(StaffPreference).count() == 0
        assert any("B. Nakamura" in w for w in rep.warnings)

    def test_a_class_only_in_another_session_is_not_linked(self, db, tmp_path):
        here = _populate(db, HERE, staff="A. Rivers", qual="Cert IV Cyber", unit="Threat data")
        there = _populate(
            db, ELSEWHERE, staff="B. Nakamura", qual="Dip Networking", unit="Routing"
        )
        path = self._filled(
            db, tmp_path, sid=HERE, staff="A. Rivers", qual="Dip Networking", unit="Routing"
        )

        import_lecturer_preferences(db, path, timetable_session_id=HERE)
        db.commit()

        for pref in db.query(StaffPreference).all():
            assert pref.staff_id == here[0].id
            # The name is kept as written so the warning is actionable, but it
            # must not point at another session's class row.
            assert pref.unit_id != there[2].id
