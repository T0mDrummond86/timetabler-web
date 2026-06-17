"""Phase 4 import/export, holding area, booking create/delete."""
from __future__ import annotations

import io
import json
import os
import sys
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

BACKEND = Path(__file__).resolve().parents[1]
DOMAIN = BACKEND.parent / "packages" / "domain"
sys.path.insert(0, str(BACKEND))
sys.path.insert(0, str(DOMAIN))

os.environ["DATABASE_URL"] = "sqlite+pysqlite:///:memory:"
os.environ["AUTO_CREATE_TABLES"] = "false"
os.environ["JWT_SECRET"] = "test-secret"

from timetable.core.models import Base  # noqa: E402
from timetable.io.backup_payload import BACKUP_SHEET_NAME, PAYLOAD_VERSION  # noqa: E402

from app.database import get_db  # noqa: E402
from app.main import app  # noqa: E402
from app.services.session_data import serialize_session  # noqa: E402


@pytest.fixture()
def client():
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, expire_on_commit=False)

    def override_get_db():
        db = SessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


def _auth(client: TestClient) -> tuple[str, int]:
    reg = client.post(
        "/auth/register",
        json={
            "username": "phase4",
            "password": "password123",
            "name": "Phase4",
            "organization_name": "Org",
        },
    )
    assert reg.status_code == 201
    token = reg.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}
    org_id = client.get("/orgs", headers=headers).json()[0]["id"]
    session_id = client.get(f"/orgs/{org_id}/sessions", headers=headers).json()[0]["id"]
    return token, session_id


def _seed(client: TestClient, session_id: int, headers: dict) -> tuple[int, int]:
    client.post(f"/sessions/{session_id}/seed-demo", headers=headers)
    course_id = client.get(f"/sessions/{session_id}/courses", headers=headers).json()[0]["id"]
    grid = client.get(
        f"/sessions/{session_id}/timetable",
        params={"course_id": course_id},
        headers=headers,
    ).json()
    return course_id, grid["bookings"][0]["id"]


def test_json_export_import_round_trip(client: TestClient):
    token, session_id = _auth(client)
    headers = {"Authorization": f"Bearer {token}"}
    _seed(client, session_id, headers)

    export = client.get(f"/sessions/{session_id}/export/json", headers=headers)
    assert export.status_code == 200
    payload = json.loads(export.content)
    assert payload["version"] == PAYLOAD_VERSION
    assert len(payload["bookings"]) == 2

    # Wipe via import of same payload (replace)
    imp = client.post(f"/sessions/{session_id}/import/json", headers=headers, json=payload)
    assert imp.status_code == 200, imp.text
    assert imp.json()["bookings"] == 2


def test_create_and_delete_booking(client: TestClient):
    token, session_id = _auth(client)
    headers = {"Authorization": f"Bearer {token}"}
    course_id, _booking_id = _seed(client, session_id, headers)

    holding = client.get(
        f"/sessions/{session_id}/holding-area",
        params={"course_id": course_id},
        headers=headers,
    )
    assert holding.status_code == 200

    units = client.get(f"/sessions/{session_id}/units", headers=headers).json()
    unit_id = units[0]["id"]

    create = client.post(
        f"/sessions/{session_id}/bookings",
        headers=headers,
        json={
            "course_id": course_id,
            "unit_id": unit_id,
            "day": 3,
            "start_slot": 12,
            "end_slot": 16,
        },
    )
    assert create.status_code == 200, create.text
    after = create.json()["change"]["after"]
    new_id = int(next(k for k, v in after.items() if v is not None))

    delete = client.delete(
        f"/sessions/{session_id}/bookings/{new_id}",
        params={"course_id": course_id},
        headers=headers,
    )
    assert delete.status_code == 200
    assert all(b["id"] != new_id for b in delete.json()["grid"]["bookings"])


def test_xlsm_import_with_backup_sheet(client: TestClient, tmp_path):
    token, session_id = _auth(client)
    headers = {"Authorization": f"Bearer {token}"}
    _seed(client, session_id, headers)

    # Build export payload from DB
    gen = app.dependency_overrides[get_db]()
    db = next(gen)
    payload = serialize_session(db, session_id)
    try:
        gen.close()
    except StopIteration:
        pass

    wb = Workbook()
    ws = wb.create_sheet(BACKUP_SHEET_NAME)
    ws["A1"] = "backup"
    ws["A2"] = json.dumps(payload, separators=(",", ":"))
    path = tmp_path / "test_export.xlsm"
    wb.save(path)
    wb.close()

    with path.open("rb") as fh:
        imp = client.post(
            f"/sessions/{session_id}/import",
            headers=headers,
            files={"file": ("export.xlsm", fh, "application/vnd.ms-excel.sheet.macroEnabled.12")},
        )
    assert imp.status_code == 200, imp.text
    assert imp.json()["bookings"] == 2


def test_import_into_second_session_avoids_global_id_clash(client: TestClient):
    """Desktop backup ids must be remapped — Postgres ids are global across sessions."""
    token, session_id_1 = _auth(client)
    headers = {"Authorization": f"Bearer {token}"}
    _seed(client, session_id_1, headers)

    export = client.get(f"/sessions/{session_id_1}/export/json", headers=headers)
    payload = json.loads(export.content)

    org_id = client.get("/orgs", headers=headers).json()[0]["id"]
    session_2 = client.post(
        f"/orgs/{org_id}/sessions",
        headers=headers,
        json={"name": "Second session"},
    )
    assert session_2.status_code == 201
    session_id_2 = session_2.json()["id"]

    imp = client.post(
        f"/sessions/{session_id_2}/import/json",
        headers=headers,
        json=payload,
    )
    assert imp.status_code == 200, imp.text
    assert imp.json()["bookings"] == 2

    courses = client.get(f"/sessions/{session_id_2}/courses", headers=headers).json()
    assert len(courses) == 1
    grid = client.get(
        f"/sessions/{session_id_2}/timetable",
        params={"course_id": courses[0]["id"]},
        headers=headers,
    )
    assert grid.status_code == 200
    assert len(grid.json()["bookings"]) == 2


def test_duplicate_session_save_as(client: TestClient):
    token, session_id = _auth(client)
    headers = {"Authorization": f"Bearer {token}"}
    _seed(client, session_id, headers)

    dup = client.post(
        f"/sessions/{session_id}/duplicate",
        headers=headers,
        json={"name": "Copied session"},
    )
    assert dup.status_code == 201, dup.text
    copy_id = dup.json()["id"]
    assert copy_id != session_id

    courses = client.get(f"/sessions/{copy_id}/courses", headers=headers).json()
    assert len(courses) == 1
    grid = client.get(
        f"/sessions/{copy_id}/timetable",
        params={"course_id": courses[0]["id"]},
        headers=headers,
    )
    assert grid.status_code == 200
    assert len(grid.json()["bookings"]) == 2

    clash = client.post(
        f"/sessions/{session_id}/duplicate",
        headers=headers,
        json={"name": "Copied session"},
    )
    assert clash.status_code == 409


def test_delete_session(client: TestClient):
    token, session_id = _auth(client)
    headers = {"Authorization": f"Bearer {token}"}
    org_id = client.get("/orgs", headers=headers).json()[0]["id"]

    created = client.post(
        f"/orgs/{org_id}/sessions",
        headers=headers,
        json={"name": "Disposable"},
    )
    assert created.status_code == 201
    disposable_id = created.json()["id"]

    deleted = client.delete(f"/sessions/{disposable_id}", headers=headers)
    assert deleted.status_code == 204

    sessions = client.get(f"/orgs/{org_id}/sessions", headers=headers).json()
    assert disposable_id not in {s["id"] for s in sessions}
