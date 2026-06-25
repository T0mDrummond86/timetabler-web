"""Admin visual import and admin export ID formatting (web)."""
from __future__ import annotations

import os
import sys
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

BACKEND = Path(__file__).resolve().parents[1]
DOMAIN = BACKEND.parent / "packages" / "domain"
sys.path.insert(0, str(BACKEND))
sys.path.insert(0, str(DOMAIN))

os.environ["DATABASE_URL"] = "sqlite+pysqlite:///:memory:"
os.environ["AUTO_CREATE_TABLES"] = "false"
os.environ["JWT_SECRET"] = "test-secret"

from timetable.core.models import Base  # noqa: E402
from timetable.io.admin_export import (  # noqa: E402
    TERM1_LABEL_COLS,
    TERM1_WEEK_COLS,
    TERM2_LABEL_COLS,
    TERM2_WEEK_COLS,
    _COURSE_TITLE_COL,
    _COURSE_TITLE_ROW,
    write_admin_export,
)

from app.database import get_db  # noqa: E402
from app.main import app  # noqa: E402


@pytest.fixture()
def client():
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, expire_on_commit=False)

    def override_get_db():
        db = SessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


def _auth(client: TestClient) -> tuple[str, int]:
    reg = client.post(
        "/auth/register",
        json={
            "username": "admin-visual",
            "password": "password123",
            "name": "Admin Visual",
            "organization_name": "Org",
        },
    )
    assert reg.status_code == 201
    token = reg.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}
    org_id = client.get("/orgs", headers=headers).json()[0]["id"]
    session_id = client.get(f"/orgs/{org_id}/sessions", headers=headers).json()[0]["id"]
    return token, session_id


def _headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _admin_visual_fixture(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "TEST1"
    ws.cell(_COURSE_TITLE_ROW, _COURSE_TITLE_COL, value="Course: TEST1")
    row = 19
    ws.cell(row, TERM1_LABEL_COLS[0], value="09:00-12:00")
    ws.cell(row, TERM1_LABEL_COLS[1], value="Alex Teacher")
    ws.cell(row, TERM1_LABEL_COLS[2], value="A101")
    label = "Python (ICTPRG435, ICTPRG434)  ID 4393283"
    ws.merge_cells(
        start_row=row,
        end_row=row,
        start_column=TERM1_WEEK_COLS[0],
        end_column=TERM1_WEEK_COLS[-1],
    )
    ws.cell(row, TERM1_WEEK_COLS[0], value=label)
    path = tmp_path / "admin_id_no_colon.xlsx"
    wb.save(path)
    return path


def test_admin_visual_import_sets_external_id(client: TestClient, tmp_path: Path):
    token, session_id = _auth(client)
    path = _admin_visual_fixture(tmp_path)
    with path.open("rb") as fh:
        res = client.post(
            f"/sessions/{session_id}/import/admin-visual",
            headers=_headers(token),
            files={"file": ("admin.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["bookings_written"] == 1
    assert body["source"] == "admin_visual"

    courses = client.get(f"/sessions/{session_id}/courses", headers=_headers(token)).json()
    course_id = next(c["id"] for c in courses if c["code"] == "TEST1")
    units = client.get(f"/sessions/{session_id}/units", headers=_headers(token)).json()
    python = next(u for u in units if u["name"] == "Python")
    assert "ID" not in python["name"]
    assert "4393283" not in python["name"]

    grid = client.get(
        f"/sessions/{session_id}/timetable",
        params={"course_id": course_id},
        headers=_headers(token),
    ).json()
    ext_ids = {b["external_id"] for b in grid["bookings"] if b.get("external_id")}
    assert ext_ids == {"4393283"}


def test_admin_visual_import_tolerates_duplicate_unit_names_across_sessions(
    client: TestClient, tmp_path: Path
):
    """Same class name in another org session must not break visual import."""
    reg_a = client.post(
        "/auth/register",
        json={
            "username": "admin-visual-a",
            "password": "password123",
            "name": "A",
            "organization_name": "Org A",
        },
    )
    assert reg_a.status_code == 201
    token_a = reg_a.json()["access_token"]
    headers_a = _headers(token_a)
    org_a = client.get("/orgs", headers=headers_a).json()[0]["id"]
    session_a = client.get(f"/orgs/{org_a}/sessions", headers=headers_a).json()[0]["id"]

    reg_b = client.post(
        "/auth/register",
        json={
            "username": "admin-visual-b",
            "password": "password123",
            "name": "B",
            "organization_name": "Org B",
        },
    )
    assert reg_b.status_code == 201
    token_b = reg_b.json()["access_token"]
    headers_b = _headers(token_b)
    org_b = client.get("/orgs", headers=headers_b).json()[0]["id"]
    session_b = client.get(f"/orgs/{org_b}/sessions", headers=headers_b).json()[0]["id"]

    for sid, headers in ((session_a, headers_a), (session_b, headers_b)):
        client.post(f"/sessions/{sid}/units", headers=headers, json={"name": "Python"})

    path = _admin_visual_fixture(tmp_path)
    with path.open("rb") as fh:
        res = client.post(
            f"/sessions/{session_b}/import/admin-visual",
            headers=headers_b,
            files={"file": ("admin.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert res.status_code == 200, res.text
    assert res.json()["bookings_written"] == 1


def test_admin_export_appends_event_id_to_week_cell(client: TestClient, tmp_path: Path):
    token, session_id = _auth(client)
    headers = _headers(token)

    staff = client.post(
        f"/sessions/{session_id}/staff",
        headers=headers,
        json={"name": "Alice"},
    ).json()
    room = client.post(
        f"/sessions/{session_id}/rooms",
        headers=headers,
        json={"code": "A101"},
    ).json()
    unit = client.post(
        f"/sessions/{session_id}/units",
        headers=headers,
        json={"name": "Python"},
    ).json()
    course = client.post(
        f"/sessions/{session_id}/courses",
        headers=headers,
        json={"code": "ID1"},
    ).json()
    created = client.post(
        f"/sessions/{session_id}/bookings",
        headers=headers,
        json={
            "course_id": course["id"],
            "unit_id": unit["id"],
            "staff_id": staff["id"],
            "room_id": room["id"],
            "day": 0,
            "start_slot": 2,
            "end_slot": 6,
            "in_term_1": 1,
            "in_term_2": 0,
        },
    )
    assert created.status_code == 200, created.text
    booking_id = created.json()["grid"]["bookings"][0]["id"]
    patched = client.patch(
        f"/sessions/{session_id}/bookings/{booking_id}",
        headers=headers,
        json={"course_id": course["id"], "external_id": "4393283"},
    )
    assert patched.status_code == 200, patched.text
    assert patched.json()["grid"]["bookings"][0]["external_id"] == "4393283"
    client.patch(
        f"/sessions/{session_id}/units/{unit['id']}",
        headers=headers,
        json={"component_codes": "ICTPRG435, ICTPRG434"},
    )

    res = client.get(f"/sessions/{session_id}/export/admin", headers=headers)
    assert res.status_code == 200
    out = tmp_path / "admin_export_id.xlsx"
    out.write_bytes(res.content)
    wb = load_workbook(out)
    ws = wb[wb.sheetnames[0]]
    assert ws["E11"].value == "Python (ICTPRG435, ICTPRG434) ID: 4393283"
    wb.close()


def test_admin_visual_import_splits_sfs_co_teacher(client: TestClient, tmp_path: Path):
    token, session_id = _auth(client)
    wb = Workbook()
    ws = wb.active
    ws.title = "CO1"
    ws.cell(_COURSE_TITLE_ROW, _COURSE_TITLE_COL, value="Course: CO1")
    row = 19
    ws.cell(row, TERM1_LABEL_COLS[0], value="09:00-12:00")
    ws.cell(row, TERM1_LABEL_COLS[1], value="Alice Teacher + Bob Co-Teacher")
    ws.cell(row, TERM1_LABEL_COLS[2], value="A101")
    label = "Team Teach (ICTPRG435)"
    ws.merge_cells(
        start_row=row,
        end_row=row,
        start_column=TERM1_WEEK_COLS[0],
        end_column=TERM1_WEEK_COLS[-1],
    )
    ws.cell(row, TERM1_WEEK_COLS[0], value=label)
    path = tmp_path / "admin_co_teacher.xlsx"
    wb.save(path)

    with path.open("rb") as fh:
        res = client.post(
            f"/sessions/{session_id}/import/admin-visual",
            headers=_headers(token),
            files={"file": ("admin.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert res.status_code == 200, res.text
    assert res.json()["bookings_written"] == 1

    staff = client.get(f"/sessions/{session_id}/staff", headers=_headers(token)).json()
    staff_by_name = {s["name"]: s["id"] for s in staff}
    assert "Alice Teacher" in staff_by_name
    assert "Bob Co-Teacher" in staff_by_name
    assert "Alice Teacher + Bob Co-Teacher" not in staff_by_name

    courses = client.get(f"/sessions/{session_id}/courses", headers=_headers(token)).json()
    course_id = next(c["id"] for c in courses if c["code"] == "CO1")
    grid = client.get(
        f"/sessions/{session_id}/timetable",
        params={"course_id": course_id},
        headers=_headers(token),
    ).json()
    booking = grid["bookings"][0]
    assert booking["staff_name"] == "Alice Teacher"
    assert booking["sfs_co_teacher_staff_id"] == staff_by_name["Bob Co-Teacher"]
    assert booking["sfs_co_teacher_name"] == "Bob Co-Teacher"


def test_admin_visual_import_term2_only_row(client: TestClient, tmp_path: Path):
    """Term-2-only classes use term-2 TIME/Lecturer/Room columns (term 1 blank)."""
    token, session_id = _auth(client)
    wb = Workbook()
    ws = wb.active
    ws.title = "T2ONLY"
    ws.cell(_COURSE_TITLE_ROW, _COURSE_TITLE_COL, value="Course: T2ONLY")
    row = 10
    ws.cell(row, TERM2_LABEL_COLS[0], value="09:30-12:00")
    ws.cell(row, TERM2_LABEL_COLS[1], value="Jeff Hoyle")
    ws.cell(row, TERM2_LABEL_COLS[2], value="A114")
    label = "Cyber lab (VU23220) ID: 4401542"
    ws.merge_cells(
        start_row=row,
        end_row=row,
        start_column=TERM2_WEEK_COLS[0],
        end_column=TERM2_WEEK_COLS[-1],
    )
    ws.cell(row, TERM2_WEEK_COLS[0], value=label)
    path = tmp_path / "admin_term2_only.xlsx"
    wb.save(path)

    with path.open("rb") as fh:
        res = client.post(
            f"/sessions/{session_id}/import/admin-visual",
            headers=_headers(token),
            files={"file": ("admin.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert res.status_code == 200, res.text
    assert res.json()["bookings_written"] == 1

    courses = client.get(f"/sessions/{session_id}/courses", headers=_headers(token)).json()
    course_id = next(c["id"] for c in courses if c["code"] == "T2ONLY")
    grid = client.get(
        f"/sessions/{session_id}/timetable",
        params={"course_id": course_id, "term": "t2"},
        headers=_headers(token),
    ).json()
    assert len(grid["bookings"]) == 1
    booking = grid["bookings"][0]
    assert booking["in_term_1"] is False
    assert booking["in_term_2"] is True
    assert booking["external_id"] == "4401542"


def test_admin_visual_import_rejects_non_admin_workbook(client: TestClient, tmp_path: Path):
    token, session_id = _auth(client)
    wb = Workbook()
    path = tmp_path / "not_admin.xlsx"
    wb.save(path)
    with path.open("rb") as fh:
        res = client.post(
            f"/sessions/{session_id}/import/admin-visual",
            headers=_headers(token),
            files={"file": ("bad.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert res.status_code == 422
