"""End-to-end export tests: class names in cells, v2 workbook opens."""
from __future__ import annotations

import io
import os
import sys
import zipfile
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

BACKEND = Path(__file__).resolve().parents[1]
DOMAIN = BACKEND.parent / "packages" / "domain"
sys.path.insert(0, str(BACKEND))
sys.path.insert(0, str(DOMAIN))

os.environ["DATABASE_URL"] = "sqlite+pysqlite:///:memory:"
os.environ["AUTO_CREATE_TABLES"] = "false"
os.environ["JWT_SECRET"] = "test-secret"

from timetable.core.models import Base  # noqa: E402
from timetable.io.xlsm_export_v2 import V2_TEMPLATE_PATH  # noqa: E402

from app.database import get_db  # noqa: E402
from app.main import app  # noqa: E402


@pytest.fixture()
def client():
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, expire_on_commit=False)

    def override_get_db():
        db = SessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


def _register(client: TestClient) -> tuple[str, int]:
    reg = client.post(
        "/auth/register",
        json={
            "username": "exports",
            "password": "password123",
            "name": "Export Tester",
            "organization_name": "Export Org",
        },
    )
    assert reg.status_code == 201
    token = reg.json()["access_token"]
    org_id = client.get("/orgs", headers={"Authorization": f"Bearer {token}"}).json()[0]["id"]
    session_id = client.get(
        f"/orgs/{org_id}/sessions",
        headers={"Authorization": f"Bearer {token}"},
    ).json()[0]["id"]
    return token, session_id


def _headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _sheet_text(wb, sheet_name: str) -> str:
    if sheet_name not in wb.sheetnames:
        return ""
    ws = wb[sheet_name]
    parts: list[str] = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is not None and str(v).strip():
                parts.append(str(v))
    return "\n".join(parts)


def test_v2_template_bundled():
    assert V2_TEMPLATE_PATH.is_file(), f"missing template: {V2_TEMPLATE_PATH}"


def test_exports_include_class_names_and_v2_opens(client: TestClient):
    token, session_id = _register(client)
    headers = _headers(token)
    client.post(f"/sessions/{session_id}/seed-demo", headers=headers)

    class_name = "Cyber Foundations"

    for variant, suffix in (("fresh", ".xlsx"), ("v2", ".xlsm")):
        res = client.get(
            f"/sessions/{session_id}/export/timetable",
            headers=headers,
            params={"variant": variant},
        )
        assert res.status_code == 200, res.text
        data = res.content
        assert len(data) > 5000, f"{variant} export too small"
        assert data[:2] == b"PK", f"{variant} not a zip/xlsx container"

        wb = load_workbook(io.BytesIO(data), read_only=True, keep_vba=(suffix == ".xlsm"))
        try:
            assert wb.sheetnames, f"{variant} has no sheets"
            body = "\n".join(_sheet_text(wb, n) for n in wb.sheetnames)
            assert class_name in body, f"{variant} export missing class name {class_name!r}"
        finally:
            wb.close()

    admin = client.get(f"/sessions/{session_id}/export/admin", headers=headers)
    assert admin.status_code == 200, admin.text
    wb = load_workbook(io.BytesIO(admin.content), read_only=True)
    try:
        body = "\n".join(_sheet_text(wb, n) for n in wb.sheetnames)
        assert class_name in body, "admin export missing class name"
    finally:
        wb.close()


def test_export_scoped_to_session_not_other_org_data(client: TestClient):
    """Second session with no bookings must not leak first session's classes into export."""
    token, session_a = _register(client)
    headers = _headers(token)
    client.post(f"/sessions/{session_a}/seed-demo", headers=headers)

    org_id = client.get("/orgs", headers=headers).json()[0]["id"]
    session_b = client.post(
        f"/orgs/{org_id}/sessions",
        headers=headers,
        json={"name": "Empty session"},
    ).json()["id"]

    res = client.get(
        f"/sessions/{session_b}/export/timetable",
        headers=headers,
        params={"variant": "fresh"},
    )
    assert res.status_code == 200, res.text
    wb = load_workbook(io.BytesIO(res.content), read_only=True)
    try:
        body = "\n".join(_sheet_text(wb, n) for n in wb.sheetnames)
        assert "Cyber Foundations" not in body
    finally:
        wb.close()
