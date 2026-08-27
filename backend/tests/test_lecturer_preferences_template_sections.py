"""The sections a lecturer fills in, and the rows the importer counts on.

The workbook is a form a human fills in by hand, so its layout is part of its
contract twice over: the importer reads preferences and the non-teaching day
from fixed rows, and finds the blocked-times grid by its time labels. Adding a
section pushes the grid down the sheet, which is exactly the change that would
quietly break the round trip if the grid were ever read by row number instead.
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

import pytest

BACKEND = Path(__file__).resolve().parents[1]
DOMAIN = BACKEND.parent / "packages" / "domain"
sys.path.insert(0, str(BACKEND))
sys.path.insert(0, str(DOMAIN))

os.environ.setdefault("DATABASE_URL", "sqlite+pysqlite:///:memory:")
os.environ.setdefault("AUTO_CREATE_TABLES", "false")
os.environ.setdefault("JWT_SECRET", "test-secret")

from openpyxl import load_workbook  # noqa: E402
from sqlalchemy import create_engine  # noqa: E402
from sqlalchemy.orm import sessionmaker  # noqa: E402
from sqlalchemy.pool import StaticPool  # noqa: E402

from timetable.constants import NUM_DAYS  # noqa: E402
from timetable.core.models import (  # noqa: E402
    Base,
    Qualification,
    Staff,
    StaffAvailability,
    StaffPreference,
    Unit,
    UnitQualification,
)
from timetable.core.tenancy_models import Organization, TimetableSession  # noqa: E402
from timetable.io.lecturer_preferences_import import import_lecturer_preferences  # noqa: E402
from timetable.io.lecturer_preferences_template import (  # noqa: E402
    DEFAULT_REQUESTED_HOURS,
    write_lecturer_preferences_template,
)

SID = 1


@pytest.fixture()
def db():
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    s = sessionmaker(bind=engine, expire_on_commit=False)()
    s.add(Organization(id=1, name="T", slug="t"))
    s.flush()
    s.add(TimetableSession(id=SID, organization_id=1, name="Joondalup"))
    s.commit()
    q = Qualification(timetable_session_id=SID, name="Cert IV Cyber")
    st = Staff(timetable_session_id=SID, name="A. Rivers")
    s.add_all([q, st])
    s.flush()
    for name in ("Threat data", "Network hardening"):
        u = Unit(timetable_session_id=SID, name=name, length_slots=4)
        s.add(u)
        s.flush()
        s.add(UnitQualification(unit_id=u.id, qualification_id=q.id))
    s.commit()
    try:
        yield s
    finally:
        s.close()


def _sheet(db, tmp_path: Path):
    out = tmp_path / "prefs.xlsx"
    write_lecturer_preferences_template(db, out, timetable_session_id=SID)
    return out, load_workbook(out)["A. Rivers"]


def _hours_cell_row(ws) -> int:
    """The input sits under its heading, which has the row to itself."""
    return _find_label(ws, "Delivery hours requested") + 1


def _find_label(ws, text: str) -> int:
    for row in range(1, 60):
        if str(ws.cell(row=row, column=1).value or "").strip().lower().startswith(text.lower()):
            return row
    raise AssertionError(f"no section labelled {text!r} in column A")


class TestNewSections:
    def test_delivery_hours_section_is_prefilled_with_a_full_load(self, db, tmp_path):
        _, ws = _sheet(db, tmp_path)

        # Pre-filled, not blank: nearly everyone wants exactly this, so the
        # section costs most lecturers no typing at all.
        assert ws.cell(row=_hours_cell_row(ws), column=2).value == DEFAULT_REQUESTED_HOURS
        assert DEFAULT_REQUESTED_HOURS == 21

    def test_delivery_hours_cell_refuses_a_nonsense_number(self, db, tmp_path):
        _, ws = _sheet(db, tmp_path)

        coord = ws.cell(row=_hours_cell_row(ws), column=2).coordinate
        ranges = [dv for dv in ws.data_validations.dataValidation if coord in dv.sqref]
        assert ranges, "the hours cell has no validation"
        assert ranges[0].type == "decimal"

    def test_section_headings_have_the_row_to_themselves(self, db, tmp_path):
        _, ws = _sheet(db, tmp_path)

        # Column A is 14 wide because the grid's time labels have to fit, so a
        # heading with anything beside it is clipped to half a word.
        for label in ("Delivery hours requested", "Additional notes"):
            row = _find_label(ws, label)
            beside = [ws.cell(row=row, column=c).value for c in range(2, 7)]
            assert not any(beside), f"{label!r} has {beside} beside it and will be clipped"

    def test_notes_section_is_one_box_big_enough_to_write_in(self, db, tmp_path):
        _, ws = _sheet(db, tmp_path)

        row = _find_label(ws, "Additional notes")
        merged = [r for r in ws.merged_cells.ranges if r.min_row > row]
        assert merged, "the notes section has no box to write in"
        box = merged[0]
        # Directly under the heading and its one-line hint.
        assert box.min_row == row + 2
        assert box.max_row - box.min_row >= 3  # several lines, not one
        assert box.max_col >= 1 + NUM_DAYS  # as wide as the grid below it
        assert ws.cell(row=box.min_row, column=1).alignment.wrap_text

    def test_the_new_sections_sit_between_the_day_and_the_grid(self, db, tmp_path):
        _, ws = _sheet(db, tmp_path)

        day = _find_label(ws, "Non-teaching day")
        hours = _find_label(ws, "Delivery hours requested")
        notes = _find_label(ws, "Additional notes")
        grid = _find_label(ws, "Blocked times")
        assert day < hours < notes < grid

    def test_the_cells_the_importer_reads_by_position_have_not_moved(self, db, tmp_path):
        _, ws = _sheet(db, tmp_path)

        # Three positions are load-bearing. Everything else on the sheet may be
        # laid out freely, so these are worth asserting on their own rather
        # than inferring from wherever a heading happens to sit.
        assert str(ws["A1"].value or "").startswith("Preferences —")
        assert [ws.cell(row=r, column=1).value for r in range(6, 12)] == [
            "First",
            "First",
            "Second",
            "Second",
            "Third",
            "Third",
        ]
        # The non-teaching day is read from B14, whatever labels it.
        day_dv = [
            dv
            for dv in ws.data_validations.dataValidation
            if "B14" in dv.sqref and dv.type == "list"
        ]
        assert day_dv, "B14 is not the non-teaching day dropdown"
        assert "Monday" in day_dv[0].formula1


class TestTheRoundTripStillWorks:
    def test_a_filled_in_workbook_imports_with_the_grid_further_down(self, db, tmp_path):
        path, ws = _sheet(db, tmp_path)
        wb = load_workbook(path)
        sheet = wb["A. Rivers"]

        sheet["B6"] = "Cert IV Cyber"
        sheet["C6"] = "Threat data"
        sheet["B14"] = "Wednesday"
        sheet.cell(row=_hours_cell_row(ws), column=2).value = 18
        notes_box = [r for r in ws.merged_cells.ranges][0]
        sheet.cell(row=notes_box.min_row, column=1).value = (
            "Job-share Fridays; away for study in weeks 5-6."
        )
        # Block Monday's first hour, wherever the grid now starts.
        grid_first = next(
            row
            for row in range(1, 80)
            if str(sheet.cell(row=row, column=1).value or "").startswith("08:00")
        )
        sheet.cell(row=grid_first, column=2).value = "X"
        sheet.cell(row=grid_first + 1, column=2).value = "X"
        wb.save(path)

        rep = import_lecturer_preferences(db, path, timetable_session_id=SID)

        assert rep.warnings == []
        assert rep.staff_updated == 1
        assert rep.preferences_imported == 1
        staff = db.query(Staff).one()
        assert staff.non_teaching_day == 2  # Wednesday
        pref = db.query(StaffPreference).one()
        assert (pref.priority, pref.class_name) == (1, "Threat data")
        monday = [a for a in db.query(StaffAvailability).all() if a.day == 0]
        # Blocked 08:00-09:00, so Monday starts at slot 2.
        assert [(a.start_slot, a.end_slot) for a in monday] == [(2, 28)]

    def test_free_text_notes_do_not_look_like_a_grid_row(self, db, tmp_path):
        path, ws = _sheet(db, tmp_path)
        wb = load_workbook(path)
        sheet = wb["A. Rivers"]
        notes_box = [r for r in ws.merged_cells.ranges][0]
        # A time range written in the notes is the one thing that could be
        # mistaken for a blocked-times row, since that is how the grid is found.
        sheet.cell(row=notes_box.min_row, column=1).value = "Please avoid 09:00–09:30 on Mondays"
        wb.save(path)

        import_lecturer_preferences(db, path, timetable_session_id=SID)

        assert db.query(StaffAvailability).count() == 0


class TestHowTheFormLooks:
    """It goes out to people who do not use the app, so this is not decoration.

    Gridlines are off and the answer cells are the only ruled, shaded things on
    the page — that pairing is what makes "where do I type" answerable at a
    glance, and either half alone would undo it.
    """

    def test_sheet_gridlines_are_off(self, db, tmp_path):
        _, ws = _sheet(db, tmp_path)

        assert ws.sheet_view.showGridLines is False

    def test_there_are_no_frozen_panes(self, db, tmp_path):
        _, ws = _sheet(db, tmp_path)

        assert not ws.freeze_panes

    def test_every_answer_cell_is_shaded_and_boxed(self, db, tmp_path):
        _, ws = _sheet(db, tmp_path)

        answer_cells = [
            ws.cell(row=r, column=c) for r in range(6, 12) for c in (2, 3)
        ] + [
            ws.cell(row=14, column=2),
            ws.cell(row=_hours_cell_row(ws), column=2),
            ws.cell(row=_find_label(ws, "Additional notes") + 2, column=1),
        ]
        for cell in answer_cells:
            assert cell.fill.fill_type == "solid", f"{cell.coordinate} is not shaded"
            assert cell.border.left.style, f"{cell.coordinate} has no box round it"

    def test_a_label_is_not_dressed_up_as_an_answer_cell(self, db, tmp_path):
        _, ws = _sheet(db, tmp_path)

        # The priority column is filled in for you; shading it like the cells
        # beside it would invite people to type over it.
        answer = ws.cell(row=6, column=2).fill.start_color.rgb
        assert ws.cell(row=6, column=1).fill.start_color.rgb != answer

    def test_the_title_reads_as_a_banner(self, db, tmp_path):
        _, ws = _sheet(db, tmp_path)

        title = ws["A1"]
        assert any(r.min_row == 1 and r.max_col >= 1 + NUM_DAYS
                   for r in ws.merged_cells.ranges)
        assert title.font.color.rgb.endswith("FFFFFF")
        assert title.fill.fill_type == "solid"

    def test_the_grid_is_banded_by_the_hour(self, db, tmp_path):
        _, ws = _sheet(db, tmp_path)

        first = next(
            r for r in range(1, 80)
            if str(ws.cell(row=r, column=1).value or "").startswith("08:00")
        )
        shade = lambda r: ws.cell(row=r, column=2).fill.start_color.rgb  # noqa: E731
        # Both halves of an hour share a shade; the next hour differs. Counting
        # half-hour lines across five columns is what this saves.
        assert shade(first) == shade(first + 1)
        assert shade(first) != shade(first + 2)
        assert shade(first + 2) == shade(first + 3)

    def test_the_day_columns_are_all_the_same_width(self, db, tmp_path):
        _, ws = _sheet(db, tmp_path)

        widths = {
            ws.column_dimensions[chr(ord("B") + d)].width for d in range(NUM_DAYS)
        }
        assert len(widths) == 1, f"the grid is lopsided: {widths}"

    def test_it_is_set_up_to_print_on_one_page_wide(self, db, tmp_path):
        _, ws = _sheet(db, tmp_path)

        assert ws.page_setup.fitToWidth == 1
        assert ws.sheet_properties.pageSetUpPr.fitToPage is True
        assert ws.print_area
