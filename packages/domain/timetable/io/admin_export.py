"""Export timetable into the admin workbook format (template-driven layout).

Use ``adminExportStyleGuide.xlsx`` (or ``templates/admin_export_base.xlsx``). The
legacy OneDrive admin template is not compatible.

This module writes **values and class fills** into these fixed ranges. Cells for
classes with a net timetabling change and a class-card event id (``external_id``)
are highlighted red on TIME / Lecturer / Room (and day headers when the weekday
moved), matching the resolved change log.

* **B9:N14**, **B19:N23**, **B28:N32**, **B37:N41**, **B46:N50**
* **P9:R14**, **P19:R23**, **P28:R32**, **P37:R41**, **P46:R50**
* **S9:AB14**, **S19:AB23**, **S28:AB32**, **S37:AB41**, **S46:AB50** (term-4 weeks)

Rows 17–18, 26–27, etc. are headers/dates — never written. Columns **O**, **AC**,
**AD**, **AE** are holidays — never written.
"""
from __future__ import annotations

import colorsys
from typing import Literal
import hashlib
import shutil
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill
from sqlalchemy.orm import Session, joinedload

from ..bundle_paths import resource_path
from ..constants import DAYS, NUM_DAYS, slot_to_time
from ..core.booking_staff import sfs_co_teacher_booking_filter
from ..core.booking_sessions import active_session_weeks_in_term
from ..core.models import Booking, Course, Week
from .backup_payload import write_backup_sheet
from .xlsm_export import _component_codes_line

ADMIN_TEMPLATE_PATH = resource_path("templates", "admin_export_base.xlsx")
ADMIN_STYLE_GUIDE_PATH = resource_path("adminExportStyleGuide.xlsx")
ADMIN_TEMPLATE_SHEET = "Course Template"
ADMIN_TEMPLATE_SHEET_ALIASES = (ADMIN_TEMPLATE_SHEET, "Sheet1")

# Class booking rows inside each style-guide band (row 10 Mon omitted — no labels).
_CLASS_DATA_ROWS_BY_DAY: dict[int, tuple[int, ...]] = {
    0: (11, 12, 13, 14),
    1: (19, 20, 21, 22, 23),
    2: (28, 29, 30, 31, 32),
    3: (37, 38, 39, 40, 41),
    4: (46, 47, 48, 49, 50),
}

# Day-name rows above each band (style guide); not in _ALLOWED_CLASS_CELLS.
_DAY_HEADER_ROW_BY_DAY: dict[int, int] = {
    day: (rows[0] - 5 if day == 0 else rows[0] - 3)
    for day, rows in _CLASS_DATA_ROWS_BY_DAY.items()
}
_TERM1_DAY_HEADER_COL = 2
_TERM2_DAY_HEADER_COL = 16
_CHANGE_HIGHLIGHT_FILL = PatternFill(
    start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
)

TERM1_LABEL_COLS = (2, 3, 4)
TERM1_WEEK_COLS = tuple(range(5, 15))
TERM2_LABEL_COLS = (16, 17, 18)
TERM2_WEEK_COLS = tuple(range(19, 29))
_RESERVED_COLS = frozenset({15, 29, 30, 31})

_COURSE_TITLE_ROW = 3
_COURSE_TITLE_COL = 2

# Ranges the style guide must contain (used to reject the old OneDrive template).
_TEMPLATE_PROBE_CELL = "N17"
_TEMPLATE_PROBE_VALUE = "Week 10"
_TEMPLATE_MIN_B_WIDTH = 15.0


def _resolve_admin_template_sheet(sheetnames: list[str]) -> str:
    for name in ADMIN_TEMPLATE_SHEET_ALIASES:
        if name in sheetnames:
            return name
    if len(sheetnames) == 1:
        return sheetnames[0]
    raise ValueError(
        f"Template missing admin course sheet (expected one of "
        f"{', '.join(ADMIN_TEMPLATE_SHEET_ALIASES)}); "
        f"found: {', '.join(sheetnames)}"
    )


def _template_is_style_guide(path: Path) -> bool:
    """Reject legacy templates (narrow columns, missing week headers)."""
    wb = load_workbook(path, data_only=True)
    try:
        sheet = _resolve_admin_template_sheet(list(wb.sheetnames))
        ws = wb[sheet]
        probe = ws[_TEMPLATE_PROBE_CELL].value
        if probe is None or str(probe).strip() != _TEMPLATE_PROBE_VALUE:
            return False
        width = ws.column_dimensions["B"].width
        if width is None or width < _TEMPLATE_MIN_B_WIDTH:
            return False
        return True
    finally:
        wb.close()


def resolve_admin_template_path(preferred: str | Path | None = None) -> Path:
    """Pick a valid style-guide workbook; ignore broken legacy template paths."""
    candidates: list[Path] = []
    if preferred:
        candidates.append(Path(preferred))
    candidates.extend([ADMIN_STYLE_GUIDE_PATH, ADMIN_TEMPLATE_PATH])
    seen: set[Path] = set()
    for path in candidates:
        key = path.resolve()
        if key in seen or not path.is_file():
            continue
        seen.add(key)
        if _template_is_style_guide(path):
            return path
    raise FileNotFoundError(
        "Admin export style guide not found or invalid. Expected "
        f"{ADMIN_STYLE_GUIDE_PATH} with {_TEMPLATE_PROBE_CELL}="
        f"{_TEMPLATE_PROBE_VALUE!r} and column B width >= {_TEMPLATE_MIN_B_WIDTH}."
    )


@dataclass(frozen=True)
class AdminLayout:
    data_rows_by_day: dict[int, tuple[int, ...]]


@dataclass
class AdminExportReport:
    out_path: str
    course_tabs: int
    bookings_written: int
    warnings: list[str]
    template_path: str | None = None


def _sheet_title(code: str, used: set[str]) -> str:
    title = (code or "Course").strip()[:31] or "Course"
    base = title
    n = 1
    while title in used:
        n += 1
        suffix = f" ({n})"
        title = (base[: 31 - len(suffix)] + suffix).strip()
    used.add(title)
    return title


def _fmt_time(start_slot: int, end_slot: int) -> str:
    end = "22:00" if end_slot >= 28 else slot_to_time(end_slot).strftime("%H:%M")
    return f"{slot_to_time(start_slot).strftime('%H:%M')}-{end}"


def _fill_for_class(unit_text: str, course_code: str = "") -> PatternFill:
    key = (unit_text or "").strip() or course_code
    h = int(hashlib.md5(key.encode("utf-8")).hexdigest()[:6], 16)
    hue = (h % 360) / 360.0
    r, g, b = colorsys.hls_to_rgb(hue, 0.86, 0.55)
    hx = f"FF{int(r * 255):02X}{int(g * 255):02X}{int(b * 255):02X}"
    return PatternFill(start_color=hx, end_color=hx, fill_type="solid")


def _layout() -> AdminLayout:
    return AdminLayout(data_rows_by_day=dict(_CLASS_DATA_ROWS_BY_DAY))


def _parse_range_cells(range_ref: str) -> frozenset[tuple[int, int]]:
    """Parse 'B9:N14' into (row, col) pairs."""
    from openpyxl.utils.cell import range_boundaries

    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    return frozenset(
        (row, col)
        for row in range(min_row, max_row + 1)
        for col in range(min_col, max_col + 1)
    )


def _build_allowed_class_cells() -> frozenset[tuple[int, int]]:
    cells: set[tuple[int, int]] = set()
    for ref in (
        "B9:N14",
        "B19:N23",
        "B28:N32",
        "B37:N41",
        "B46:N50",
        "P9:R14",
        "P19:R23",
        "P28:R32",
        "P37:R41",
        "P46:R50",
        "S9:AB14",
        "S19:AB23",
        "S28:AB32",
        "S37:AB41",
        "S46:AB50",
    ):
        cells.update(_parse_range_cells(ref))
    return frozenset(cells)


_ALLOWED_CLASS_CELLS = _build_allowed_class_cells()


def _cell_allowed(row: int, col: int) -> bool:
    return (row, col) in _ALLOWED_CLASS_CELLS


def _is_merged_non_anchor(ws, row: int, col: int) -> bool:
    return isinstance(ws.cell(row=row, column=col), MergedCell)


def _unmerge_overlapping_on_row(ws, row: int, min_col: int, max_col: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= row <= merged.max_row and merged.min_col <= max_col and merged.max_col >= min_col:
            if merged.max_row > merged.min_row or merged.max_col > merged.min_col:
                ws.unmerge_cells(str(merged))


def _set_cell_value(ws, row: int, col: int, value) -> None:
    if row == _COURSE_TITLE_ROW and col == _COURSE_TITLE_COL:
        ws.cell(row=row, column=col).value = value
        return
    if not _cell_allowed(row, col) or _is_merged_non_anchor(ws, row, col):
        return
    ws.cell(row=row, column=col).value = value


def _admin_class_title_line(b: Booking) -> str:
    """Class title for admin week cells (event id is appended separately)."""
    name = (b.unit.name if b.unit else "") or ""
    part = getattr(b, "session_part", 1) or 1
    if part > 1:
        name = f"{name} ({part}/2)".strip()
    elif b.unit and getattr(b.unit, "double_session", 0):
        name = f"{name} (1/2)".strip()
    return name


def _admin_week_cell_text(b: Booking) -> str:
    title = _admin_class_title_line(b)
    codes = _component_codes_line(b)
    if codes:
        text = f"{title} ({codes})" if title else codes
    else:
        text = title
    eid = (b.external_id or "").strip()
    if eid:
        text = f"{text} ID: {eid}".strip()
    return text


def _writable_week_cols(ws, row: int, week_cols: tuple[int, ...]) -> list[int]:
    return [
        col
        for col in week_cols
        if _cell_allowed(row, col) and not _is_merged_non_anchor(ws, row, col)
    ]


def _row_accepts_booking(
    ws,
    row: int,
    in_term_1: bool,
    in_term_2: bool,
) -> bool:
    if in_term_1 and not _writable_week_cols(ws, row, TERM1_WEEK_COLS):
        return False
    if in_term_2 and not _writable_week_cols(ws, row, TERM2_WEEK_COLS):
        return False
    return True


def _semester_week_col(week_num: int) -> int | None:
    if 1 <= week_num <= 10:
        return TERM1_WEEK_COLS[week_num - 1]
    if 11 <= week_num <= 20:
        return TERM2_WEEK_COLS[week_num - 11]
    return None


def _contiguous_col_ranges(cols: list[int]) -> list[tuple[int, int]]:
    if not cols:
        return []
    ordered = sorted(cols)
    ranges: list[tuple[int, int]] = []
    start = prev = ordered[0]
    for col in ordered[1:]:
        if col == prev + 1:
            prev = col
            continue
        ranges.append((start, prev))
        start = prev = col
    ranges.append((start, prev))
    return ranges


def _write_week_band_row(
    ws,
    row: int,
    week_cols: tuple[int, ...],
    unit_text: str,
    fill: PatternFill,
) -> None:
    merge_cols = _writable_week_cols(ws, row, week_cols)
    if not merge_cols:
        return

    _unmerge_overlapping_on_row(ws, row, merge_cols[0], merge_cols[-1])

    for col in merge_cols:
        cell = ws.cell(row=row, column=col)
        cell.value = None
        cell.fill = fill

    if len(merge_cols) > 1:
        ws.merge_cells(
            start_row=row,
            end_row=row,
            start_column=merge_cols[0],
            end_column=merge_cols[-1],
        )

    anchor = ws.cell(row=row, column=merge_cols[0])
    if isinstance(anchor, MergedCell):
        return
    ut = (unit_text or "").strip()
    anchor.value = ut if ut else None
    anchor.fill = fill


def _write_active_week_band_row(
    ws,
    row: int,
    term: int,
    booking: Booking,
    unit_text: str,
    fill: PatternFill,
) -> None:
    """Fill only columns for active session weeks in this term band."""
    active = active_session_weeks_in_term(booking, term)
    if not active:
        return
    cols = sorted(
        col
        for week_num in active
        if (col := _semester_week_col(week_num)) is not None
    )
    cols = [
        col
        for col in cols
        if _cell_allowed(row, col) and not _is_merged_non_anchor(ws, row, col)
    ]
    if not cols:
        return

    for start_col, end_col in _contiguous_col_ranges(cols):
        _write_week_band_row(ws, row, tuple(range(start_col, end_col + 1)), unit_text, fill)


def _write_label_triplet(
    ws,
    row: int,
    cols: tuple[int, ...],
    t: str,
    staff: str,
    room: str,
) -> None:
    for col, text in zip(cols, (t, staff, room)):
        _set_cell_value(ws, row, col, text)


def _set_label_cell_fill(ws, row: int, col: int, fill: PatternFill) -> None:
    if not _cell_allowed(row, col) or _is_merged_non_anchor(ws, row, col):
        return
    ws.cell(row=row, column=col).fill = fill


def _set_day_header_fill(ws, row: int, col: int, fill: PatternFill) -> None:
    if _is_merged_non_anchor(ws, row, col):
        return
    ws.cell(row=row, column=col).fill = fill


def _apply_row_change_highlights(
    ws,
    row: int,
    flags,
    *,
    in_term_1: bool,
    in_term_2: bool,
) -> None:
    """Red-fill TIME / Lecturer / Room cells that changed (per resolved change log)."""
    if in_term_1:
        cols = TERM1_LABEL_COLS
        if flags.time:
            _set_label_cell_fill(ws, row, cols[0], _CHANGE_HIGHLIGHT_FILL)
        if flags.lecturer:
            _set_label_cell_fill(ws, row, cols[1], _CHANGE_HIGHLIGHT_FILL)
        if flags.room:
            _set_label_cell_fill(ws, row, cols[2], _CHANGE_HIGHLIGHT_FILL)
    if in_term_2:
        cols = TERM2_LABEL_COLS
        if flags.time:
            _set_label_cell_fill(ws, row, cols[0], _CHANGE_HIGHLIGHT_FILL)
        if flags.lecturer:
            _set_label_cell_fill(ws, row, cols[1], _CHANGE_HIGHLIGHT_FILL)
        if flags.room:
            _set_label_cell_fill(ws, row, cols[2], _CHANGE_HIGHLIGHT_FILL)


def _apply_day_header_highlights(ws, days: set[int]) -> None:
    for day in days:
        hdr_row = _DAY_HEADER_ROW_BY_DAY.get(day)
        if hdr_row is None:
            continue
        _set_day_header_fill(ws, hdr_row, _TERM1_DAY_HEADER_COL, _CHANGE_HIGHLIGHT_FILL)
        _set_day_header_fill(ws, hdr_row, _TERM2_DAY_HEADER_COL, _CHANGE_HIGHLIGHT_FILL)


def _copy_column_dimensions(src, dst) -> None:
    """openpyxl copy_worksheet sometimes drops widths; mirror explicitly."""
    for key, dim in src.column_dimensions.items():
        if dim.width is not None:
            dst.column_dimensions[key].width = dim.width


def _admin_lecturer_cell(b: Booking, *, term: Literal["t1", "t2"] | None = None) -> str:
    from ..core.booking_staff import export_lecturer_label

    return export_lecturer_label(b, term=term)


def _write_course_admin_sheet(
    session: Session,
    ws,
    layout: AdminLayout,
    week: Week,
    course: Course,
    *,
    co_teach_only: bool = False,
    change_highlights: dict | None = None,
) -> tuple[int, list[str]]:
    """Populate one admin template sheet for ``course``; return (rows written, warnings)."""
    row_slot: dict[int, int] = {d: 0 for d in layout.data_rows_by_day}
    cc = (course.code or "") or ""
    title = f"Course: {course.code or ''}"
    if co_teach_only:
        title = f"{title} (SFS co-teach)"
    _set_cell_value(ws, _COURSE_TITLE_ROW, _COURSE_TITLE_COL, title)

    q = (
        session.query(Booking)
        .options(
            joinedload(Booking.staff),
            joinedload(Booking.sfs_co_teacher),
            joinedload(Booking.unit),
            joinedload(Booking.room),
        )
        .filter(Booking.week_id == week.id, Booking.course_id == course.id)
        .order_by(Booking.day, Booking.start_slot)
    )
    if co_teach_only:
        q = q.filter(sfs_co_teacher_booking_filter())
    bookings_written = 0
    warnings: list[str] = []
    day_headers_to_mark: set[int] = set()
    highlights = change_highlights or {}
    for b in q.all():
        rows = layout.data_rows_by_day.get(b.day)
        if not rows:
            continue
        in_term_1 = int(getattr(b, "in_term_1", 1))
        in_term_2 = int(getattr(b, "in_term_2", 1))
        idx = row_slot[b.day]
        while idx < len(rows) and not _row_accepts_booking(ws, rows[idx], in_term_1, in_term_2):
            idx += 1
        if idx >= len(rows):
            warnings.append(
                f"{course.code}: too many rows on {DAYS[b.day]}, skipped booking {b.id}"
            )
            continue
        cur = rows[idx]
        unit_text = _admin_week_cell_text(b)
        t = _fmt_time(b.start_slot, b.end_slot)
        room = b.room.code if b.room else ""
        fill = _fill_for_class(cc, unit_text)

        if in_term_1:
            _write_label_triplet(
                ws, cur, TERM1_LABEL_COLS, t, _admin_lecturer_cell(b, term="t1"), room
            )
            _write_active_week_band_row(ws, cur, 1, b, unit_text, fill)
        if in_term_2:
            _write_label_triplet(
                ws, cur, TERM2_LABEL_COLS, t, _admin_lecturer_cell(b, term="t2"), room
            )
            _write_active_week_band_row(ws, cur, 2, b, unit_text, fill)

        eid = (b.external_id or "").strip()
        flags = highlights.get(eid) if eid else None
        if flags is not None:
            _apply_row_change_highlights(
                ws, cur, flags, in_term_1=bool(in_term_1), in_term_2=bool(in_term_2)
            )
            day_headers_to_mark.update(flags.day_header_days)

        row_slot[b.day] = idx + 1
        bookings_written += 1
    if day_headers_to_mark:
        _apply_day_header_highlights(ws, day_headers_to_mark)
    return bookings_written, warnings


def _courses_for_admin_export(
    session: Session,
    week: Week,
    *,
    co_teach_only: bool,
    timetable_session_id: int | None = None,
) -> list[Course]:
    if not co_teach_only:
        q = session.query(Course).order_by(Course.code)
        if timetable_session_id is not None:
            q = q.filter(Course.timetable_session_id == timetable_session_id)
        return q.all()
    ids = {
        row[0]
        for row in session.query(Booking.course_id)
        .filter(Booking.week_id == week.id, sfs_co_teacher_booking_filter())
        .distinct()
        .all()
    }
    if not ids:
        return []
    return (
        session.query(Course)
        .filter(Course.id.in_(ids))
        .order_by(Course.code)
        .all()
    )


def _write_admin_workbook(
    session: Session,
    out_path: Path,
    template_path: Path,
    *,
    co_teach_only: bool = False,
    week_id: int | None = None,
    timetable_session_id: int | None = None,
) -> AdminExportReport:
    layout = _layout()
    probe = load_workbook(template_path, read_only=True)
    try:
        template_sheet = _resolve_admin_template_sheet(list(probe.sheetnames))
    finally:
        probe.close()

    if template_path.resolve() != out_path.resolve():
        shutil.copyfile(template_path, out_path)

    wb = load_workbook(out_path)
    base_ws = wb[template_sheet]
    if week_id is not None:
        week = session.get(Week, week_id)
        if week is None:
            raise RuntimeError(f"Week id {week_id} not found.")
    else:
        week = session.query(Week).order_by(Week.id).first()
    if week is None:
        raise RuntimeError("No week exists in session.")
    # Full highlight logic lives in desktop UI; web export uses no highlights for now.
    change_highlights: dict = {}
    courses = _courses_for_admin_export(
        session,
        week,
        co_teach_only=co_teach_only,
        timetable_session_id=timetable_session_id,
    )
    used_titles: set[str] = set()
    warnings: list[str] = []
    bookings_written = 0
    new_sheets = []
    for c in courses:
        ws = wb.copy_worksheet(base_ws)
        _copy_column_dimensions(base_ws, ws)
        ws.title = _sheet_title(c.code or f"Course {c.id}", used_titles)
        n, w = _write_course_admin_sheet(
            session,
            ws,
            layout,
            week,
            c,
            co_teach_only=co_teach_only,
            change_highlights=change_highlights,
        )
        if co_teach_only and n == 0:
            wb.remove(ws)
            continue
        bookings_written += n
        warnings.extend(w)
        new_sheets.append(ws)

    wb.remove(wb[template_sheet])
    if not new_sheets:
        label = "No SFS co-teach classes" if co_teach_only else "No courses"
        ws = wb.create_sheet(label[:31])
        ws["A1"] = f"{label} in session."
    write_backup_sheet(wb, session, timetable_session_id=timetable_session_id)
    wb.save(out_path)
    return AdminExportReport(
        out_path=str(out_path),
        course_tabs=len(new_sheets),
        bookings_written=bookings_written,
        warnings=warnings,
        template_path=str(template_path),
    )


def write_admin_export(
    session: Session,
    out_path: str | Path,
    template_path: str | Path | None = None,
    *,
    week_id: int | None = None,
    timetable_session_id: int | None = None,
) -> AdminExportReport:
    """Export one sheet per course; template supplies all formatting."""
    template_path = resolve_admin_template_path(template_path)
    return _write_admin_workbook(
        session,
        Path(out_path),
        template_path,
        co_teach_only=False,
        week_id=week_id,
        timetable_session_id=timetable_session_id,
    )


def write_co_teach_admin_export(
    session: Session,
    out_path: str | Path,
    template_path: str | Path | None = None,
    *,
    week_id: int | None = None,
    timetable_session_id: int | None = None,
) -> AdminExportReport:
    """Export one sheet per course that has SFS co-teach classes (admin template)."""
    template_path = resolve_admin_template_path(template_path)
    return _write_admin_workbook(
        session,
        Path(out_path),
        template_path,
        co_teach_only=True,
        week_id=week_id,
        timetable_session_id=timetable_session_id,
    )
