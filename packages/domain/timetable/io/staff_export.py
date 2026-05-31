"""Write Staff tab (hours summary) to Excel."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..core.staff_export_data import STAFF_TAB_EXPORT_HEADERS, gather_staff_tab_main_rows


def _write_sheet(
    ws,
    headers: tuple[str, ...],
    rows: list[dict[str, str]],
    *,
    header_fill: str,
    col_widths: tuple[int, ...],
) -> None:
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill_obj = PatternFill(start_color=header_fill, end_color=header_fill, fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = header_font
        c.fill = header_fill_obj
        c.alignment = header_align
    body_align = Alignment(vertical="top", wrap_text=True)
    for r, row in enumerate(rows, start=2):
        for col, key in enumerate(headers, start=1):
            ws.cell(row=r, column=col, value=row.get(key, "")).alignment = body_align
    for i, w in enumerate(col_widths, start=1):
        if i <= len(headers):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def write_staff_tab_xlsx(
    session: Session,
    path: str | Path,
    *,
    workbook_title: str = "Staff",
) -> str:
    """Single sheet: main Staff hours table."""
    path = Path(path)
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Staff hours"[:31]
    main_rows = gather_staff_tab_main_rows(session)
    _write_sheet(
        ws0,
        STAFF_TAB_EXPORT_HEADERS,
        main_rows,
        header_fill="FF374151",
        col_widths=(18, 8, 14, 18, 10, 48, 14, 14, 28, 12, 12, 10),
    )

    wb.properties.title = workbook_title[:31] or "Staff"
    wb.save(path)
    return str(path)
