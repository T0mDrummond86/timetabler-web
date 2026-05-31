"""Write warnings / clashes report to Excel."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .export_headers import VIOLATIONS_EXPORT_HEADERS


def write_violations_report_xlsx(
    path: str | Path,
    rows: list[dict[str, str]],
    *,
    sheet_title: str = "Warnings",
) -> str:
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Warnings"

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(start_color="FF92400e", end_color="FF92400e", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col, title in enumerate(VIOLATIONS_EXPORT_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    body_align = Alignment(vertical="top", wrap_text=True)
    for r, row in enumerate(rows, start=2):
        for col, key in enumerate(VIOLATIONS_EXPORT_HEADERS, start=1):
            ws.cell(row=r, column=col, value=row.get(key, "")).alignment = body_align

    widths = (10, 22, 14, 22, 36, 20, 12, 16, 10, 50)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(path)
    return str(path)
