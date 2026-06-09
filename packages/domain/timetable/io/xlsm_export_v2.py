"""Export v2 — per-entity timetables using pre-styled v2 Excel templates (no Overall sheet).

Grid styling (borders, fills, headers, time column) lives in ``templates/export_v2_base.xlsm``
and must be edited in Excel. This module only writes data: sheet titles and merged placecards.

Layout and grid borders come from ``timetable_template_styleguide.xlsx``. Placecards restore
per-cell template borders before applying a coloured edge; empty grid cells are reset from the
template after painting. The staff summary table clones row styles from the styleguide as it grows.

Each weekday block is four columns wide. Course placecards use two columns (T1 left, T2 right)
or all four when the booking is in both terms.
"""
from __future__ import annotations

import shutil
from copy import copy
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from ..bundle_paths import resource_path
from ..constants import NUM_DAYS, NUM_SLOTS
from ..core.class_colour import booking_colour_key
from ..core.models import Booking, Course, Room, Staff
from ..core.room_types import room_is_physical
from .backup_payload import write_backup_sheet
from ..core.screen_colours import assign_screen_colours, screen_border_xlsx, screen_fill_xlsx
from .xlsm_export import (
    ExportReport,
    _bookings_by_slot,
    _bookings_for_staff_tab,
    _copy_worksheet_layout,
    _entity_values,
    _staff_lane_subgroups,
    _staff_placecard_lines,
    _lecturer_label,
    placecard_subject_block,
    _resolve_week,
    _safe_sheet_name,
    _terms_of,
    _week_bookings,
)

V2_TEMPLATE_PATH = resource_path("templates", "export_v2_base.xlsm")

V2_TEMPLATE_SHEETS = (
    "Course Template v2",
    "Staff Template v2",
    "Room Template v2",
)

V2_TITLE_ROW = 1
V2_FIRST_SLOT_ROW = 3
V2_DAY_STRIDE = 4  # 4 contiguous day columns (Mon–Fri), no spacer
V2_NUM_DAYS = NUM_DAYS  # Mon–Fri (no Saturday)
V2_DAY_FIRST_COL = 2
V2_DAY_FIRST_COLS = tuple(V2_DAY_FIRST_COL + day * V2_DAY_STRIDE for day in range(V2_NUM_DAYS))
V2_DAY_BLOCK_WIDTH = 4
V2_LANE_WIDTH = 2  # T1-only or T2-only placecards span two columns within the day block
V2_LAST_COL = V2_DAY_FIRST_COLS[-1] + V2_DAY_BLOCK_WIDTH - 1
V2_GRID_LAST_ROW = V2_FIRST_SLOT_ROW + NUM_SLOTS - 1
V2_FOOTER_START_ROW = V2_GRID_LAST_ROW + 1

# Staff sheet: hours summary beside the grid — title row 2 (cols W/X), table from row 3.
V2_SUMMARY_TITLE_ROW = 2
V2_SUMMARY_CLASS_COL = 23
V2_SUMMARY_HOURS_COL = 24
V2_STAFF_HOURS_SUMMARY_HEADER_ROW = 3
V2_SUMMARY_HEADER_ROW = V2_STAFF_HOURS_SUMMARY_HEADER_ROW
V2_SUMMARY_FIRST_DATA_ROW = V2_STAFF_HOURS_SUMMARY_HEADER_ROW + 1
# Template rows whose cell styles are cloned when the summary table grows.
V2_SUMMARY_DATA_STYLE_ROW = 4
V2_SUMMARY_TOTAL_STYLE_ROW = 9

# Placecard styling only (booking cards, not the grid template).
V2_PLACECARD_FONT = Font(name="Calibri", size=12, bold=True, color="FF1A1A1A")
V2_PLACECARD_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
V2_SUMMARY_HEADER_FILL = PatternFill(
    start_color="FF374151", end_color="FF374151", fill_type="solid"
)
V2_SUMMARY_DATA_FILL = PatternFill(
    start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid"
)
V2_SUMMARY_DATA_ALT_FILL = PatternFill(
    start_color="FFF9FAFB", end_color="FFF9FAFB", fill_type="solid"
)
V2_SUMMARY_TOTAL_FILL = PatternFill(
    start_color="FFF3F4F6", end_color="FFF3F4F6", fill_type="solid"
)
V2_SUMMARY_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
V2_SUMMARY_BODY_FONT = Font(name="Calibri", size=11, bold=False, color="FF1A1A1A")
V2_SUMMARY_TOTAL_FONT = Font(name="Calibri", size=11, bold=True, color="FF1A1A1A")
V2_SUMMARY_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
V2_SUMMARY_HOURS_ALIGN = Alignment(horizontal="right", vertical="center")
V2_SUMMARY_CLASS_COL_WIDTH = 32
V2_SUMMARY_HOURS_COL_WIDTH = 11

_SUMMARY_BORDER_COLOR = "FFD1D5DB"
_SUMMARY_BORDER_STRONG = "FF9CA3AF"


def _ranges_intersect(a_min: int, a_max: int, b_min: int, b_max: int) -> bool:
    return a_min <= b_max and a_max >= b_min


@dataclass
class ExportV2Report(ExportReport):
    """Same counters as v1 export; no Overall courses_written fields used."""

    template_path: Path | None = None
    colour_by_class: bool = True


def _v2_slot_row(slot: int) -> int:
    return V2_FIRST_SLOT_ROW + slot


def _v2_booking_rows(b: Booking) -> tuple[int, int]:
    """Map booking slots to worksheet rows (end_slot is exclusive)."""
    top_row = _v2_slot_row(b.start_slot)
    end_slot = b.end_slot if b.end_slot > b.start_slot else b.start_slot + 1
    bot_row = _v2_slot_row(end_slot - 1)
    return top_row, max(top_row, bot_row)


def _v2_lane_cols(day: int, t1: bool, t2: bool) -> tuple[int, int]:
    """Inclusive column range for a placecard within a 4-column day block.

    - Semester (T1 and T2): all four columns.
    - T1 only: left two columns of the block.
    - T2 only: right two columns of the block.
    """
    if not (0 <= day < V2_NUM_DAYS):
        start = V2_DAY_FIRST_COLS[0]
        return start, start + V2_DAY_BLOCK_WIDTH - 1
    start = V2_DAY_FIRST_COLS[day]
    if t1 and t2:
        return start, start + V2_DAY_BLOCK_WIDTH - 1
    if t1:
        return start, start + V2_LANE_WIDTH - 1
    if t2:
        return start + V2_LANE_WIDTH, start + V2_DAY_BLOCK_WIDTH - 1
    return start, start + V2_DAY_BLOCK_WIDTH - 1


def _v2_last_slot_row() -> int:
    return V2_FIRST_SLOT_ROW + NUM_SLOTS - 1


def _v2_body_col_range() -> tuple[int, int]:
    first = V2_DAY_FIRST_COLS[0]
    last_day = V2_DAY_FIRST_COLS[-1]
    return first, last_day + V2_DAY_BLOCK_WIDTH - 1


def _v2_slot_content_cols() -> tuple[int, ...]:
    """Day-block content columns only (excludes gap/spacer columns)."""
    cols: list[int] = []
    for day in range(V2_NUM_DAYS):
        start = V2_DAY_FIRST_COLS[day]
        cols.extend(range(start, start + V2_DAY_BLOCK_WIDTH))
    return tuple(cols)


def _unmerge_intersecting(
    ws,
    top_row: int,
    bot_row: int,
    col_start: int,
    col_end: int,
) -> None:
    """Split merges overlapping the target rectangle."""
    for merged in list(ws.merged_cells.ranges):
        if _ranges_intersect(merged.min_row, merged.max_row, top_row, bot_row) and _ranges_intersect(
            merged.min_col, merged.max_col, col_start, col_end
        ):
            ws.unmerge_cells(str(merged))


def _clear_v2_cell_value(ws, row: int, col: int) -> None:
    """Clear cell value only; preserve template formatting."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = None


def _copy_cell_style(ref_cell, cell) -> None:
    """Copy one cell's display style (openpyxl does not preserve these when values change)."""
    if not ref_cell.has_style:
        return
    cell.font = copy(ref_cell.font)
    cell.border = copy(ref_cell.border)
    cell.fill = copy(ref_cell.fill)
    cell.number_format = copy(ref_cell.number_format)
    cell.protection = copy(ref_cell.protection)
    cell.alignment = copy(ref_cell.alignment)


def _set_cell_border(
    cell,
    *,
    top: Side | None = None,
    bottom: Side | None = None,
    left: Side | None = None,
    right: Side | None = None,
) -> None:
    """Update selected border sides without clearing the rest."""
    b = cell.border
    cell.border = Border(
        top=top if top is not None else b.top,
        bottom=bottom if bottom is not None else b.bottom,
        left=left if left is not None else b.left,
        right=right if right is not None else b.right,
        diagonal=b.diagonal,
        diagonal_direction=b.diagonal_direction,
        outline=b.outline,
        vertical=b.vertical,
        horizontal=b.horizontal,
    )


def _placecard_anchor(ws, row: int, col: int) -> bool:
    """True when this cell is the anchor of a painted booking placecard."""
    cell = ws.cell(row=row, column=col)
    if cell.value in (None, ""):
        return False
    if cell.fill.fill_type != "solid":
        return False
    rgb = str(getattr(cell.fill.start_color, "rgb", None) or "").upper()
    if rgb in ("", "00000000", "FFFFFFFF"):
        return False
    if rgb.endswith("F3E8FF") or "F3E8FF" in rgb:
        return False
    for merged in ws.merged_cells.ranges:
        if merged.min_row == row and merged.min_col == col:
            return merged.max_row > merged.min_row or merged.max_col > merged.min_col
    return False


def _reapply_template_grid_borders(ws, ref_ws) -> None:
    """Restore timetable grid borders from the template except on placecard anchors."""
    cols = (1,) + _v2_slot_content_cols()
    for row in range(V2_FIRST_SLOT_ROW, _v2_last_slot_row() + 1):
        for col in cols:
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            if _placecard_anchor(ws, row, col):
                continue
            cell.border = copy(ref_ws.cell(row=row, column=col).border)


def _clone_v2_template_sheet(wb, template_ws, title: str, used_names: set[str]):
    """Copy a v2 template tab with full grid styling (copy_worksheet can drop layout)."""
    ws = wb.create_sheet(_safe_sheet_name(title, used_names))
    _copy_worksheet_layout(template_ws, ws)
    return ws


def _clear_v2_body(ws) -> None:
    """Clear prior booking text from day columns; do not alter grid styles."""
    for row in range(V2_FIRST_SLOT_ROW, _v2_last_slot_row() + 1):
        for col in _v2_slot_content_cols():
            _clear_v2_cell_value(ws, row, col)


def _clear_v2_footer(ws) -> None:
    """Clear footer text only (values), preserving template formatting."""
    last_col = _v2_body_col_range()[1]
    for row in range(V2_FOOTER_START_ROW, ws.max_row + 1):
        for col in range(1, last_col + 1):
            _clear_v2_cell_value(ws, row, col)


def _unmerge_v2_summary_cols(ws, *, keep_title_merge: bool) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_col >= V2_SUMMARY_CLASS_COL and merged.max_col <= V2_SUMMARY_HOURS_COL:
            if (
                keep_title_merge
                and merged.min_row <= V2_SUMMARY_TITLE_ROW <= merged.max_row
            ):
                continue
            ws.unmerge_cells(str(merged))


def _unmerge_v2_staff_summary(ws) -> None:
    _unmerge_v2_summary_cols(ws, keep_title_merge=True)


def _copy_summary_row_style(
    ws,
    ref_ws,
    template_row: int,
    target_row: int,
    *,
    clear_values: bool = True,
) -> None:
    """Clone summary table formatting from the styleguide row."""
    for col in (V2_SUMMARY_CLASS_COL, V2_SUMMARY_HOURS_COL):
        dst = ws.cell(row=target_row, column=col)
        if clear_values:
            dst.value = None
        _copy_cell_style(ref_ws.cell(template_row, col), dst)


def _clear_v2_staff_summary(ws, ref_ws) -> None:
    """Clear summary body values and reset row styles from the template band."""
    _unmerge_v2_staff_summary(ws)
    last_row = max(ws.max_row, V2_SUMMARY_FIRST_DATA_ROW + 40)
    for row in range(V2_SUMMARY_FIRST_DATA_ROW, last_row + 1):
        style_row = (
            V2_SUMMARY_TOTAL_STYLE_ROW
            if row == V2_SUMMARY_TOTAL_STYLE_ROW
            else V2_SUMMARY_DATA_STYLE_ROW
        )
        _copy_summary_row_style(ws, ref_ws, style_row, row)


def _summary_cols() -> tuple[int, int]:
    return V2_SUMMARY_CLASS_COL, V2_SUMMARY_HOURS_COL


def _blank_v2_summary_cell(ws, row: int, col: int) -> None:
    """Remove summary values, borders, and template fills from one cell."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = None
    cell.border = Border()
    cell.fill = PatternFill()
    cell.font = Font(name="Calibri", size=11)


def _clear_v2_hours_summary_panel(ws) -> None:
    """Remove the hours summary table (course/room tabs must not show lecturer summaries)."""
    _unmerge_v2_summary_cols(ws, keep_title_merge=False)
    last_row = max(ws.max_row, V2_SUMMARY_FIRST_DATA_ROW + 40)
    for row in range(V2_SUMMARY_TITLE_ROW, last_row + 1):
        for col in (V2_SUMMARY_CLASS_COL, V2_SUMMARY_HOURS_COL):
            _blank_v2_summary_cell(ws, row, col)


def _trim_v2_summary_below(ws, last_row: int) -> None:
    """Clear template summary borders/values below the last written row."""
    end_row = max(ws.max_row, V2_SUMMARY_FIRST_DATA_ROW + 40)
    for row in range(last_row + 1, end_row + 1):
        for col in (V2_SUMMARY_CLASS_COL, V2_SUMMARY_HOURS_COL):
            _blank_v2_summary_cell(ws, row, col)


def _colour_map_for_bookings(
    bookings: list[Booking],
    *,
    colour_by_class: bool,
) -> dict[str, tuple[str, str]]:
    keys = {booking_colour_key(b, by_class=colour_by_class) for b in bookings}
    return assign_screen_colours(keys)


def _v2_summary_spacer_row(
    ws, ref_ws, row: int, col_l: int, col_r: int
) -> int:
    _copy_summary_row_style(ws, ref_ws, V2_SUMMARY_DATA_STYLE_ROW, row)
    _clear_v2_cell_value(ws, row, col_l)
    _clear_v2_cell_value(ws, row, col_r)
    return row + 1


def _v2_summary_hours_row(
    ws,
    row: int,
    col_l: int,
    col_r: int,
    label: str,
    hours: float | None,
) -> int:
    ws.cell(row=row, column=col_l, value=label)
    h_cell = ws.cell(row=row, column=col_r, value=None if hours is None else round(hours, 2))
    if hours is not None:
        h_cell.number_format = "0.00"
    return row + 1


def _render_v2_staff_hours_summary(
    ws,
    session: Session,
    staff_id: int,
    bookings: list[Booking],
    *,
    grid_ref_ws,
) -> int:
    """Class / hours table in columns W/X; grows using template row borders."""
    from ..core.staff_hours import (
        class_hours_summary_for_staff_export,
        staff_v2_hours_summary_footer,
    )

    _clear_v2_staff_summary(ws, grid_ref_ws)
    class_rows, _class_total = class_hours_summary_for_staff_export(
        session, staff_id, bookings
    )
    extra_rows, grand_total, variance = staff_v2_hours_summary_footer(
        session, staff_id, bookings
    )

    col_l, col_r = _summary_cols()
    row = V2_SUMMARY_FIRST_DATA_ROW
    for label, hours in class_rows:
        _copy_summary_row_style(ws, grid_ref_ws, V2_SUMMARY_DATA_STYLE_ROW, row, clear_values=False)
        row = _v2_summary_hours_row(ws, row, col_l, col_r, label, hours)

    if class_rows:
        row = _v2_summary_spacer_row(ws, grid_ref_ws, row, col_l, col_r)

    for label, hours in extra_rows:
        _copy_summary_row_style(ws, grid_ref_ws, V2_SUMMARY_DATA_STYLE_ROW, row, clear_values=False)
        row = _v2_summary_hours_row(ws, row, col_l, col_r, label, hours)

    row = _v2_summary_spacer_row(ws, grid_ref_ws, row, col_l, col_r)
    _copy_summary_row_style(ws, grid_ref_ws, V2_SUMMARY_TOTAL_STYLE_ROW, row, clear_values=False)
    row = _v2_summary_hours_row(ws, row, col_l, col_r, "Total", grand_total)

    if variance is not None:
        _copy_summary_row_style(ws, grid_ref_ws, V2_SUMMARY_TOTAL_STYLE_ROW, row, clear_values=False)
        row = _v2_summary_hours_row(ws, row, col_l, col_r, "Over / under", variance)

    _trim_v2_summary_below(ws, row - 1)
    return row - 1


def _draw_v2_placecard(
    ws,
    top_row: int,
    bot_row: int,
    col_start: int,
    col_end: int,
    lines: tuple[str, str, str],
    tint_key: str,
    *,
    grid_ref_ws,
    colour_map: dict[str, tuple[str, str]] | None = None,
) -> None:
    """Paint a merged placecard: fill, text, coloured edge, template grid borders preserved."""
    top_row, bot_row = min(top_row, bot_row), max(top_row, bot_row)
    col_start, col_end = min(col_start, col_end), max(col_start, col_end)
    text = "\n".join(x for x in lines if x).strip()
    if not text:
        return
    _unmerge_intersecting(ws, top_row, bot_row, col_start, col_end)

    template_borders = {
        (row, col): copy(grid_ref_ws.cell(row=row, column=col).border)
        for row in range(top_row, bot_row + 1)
        for col in range(col_start, col_end + 1)
    }
    fill_hex = screen_fill_xlsx(tint_key, colour_map)
    border_hex = screen_border_xlsx(tint_key, colour_map)
    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    edge = Side(style="thin", color=border_hex)

    for row in range(top_row, bot_row + 1):
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = template_borders[(row, col)]

    anchor = ws.cell(row=top_row, column=col_start)
    anchor.value = text
    anchor.font = V2_PLACECARD_FONT
    anchor.alignment = V2_PLACECARD_ALIGN

    for col in range(col_start, col_end + 1):
        _set_cell_border(ws.cell(row=top_row, column=col), top=edge)
        _set_cell_border(ws.cell(row=bot_row, column=col), bottom=edge)
    for row in range(top_row, bot_row + 1):
        _set_cell_border(ws.cell(row=row, column=col_start), left=edge)
        _set_cell_border(ws.cell(row=row, column=col_end), right=edge)

    ws.merge_cells(
        start_row=top_row,
        end_row=bot_row,
        start_column=col_start,
        end_column=col_end,
    )


def _set_v2_title(ws, title: str) -> None:
    """Set sheet title text in the template's merged title cell (no style changes)."""
    ws.cell(row=V2_TITLE_ROW, column=1, value=title)


def _strip_saturday_from_sheet(ws) -> None:
    """Clear spacer column values between the grid and summary; keep template layout."""
    clear_start = V2_LAST_COL + 1
    clear_end = V2_SUMMARY_CLASS_COL - 1
    if clear_start > clear_end:
        return
    for row in range(1, ws.max_row + 1):
        for col in range(clear_start, clear_end + 1):
            _clear_v2_cell_value(ws, row, col)


def _render_v2_course_tab(
    ws,
    title: str,
    bookings: list[Booking],
    *,
    grid_ref_ws,
    colour_by_class: bool,
    colour_map: dict[str, tuple[str, str]],
) -> None:
    _strip_saturday_from_sheet(ws)
    _clear_v2_hours_summary_panel(ws)
    _set_v2_title(ws, title)
    _clear_v2_body(ws)
    _clear_v2_footer(ws)

    for b in bookings:
        if not (0 <= b.day < V2_NUM_DAYS):
            continue
        top_row, bot_row = _v2_booking_rows(b)
        t1, t2 = _terms_of(b)
        if not t1 and not t2:
            continue
        col_start, col_end = _v2_lane_cols(b.day, t1, t2)
        lane_term = "t1" if t1 and not t2 else ("t2" if t2 and not t1 else None)
        lines = (
            placecard_subject_block(b, term=lane_term),
            (b.room.code if b.room else "") or "",
            _lecturer_label(b),
        )
        tint = booking_colour_key(b, by_class=colour_by_class)
        _draw_v2_placecard(
            ws,
            top_row,
            bot_row,
            col_start,
            col_end,
            lines,
            tint,
            grid_ref_ws=grid_ref_ws,
            colour_map=colour_map,
        )
    _reapply_template_grid_borders(ws, grid_ref_ws)
    _clear_v2_hours_summary_panel(ws)


def _render_v2_staff_tab(
    ws,
    title: str,
    bookings: list[Booking],
    *,
    grid_ref_ws,
    view_staff_id: int,
    session: Session,
    colour_by_class: bool,
    colour_map: dict[str, tuple[str, str]],
) -> None:
    _strip_saturday_from_sheet(ws)
    _set_v2_title(ws, title)
    _clear_v2_body(ws)
    _clear_v2_footer(ws)

    for group in _bookings_by_slot(bookings):
        b0 = group[0]
        if not (0 <= b0.day < V2_NUM_DAYS):
            continue
        top_row, bot_row = _v2_booking_rows(b0)
        for (t1, t2), lane_bookings in _staff_lane_subgroups(group, view_staff_id):
            if not t1 and not t2:
                continue
            col_start, col_end = _v2_lane_cols(b0.day, t1, t2)
            lane_term = "t1" if t1 and not t2 else ("t2" if t2 and not t1 else None)
            lines, tint = _staff_placecard_lines(
                lane_bookings,
                view_staff_id=view_staff_id,
                colour_by_class=colour_by_class,
                term=lane_term,
            )
            _draw_v2_placecard(
                ws,
                top_row,
                bot_row,
                col_start,
                col_end,
                lines,
                tint,
                grid_ref_ws=grid_ref_ws,
                colour_map=colour_map,
            )

    _reapply_template_grid_borders(ws, grid_ref_ws)
    _render_v2_staff_hours_summary(
        ws, session, view_staff_id, bookings, grid_ref_ws=grid_ref_ws
    )


def _render_v2_room_tab(
    ws,
    title: str,
    bookings: list[Booking],
    *,
    grid_ref_ws,
    colour_by_class: bool,
    colour_map: dict[str, tuple[str, str]],
) -> None:
    _strip_saturday_from_sheet(ws)
    _clear_v2_hours_summary_panel(ws)
    _set_v2_title(ws, title)
    _clear_v2_body(ws)
    _clear_v2_footer(ws)

    for b in bookings:
        if not (0 <= b.day < V2_NUM_DAYS):
            continue
        top_row, bot_row = _v2_booking_rows(b)
        col_start, col_end = _v2_lane_cols(b.day, True, True)
        lines = (
            placecard_subject_block(b),
            _lecturer_label(b),
            (b.course.code if b.course else "") or "",
        )
        tint = booking_colour_key(b, by_class=colour_by_class)
        _draw_v2_placecard(
            ws,
            top_row,
            bot_row,
            col_start,
            col_end,
            lines,
            tint,
            grid_ref_ws=grid_ref_ws,
            colour_map=colour_map,
        )
    _reapply_template_grid_borders(ws, grid_ref_ws)
    _clear_v2_hours_summary_panel(ws)


def _generate_v2_entity_tabs(
    wb,
    session: Session,
    week_id: int,
    report: ExportV2Report,
    *,
    colour_by_class: bool,
    timetable_session_id: int | None = None,
) -> None:
    from ..core.sidebar_order import ordered_courses, ordered_staff

    bookings = _week_bookings(session, week_id)
    session_colour_map = _colour_map_for_bookings(
        bookings, colour_by_class=colour_by_class
    )
    used_names = set(wb.sheetnames)

    course_t = wb["Course Template v2"] if "Course Template v2" in wb.sheetnames else None
    staff_t = wb["Staff Template v2"] if "Staff Template v2" in wb.sheetnames else None
    room_t = wb["Room Template v2"] if "Room Template v2" in wb.sheetnames else None

    if course_t is not None:
        for c in ordered_courses(
            session, include_block_cohorts=True, timetable_session_id=timetable_session_id
        ):
            ws = _clone_v2_template_sheet(wb, course_t, c.code, used_names)
            _render_v2_course_tab(
                ws,
                c.code,
                [b for b in bookings if b.course_id == c.id],
                grid_ref_ws=course_t,
                colour_by_class=colour_by_class,
                colour_map=session_colour_map,
            )
            report.course_tabs += 1

    if staff_t is not None:
        for s in ordered_staff(session, timetable_session_id=timetable_session_id):
            ws = _clone_v2_template_sheet(wb, staff_t, s.name, used_names)
            staff_bookings = _bookings_for_staff_tab(bookings, s.id)
            _render_v2_staff_tab(
                ws,
                s.name,
                staff_bookings,
                grid_ref_ws=staff_t,
                view_staff_id=s.id,
                session=session,
                colour_by_class=colour_by_class,
                colour_map=session_colour_map,
            )
            report.staff_tabs += 1

    if room_t is not None:
        room_q = session.query(Room).order_by(Room.code)
        if timetable_session_id is not None and "timetable_session_id" in Room.__table__.columns:
            room_q = room_q.filter(Room.timetable_session_id == timetable_session_id)
        rooms = [r for r in room_q.all() if room_is_physical(r)]
        for r in rooms:
            ws = _clone_v2_template_sheet(wb, room_t, r.code, used_names)
            _render_v2_room_tab(
                ws,
                r.code,
                [b for b in bookings if b.room_id == r.id],
                grid_ref_ws=room_t,
                colour_by_class=colour_by_class,
                colour_map=session_colour_map,
            )
            report.room_tabs += 1

    for tpl_name in V2_TEMPLATE_SHEETS:
        if tpl_name in wb.sheetnames:
            del wb[tpl_name]


def _apply_v2_readonly_protection(wb) -> None:
    """Protect all worksheets so exported timetables open read-only in Excel."""
    from openpyxl.workbook.protection import WorkbookProtection

    for ws in wb.worksheets:
        prot = ws.protection
        prot.enable()
        # Allow viewing/selecting cells; locked cells cannot be edited once protected.
        prot.selectLockedCells = True
        prot.selectUnlockedCells = True
    wb.security = WorkbookProtection(lockStructure=True)


def write_v2(
    session: Session,
    out_path: str | Path,
    *,
    template_path: str | Path | None = None,
    week_id: int | None = None,
    colour_by_class: bool | None = None,
    timetable_session_id: int | None = None,
) -> ExportV2Report:
    """Export course / staff / room tabs using the v2 templates (no Overall)."""
    from ..core.display_settings import resolve_export_colour_by_class

    colour_by_class = resolve_export_colour_by_class(colour_by_class)
    template_path = Path(template_path or V2_TEMPLATE_PATH)
    out_path = Path(out_path)
    if not template_path.is_file():
        raise FileNotFoundError(f"v2 export template not found: {template_path}")

    probe = load_workbook(template_path, read_only=True)
    try:
        missing = [n for n in V2_TEMPLATE_SHEETS if n not in probe.sheetnames]
        if missing:
            raise ValueError(f"Template missing sheet(s): {', '.join(missing)}")
    finally:
        probe.close()

    if template_path.resolve() != out_path.resolve():
        shutil.copyfile(template_path, out_path)

    wb = load_workbook(out_path, keep_vba=True)
    report = ExportV2Report(
        out_path=out_path,
        template_path=template_path,
        colour_by_class=colour_by_class,
    )

    for name in list(wb.sheetnames):
        if name not in V2_TEMPLATE_SHEETS:
            del wb[name]
            report.sheets_removed.append(name)

    week = _resolve_week(session, week_id)

    for tpl_name in V2_TEMPLATE_SHEETS:
        _strip_saturday_from_sheet(wb[tpl_name])

    _generate_v2_entity_tabs(
        wb,
        session,
        week.id,
        report,
        colour_by_class=colour_by_class,
        timetable_session_id=timetable_session_id,
    )
    report.bookings = len(_week_bookings(session, week.id))
    write_backup_sheet(wb, session, timetable_session_id=timetable_session_id)
    _apply_v2_readonly_protection(wb)
    wb.save(out_path)
    return report
