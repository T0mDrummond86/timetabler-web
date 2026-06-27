"""Serialize the entire session to a JSON-safe dict and back.

This payload is embedded in a hidden Excel sheet on export (timetable workbook,
export v2, admin export) so a round-trip through Excel becomes a true
backup/restore — every entity, every linkage, every constraint and every term
tag survives.
"""
from __future__ import annotations

import json
from typing import Any

from sqlalchemy.orm import Session

from ..core.models import (
    Booking,
    Course,
    CourseUnit,
    Qualification,
    QualificationTimeWindow,
    Room,
    Staff,
    StaffAvailability,
    StaffCompetency,
    StaffPreference,
    StaffQualificationOnlineStudents,
    StaffUnitOnlineStudents,
    Unit,
    UnitAllowedRoom,
    UnitQualification,
    Week,
)
from ..core.unit_brackets import (
    apply_unit_bracket_fields_from_names,
    normalize_component_codes_commas,
    split_class_title_and_unit_codes,
)


PAYLOAD_VERSION = 1
BACKUP_SHEET_NAME = "__timetable_data__"
_BACKUP_CHUNK_SIZE = 30000


def _unit_row_for_restore(u: dict[str, Any]) -> dict[str, Any]:
    """Move trailing ``(UNITCODE)`` from ``name`` into ``component_codes`` when restoring.

    Older exports stored the full spreadsheet label in ``name``; newer imports
    keep the title in ``name`` and codes in ``component_codes``. This keeps
    restore consistent without mutating rows that already have explicit codes.
    """
    out = dict(u)

    def normalize_units_field() -> dict[str, Any]:
        cc = (out.get("component_codes") or "").strip()
        if cc:
            n = normalize_component_codes_commas(cc)
            if n:
                out["component_codes"] = n
        return out

    name = (out.get("name") or "").strip()
    if not name:
        return normalize_units_field()
    existing_codes = (out.get("component_codes") or "").strip()
    title, suffix = split_class_title_and_unit_codes(name)
    if suffix is None:
        return normalize_units_field()
    if existing_codes and existing_codes != suffix:
        return normalize_units_field()
    out["name"] = title
    out["component_codes"] = suffix
    return normalize_units_field()


def _prepare_units_for_restore(units: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Normalize unit rows and ensure ``name`` values are unique for insert."""
    prepared: list[dict[str, Any]] = []
    seen: dict[str, int] = {}
    for raw in units:
        row = _unit_row_for_restore(raw)
        base = (row.get("name") or "").strip() or f"Class #{row['id']}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        if count:
            row = dict(row)
            row["name"] = f"{base} ({count + 1})"
        prepared.append(row)
    return prepared


def serialize(session: Session, *, timetable_session_id: int | None = None) -> dict[str, Any]:
    """Build a plain-dict snapshot of the session's data.

    When ``timetable_session_id`` is set (web multi-session DB), only rows for that
    timetable session are included.
    """
    qual_q = session.query(Qualification).order_by(Qualification.id)
    course_q = session.query(Course).order_by(Course.id)
    unit_q = session.query(Unit).order_by(Unit.id)
    staff_q = session.query(Staff).order_by(Staff.id)
    room_q = session.query(Room).order_by(Room.id)
    if timetable_session_id is not None:
        qual_q = qual_q.filter(Qualification.timetable_session_id == timetable_session_id)
        course_q = course_q.filter(Course.timetable_session_id == timetable_session_id)
        unit_q = unit_q.filter(Unit.timetable_session_id == timetable_session_id)
        staff_q = staff_q.filter(Staff.timetable_session_id == timetable_session_id)
        room_q = room_q.filter(Room.timetable_session_id == timetable_session_id)

    qualifications = qual_q.all()
    courses = course_q.all()
    units = unit_q.all()
    staff_rows = staff_q.all()
    rooms = room_q.all()

    qual_ids = {q.id for q in qualifications}
    course_ids = {c.id for c in courses}
    unit_ids = {u.id for u in units}
    staff_ids = {s.id for s in staff_rows}

    week_ids: set[int] | None = None
    if timetable_session_id is not None:
        from ..core.models import Semester, Week

        sem = (
            session.query(Semester)
            .filter(Semester.timetable_session_id == timetable_session_id)
            .first()
        )
        if sem is not None:
            week_ids = {
                wid
                for (wid,) in session.query(Week.id).filter(Week.semester_id == sem.id).all()
            }

    booking_q = session.query(Booking).order_by(Booking.id)
    if week_ids is not None:
        booking_q = booking_q.filter(
            Booking.week_id.in_(week_ids or [-1]),
            Booking.course_id.in_(course_ids or [-1]),
        )

    return {
        "version": PAYLOAD_VERSION,
        "qualifications": [
            {
                "id": q.id,
                "name": q.name,
                "num_groups": getattr(q, "num_groups", 1) or 1,
                "schedule_period": getattr(q, "schedule_period", None) or "day",
                "delivery_mode": getattr(q, "delivery_mode", None) or "regular",
                "block_week_count": getattr(q, "block_week_count", None),
                "block_start_semester_week": getattr(q, "block_start_semester_week", None),
            }
            for q in qualifications
        ],
        "qualification_time_windows": [
            {
                "qualification_id": w.qualification_id,
                "day": w.day,
                "start_slot": w.start_slot,
                "end_slot": w.end_slot,
            }
            for w in session.query(QualificationTimeWindow).all()
            if w.qualification_id in qual_ids
        ],
        "courses": [
            {
                "id": c.id,
                "code": c.code,
                "name": c.name,
                "qualification_id": c.qualification_id,
                "timetable_locked": getattr(c, "timetable_locked", 0) or 0,
                "sidebar_order": getattr(c, "sidebar_order", 0) or 0,
                "block_week_count": getattr(c, "block_week_count", None),
                "block_start_semester_week": getattr(c, "block_start_semester_week", None),
                "is_block_cohort": getattr(c, "is_block_cohort", 0) or 0,
            }
            for c in courses
        ],
        "units": [
            {
                "id": u.id, "name": u.name,
                "length_slots": u.length_slots,
                "component_codes": u.component_codes,
                "required_room_type": u.required_room_type,
                "required_capacity": u.required_capacity,
                "double_session": getattr(u, "double_session", 0) or 0,
                "double_session_same_day": getattr(u, "double_session_same_day", None),
                "double_session_first_slots": getattr(u, "double_session_first_slots", None),
            }
            for u in units
        ],
        "unit_qualifications": [
            {"unit_id": uq.unit_id, "qualification_id": uq.qualification_id}
            for uq in session.query(UnitQualification).all()
            if uq.unit_id in unit_ids and uq.qualification_id in qual_ids
        ],
        "unit_allowed_rooms": [
            {"unit_id": ur.unit_id, "room_id": ur.room_id}
            for ur in session.query(UnitAllowedRoom).all()
            if ur.unit_id in unit_ids
        ],
        "course_units": [
            {"course_id": cu.course_id, "unit_id": cu.unit_id}
            for cu in session.query(CourseUnit).all()
            if cu.course_id in course_ids and cu.unit_id in unit_ids
        ],
        "staff": [
            {
                "id": s.id, "name": s.name,
                "cost_centre": getattr(s, "cost_centre", None),
                "staff_identifier": getattr(s, "staff_identifier", None),
                "max_hours_per_week": s.max_hours_per_week,
                "non_teaching_day": s.non_teaching_day,
                "fte": getattr(s, "fte", None),
                "ot_hours": getattr(s, "ot_hours", None),
                "development_project_hours": getattr(s, "development_project_hours", None),
                "development_project_description": getattr(
                    s, "development_project_description", None
                ),
                "tae_hours": getattr(s, "tae_hours", None),
                "supervision_hours": getattr(s, "supervision_hours", None),
                "default_online_students_per_class": getattr(
                    s, "default_online_students_per_class", None
                ),
                "timetable_locked": getattr(s, "timetable_locked", 0) or 0,
                "sidebar_order": getattr(s, "sidebar_order", 0) or 0,
            }
            for s in staff_rows
        ],
        "staff_qualification_online_students": [
            {
                "staff_id": row.staff_id,
                "qualification_id": row.qualification_id,
                "student_count": row.student_count,
            }
            for row in session.query(StaffQualificationOnlineStudents).all()
            if row.staff_id in staff_ids and row.qualification_id in qual_ids
        ],
        "staff_unit_online_students": [
            {
                "staff_id": row.staff_id,
                "unit_id": row.unit_id,
                "student_count": row.student_count,
            }
            for row in session.query(StaffUnitOnlineStudents).all()
            if row.staff_id in staff_ids and row.unit_id in unit_ids
        ],
        "staff_preferences": [
            {
                "id": p.id,
                "staff_id": p.staff_id,
                "priority": p.priority,
                "slot_number": p.slot_number,
                "qualification_name": p.qualification_name,
                "class_name": p.class_name,
                "unit_id": p.unit_id,
            }
            for p in session.query(StaffPreference).order_by(StaffPreference.id).all()
            if p.staff_id in staff_ids
        ],
        "staff_competencies": [
            {"staff_id": sc.staff_id, "unit_id": sc.unit_id}
            for sc in session.query(StaffCompetency).all()
            if sc.staff_id in staff_ids and sc.unit_id in unit_ids
        ],
        "staff_availability": [
            {
                "id": a.id, "staff_id": a.staff_id,
                "day": a.day, "start_slot": a.start_slot, "end_slot": a.end_slot,
            }
            for a in session.query(StaffAvailability).all()
            if a.staff_id in staff_ids
        ],
        "rooms": [
            {
                "id": r.id, "code": r.code, "name": r.name,
                "room_type": r.room_type, "capacity": r.capacity,
            }
            for r in rooms
        ],
        "bookings": [
            {
                "id": b.id, "course_id": b.course_id,
                "unit_id": b.unit_id, "staff_id": b.staff_id,
                "sfs_co_teacher_staff_id": getattr(b, "sfs_co_teacher_staff_id", None),
                "sfs_co_teacher_in_term_1": int(getattr(b, "sfs_co_teacher_in_term_1", 0)),
                "sfs_co_teacher_in_term_2": int(getattr(b, "sfs_co_teacher_in_term_2", 0)),
                "cover_staff_id": getattr(b, "cover_staff_id", None),
                "room_id": b.room_id,
                "day": b.day, "start_slot": b.start_slot, "end_slot": b.end_slot,
                "notes": b.notes,
                "external_id": b.external_id,
                "in_term_1": int(getattr(b, "in_term_1", 1)),
                "in_term_2": int(getattr(b, "in_term_2", 1)),
                "online_student_count": getattr(b, "online_student_count", None),
                "lock_time": getattr(b, "lock_time", 0) or 0,
                "lock_staff": getattr(b, "lock_staff", 0) or 0,
                "session_part": getattr(b, "session_part", 1) or 1,
                "session_weeks": getattr(b, "session_weeks", None),
                "block_week_index": getattr(b, "block_week_index", None),
                "combined_class_group_id": getattr(b, "combined_class_group_id", None),
            }
            for b in booking_q.all()
        ],
    }


def deserialize(session: Session, payload: dict[str, Any]) -> None:
    """Replace the session's content with the payload.

    Preserves the current Week's id so anything elsewhere referencing it
    stays valid; bookings are remapped to that week.
    """
    if not isinstance(payload, dict) or payload.get("version") not in (PAYLOAD_VERSION,):
        raise ValueError(
            f"Unknown payload version: {payload.get('version') if isinstance(payload, dict) else '?'}"
        )

    # 1. Wipe existing data in dependency order.
    for model in (
        Booking,
        StaffAvailability,
        StaffPreference,
        StaffCompetency,
        StaffQualificationOnlineStudents,
        StaffUnitOnlineStudents,
        UnitAllowedRoom,
        UnitQualification,
        QualificationTimeWindow,
        CourseUnit,
    ):
        session.query(model).delete()
    # Order matters for FKs (CourseUnit must go before Course / Unit).
    session.query(Course).update({Course.qualification_id: None})
    session.flush()
    session.query(Course).delete()
    session.query(Unit).delete()
    session.query(Qualification).delete()
    session.query(Staff).delete()
    session.query(Room).delete()
    session.flush()

    # 2. Re-insert with original ids. Flush after each table so FK-dependent
    # rows that follow (e.g. Course → Qualification) see their parents.
    for q in payload.get("qualifications", []):
        session.add(Qualification(
            id=q["id"],
            name=q["name"],
            num_groups=q.get("num_groups", 1) or 1,
            schedule_period=q.get("schedule_period") or "day",
            delivery_mode=q.get("delivery_mode") or "regular",
            block_week_count=q.get("block_week_count"),
            block_start_semester_week=q.get("block_start_semester_week"),
        ))
    session.flush()
    for u in _prepare_units_for_restore(payload.get("units", [])):
        session.add(Unit(
            id=u["id"], name=u["name"],
            length_slots=u.get("length_slots"),
            component_codes=u.get("component_codes"),
            required_room_type=u.get("required_room_type"),
            required_capacity=u.get("required_capacity"),
            double_session=u.get("double_session", 0) or 0,
            double_session_same_day=u.get("double_session_same_day"),
            double_session_first_slots=u.get("double_session_first_slots"),
        ))
    session.flush()
    for c in payload.get("courses", []):
        session.add(Course(
            id=c["id"],
            code=c["code"],
            name=c.get("name"),
            qualification_id=c.get("qualification_id"),
            timetable_locked=c.get("timetable_locked", 0) or 0,
            sidebar_order=c.get("sidebar_order", 0) or 0,
            block_week_count=c.get("block_week_count"),
            block_start_semester_week=c.get("block_start_semester_week"),
            is_block_cohort=c.get("is_block_cohort", 0) or 0,
        ))
    session.flush()
    for s in payload.get("staff", []):
        session.add(Staff(
            id=s["id"], name=s["name"],
            cost_centre=s.get("cost_centre"),
            staff_identifier=s.get("staff_identifier"),
            max_hours_per_week=s.get("max_hours_per_week"),
            non_teaching_day=s.get("non_teaching_day"),
            fte=s.get("fte"),
            ot_hours=s.get("ot_hours"),
            development_project_hours=s.get("development_project_hours"),
            development_project_description=s.get("development_project_description"),
            tae_hours=s.get("tae_hours"),
            supervision_hours=s.get("supervision_hours"),
            default_online_students_per_class=s.get("default_online_students_per_class"),
            timetable_locked=s.get("timetable_locked", 0) or 0,
            sidebar_order=s.get("sidebar_order", 0) or 0,
        ))
    for r in payload.get("rooms", []):
        session.add(Room(
            id=r["id"], code=r["code"], name=r.get("name"),
            room_type=r.get("room_type"), capacity=r.get("capacity"),
        ))
    session.flush()

    # 3. Linkage tables.
    for w in payload.get("qualification_time_windows", []):
        session.add(QualificationTimeWindow(
            qualification_id=w["qualification_id"],
            day=w["day"], start_slot=w["start_slot"], end_slot=w["end_slot"],
        ))
    for uq in payload.get("unit_qualifications", []):
        session.add(UnitQualification(unit_id=uq["unit_id"], qualification_id=uq["qualification_id"]))
    for ur in payload.get("unit_allowed_rooms", []):
        session.add(UnitAllowedRoom(unit_id=ur["unit_id"], room_id=ur["room_id"]))
    for cu in payload.get("course_units", []):
        session.add(CourseUnit(course_id=cu["course_id"], unit_id=cu["unit_id"]))
    for sc in payload.get("staff_competencies", []):
        session.add(StaffCompetency(staff_id=sc["staff_id"], unit_id=sc["unit_id"]))
    for row in payload.get("staff_qualification_online_students", []):
        session.add(
            StaffQualificationOnlineStudents(
                staff_id=row["staff_id"],
                qualification_id=row["qualification_id"],
                student_count=row.get("student_count"),
            )
        )
    for row in payload.get("staff_unit_online_students", []):
        session.add(
            StaffUnitOnlineStudents(
                staff_id=row["staff_id"],
                unit_id=row["unit_id"],
                student_count=row.get("student_count"),
            )
        )
    for a in payload.get("staff_availability", []):
        session.add(StaffAvailability(
            id=a["id"], staff_id=a["staff_id"],
            day=a["day"], start_slot=a["start_slot"], end_slot=a["end_slot"],
        ))
    for p in payload.get("staff_preferences", []):
        session.add(StaffPreference(
            id=p.get("id"),
            staff_id=p["staff_id"],
            priority=p["priority"],
            slot_number=p["slot_number"],
            qualification_name=p.get("qualification_name"),
            class_name=p.get("class_name"),
            unit_id=p.get("unit_id"),
        ))

    # 4. Bookings — remapped to the current Week.
    week = session.query(Week).order_by(Week.id).first()
    if week is None:
        raise RuntimeError("No Week row in this session; can't restore bookings.")
    for b in payload.get("bookings", []):
        session.add(Booking(
            id=b["id"], week_id=week.id,
            course_id=b["course_id"], unit_id=b.get("unit_id"),
            staff_id=b.get("staff_id"),
            sfs_co_teacher_staff_id=b.get("sfs_co_teacher_staff_id"),
            sfs_co_teacher_in_term_1=(
                b.get("sfs_co_teacher_in_term_1", b.get("in_term_1", 1))
                if b.get("sfs_co_teacher_staff_id")
                else 0
            ),
            sfs_co_teacher_in_term_2=(
                b.get("sfs_co_teacher_in_term_2", b.get("in_term_2", 1))
                if b.get("sfs_co_teacher_staff_id")
                else 0
            ),
            cover_staff_id=b.get("cover_staff_id"),
            room_id=b.get("room_id"),
            day=b["day"], start_slot=b["start_slot"], end_slot=b["end_slot"],
            notes=b.get("notes"),
            external_id=b.get("external_id"),
            in_term_1=b.get("in_term_1", 1),
            in_term_2=b.get("in_term_2", 1),
            online_student_count=b.get("online_student_count"),
            lock_time=b.get("lock_time", 0) or 0,
            lock_staff=b.get("lock_staff", 0) or 0,
            session_part=b.get("session_part", 1) or 1,
            session_weeks=b.get("session_weeks"),
            block_week_index=b.get("block_week_index"),
            combined_class_group_id=b.get("combined_class_group_id"),
        ))
    apply_unit_bracket_fields_from_names(session)
    session.commit()


def write_backup_sheet(
    wb, session: Session, *, timetable_session_id: int | None = None
) -> None:
    """Embed a hidden sheet with the full structured session payload."""
    if BACKUP_SHEET_NAME in wb.sheetnames:
        del wb[BACKUP_SHEET_NAME]
    ws = wb.create_sheet(BACKUP_SHEET_NAME)
    ws["A1"] = (
        "DO NOT EDIT — round-trip backup for the timetabling app. "
        f"Payload v{PAYLOAD_VERSION}. Visual sheets are regenerated on import."
    )
    payload = json.dumps(
        serialize(session, timetable_session_id=timetable_session_id),
        separators=(",", ":"),
    )
    for idx, start in enumerate(range(0, len(payload), _BACKUP_CHUNK_SIZE)):
        ws.cell(row=idx + 2, column=1, value=payload[start : start + _BACKUP_CHUNK_SIZE])
    ws.sheet_state = "hidden"


def read_backup_payload(xlsm_path: str) -> dict[str, Any]:
    """Load the embedded session payload from an export workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsm_path, data_only=True, keep_vba=False)
    if BACKUP_SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(
            "Workbook is not a valid timetable export: missing embedded "
            f"'{BACKUP_SHEET_NAME}' backup sheet."
        )
    ws_meta = wb[BACKUP_SHEET_NAME]
    chunks: list[str] = []
    row = 2
    while True:
        v = ws_meta.cell(row=row, column=1).value
        if v in (None, ""):
            break
        chunks.append(str(v))
        row += 1
    wb.close()
    return json.loads("".join(chunks))
