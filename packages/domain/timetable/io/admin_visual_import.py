"""Import timetable data from admin-export visual spreadsheets.

Reads one course tab per sheet (style-guide layout): TIME / Lecturer / Room label
columns and merged week bands for Term 3–4. Class cells use the placecard format::

    Class title (UNIT1, UNIT2) ID: 1234567

This path parses visible cells only (no embedded ``__timetable_data__`` backup).
"""
from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import time

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from ..constants import NUM_DAYS, time_to_slot
from ..core.booking_staff import apply_parsed_lecturers_to_booking, parse_import_lecturer_label
from ..core.booking_sessions import initialize_session_weeks, serialize_session_weeks
from ..core.models import Booking, Course, CourseUnit, Qualification, Room, Staff, Unit, UnitQualification
from ..core.qualification_schedule import SCHEDULE_PERIOD_DAY, replace_qualification_time_windows
from ..core.unit_brackets import (
    apply_unit_bracket_fields_from_names,
    normalize_class_label_for_parse,
    split_class_title_and_unit_codes,
)
from .admin_export import (
    TERM1_LABEL_COLS,
    TERM1_WEEK_COLS,
    TERM2_LABEL_COLS,
    TERM2_WEEK_COLS,
    _COURSE_TITLE_COL,
    _COURSE_TITLE_ROW,
)
from .backup_payload import BACKUP_SHEET_NAME
from .overall_visual_import import (
    OverallVisualImportReport,
    _base_qualification_name,
    _extract_placecard_id,
    _join_unit_parts,
    _lecturer_placeholder,
    _normalize_lecturer_name,
    _parse_external_and_terms,
    _resolve_unit_storage_name,
)
from .visual_import_session import VisualImportContext

_TIME_RANGE_RE = re.compile(r"^(\d{1,2}):(\d{2})-(\d{1,2}):(\d{2})$")
_WEEK_NOISE_RE = re.compile(
    r"^(week\s*\d+|holiday|public\s+holiday|term\s*break|pd\s*week|accrued\s+pd|section|professional|development)$",
    re.I,
)

_DAY_ROW_BANDS: tuple[tuple[int, range], ...] = (
    (0, range(9, 15)),
    (1, range(18, 24)),
    (2, range(27, 33)),
    (3, range(36, 42)),
    (4, range(45, 51)),
)


@dataclass
class AdminVisualImportReport(OverallVisualImportReport):
    """Alias report type for admin visual imports."""


def is_admin_visual_workbook(path: str) -> bool:
    """True when the workbook looks like a course-per-tab admin export grid."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = [n for n in wb.sheetnames if n != BACKUP_SHEET_NAME]
        if not sheets:
            return False
        for name in sheets[:5]:
            ws = wb[name]
            title = ws.cell(_COURSE_TITLE_ROW, _COURSE_TITLE_COL).value
            if title and str(title).strip().lower().startswith("course:"):
                return True
            probe = ws["N17"].value
            if probe is not None and str(probe).strip() == "Week 10":
                return True
        return False
    finally:
        wb.close()


def _strip_cell(val: object | None) -> str:
    if val is None:
        return ""
    if hasattr(val, "strftime"):
        return ""
    return str(val).strip()


def _day_for_row(row: int) -> int | None:
    for day, rows in _DAY_ROW_BANDS:
        if row in rows:
            return day
    return None


def _parse_time_range(text: str) -> tuple[int, int] | None:
    m = _TIME_RANGE_RE.match(text.strip())
    if not m:
        return None
    try:
        start = time(int(m.group(1)), int(m.group(2)))
        end = time(int(m.group(3)), int(m.group(4)))
        start_slot = time_to_slot(start)
        end_slot = time_to_slot(end)
    except ValueError:
        return None
    if end_slot <= start_slot:
        return None
    return start_slot, end_slot


def _col_to_semester_week(col: int) -> int | None:
    if col in TERM1_WEEK_COLS:
        return col - TERM1_WEEK_COLS[0] + 1
    if col in TERM2_WEEK_COLS:
        return col - TERM2_WEEK_COLS[0] + 11
    return None


def _is_week_noise(text: str) -> bool:
    t = text.strip()
    if not t:
        return True
    if _WEEK_NOISE_RE.match(t):
        return True
    if t.lower().startswith("week "):
        return True
    return False


def _class_text_from_band(text: str) -> str | None:
    s = text.strip()
    if not s or _is_week_noise(s):
        return None
    if _TIME_RANGE_RE.match(s):
        return None
    return s


def _merged_row_text_by_col(ws: Worksheet, row: int) -> dict[int, str]:
    out: dict[int, str] = {}
    week_cols = set(TERM1_WEEK_COLS) | set(TERM2_WEEK_COLS)
    for merge in ws.merged_cells.ranges:
        if merge.min_row != row or merge.max_row != row:
            continue
        raw = _strip_cell(ws.cell(merge.min_row, merge.min_col).value)
        cls = _class_text_from_band(raw)
        if not cls:
            continue
        for col in range(merge.min_col, merge.max_col + 1):
            if col in week_cols:
                out[col] = cls
    for col in week_cols:
        if col in out:
            continue
        cls = _class_text_from_band(_strip_cell(ws.cell(row, col).value))
        if cls:
            out[col] = cls
    return out


def _active_semester_weeks(text_by_col: dict[int, str]) -> list[int]:
    weeks: set[int] = set()
    for col in text_by_col:
        week = _col_to_semester_week(col)
        if week is not None:
            weeks.add(week)
    return sorted(weeks)


def _course_code_from_sheet(ws: Worksheet, sheet_name: str) -> str:
    if sheet_name and sheet_name.strip():
        return sheet_name.strip()[:80]
    title = _strip_cell(ws.cell(_COURSE_TITLE_ROW, _COURSE_TITLE_COL).value)
    if title.lower().startswith("course:"):
        title = title.split(":", 1)[1].strip()
    return title[:80] if title else "Untitled"


def _parse_course_tab(
    session: Session,
    ws: Worksheet,
    *,
    ctx: VisualImportContext,
    course_code: str,
    week_id: int,
    report: AdminVisualImportReport,
    staff_cache: dict[str, int],
    room_cache: dict[str, int],
    unit_cache: dict[str, int],
    course_unit_sets: dict[str, set[int]],
) -> None:
    course = ctx.course_by_code(session, course_code)
    if course is None:
        course = ctx.new_course(course_code)
        session.add(course)
        session.flush()

    def get_staff(name: str) -> int:
        key = name.strip()
        if key in staff_cache:
            return staff_cache[key]
        row = ctx.staff_by_name(session, key)
        if row is None:
            row = ctx.new_staff(key)
            session.add(row)
            session.flush()
            report.staff_created += 1
        staff_cache[key] = row.id
        return row.id

    def resolve_staff_id(name: str | None) -> int | None:
        if name is None:
            return None
        return get_staff(name)

    def get_room(code: str) -> int | None:
        if not code.strip():
            return None
        key = code.strip()
        if key in room_cache:
            return room_cache[key]
        row = ctx.room_by_code(session, key)
        if row is None:
            from ..core.room_types import ROOM_TYPE_ON_CAMPUS

            row = ctx.new_room(key, room_type=ROOM_TYPE_ON_CAMPUS)
            session.add(row)
            session.flush()
            report.rooms_created += 1
        room_cache[key] = row.id
        return row.id

    def get_unit(display_name: str, span_slots: int) -> int | None:
        display_name = display_name.strip()
        if not display_name:
            return None
        split_title, suffix = split_class_title_and_unit_codes(display_name)
        name_key, comp_codes = _resolve_unit_storage_name(
            session, split_title, suffix, timetable_session_id=ctx.timetable_session_id
        )
        for key in (display_name, name_key):
            if key in unit_cache:
                uid = unit_cache[key]
                u = session.get(Unit, uid)
                if u is not None:
                    need = max(span_slots, 1)
                    if u.length_slots is None or u.length_slots < need:
                        u.length_slots = need
                    if comp_codes is not None and not (u.component_codes or "").strip():
                        u.component_codes = comp_codes
                unit_cache[display_name] = uid
                unit_cache[name_key] = uid
                return uid
        row = ctx.unit_by_name(session, name_key)
        if row is None:
            row = ctx.new_unit(
                name_key,
                length_slots=max(span_slots, 1),
                component_codes=comp_codes,
            )
            session.add(row)
            session.flush()
            report.units_created += 1
        else:
            need = max(span_slots, 1)
            if row.length_slots is None or row.length_slots < need:
                row.length_slots = need
            if comp_codes is not None and not (row.component_codes or "").strip():
                row.component_codes = comp_codes
        unit_cache[name_key] = row.id
        unit_cache[display_name] = row.id
        return row.id

    linked_course_units: set[tuple[int, int]] = set()

    def ensure_course_unit(unit_id: int | None) -> None:
        if unit_id is None:
            return
        key = (course.id, unit_id)
        if key in linked_course_units:
            return
        exists = (
            session.query(CourseUnit)
            .filter_by(course_id=course.id, unit_id=unit_id)
            .first()
        )
        if not exists:
            session.add(CourseUnit(course_id=course.id, unit_id=unit_id))
        linked_course_units.add(key)

    import_rows: set[int] = set()
    for _day, rows in _DAY_ROW_BANDS:
        import_rows.update(rows)
    for row in sorted(import_rows):
        day = _day_for_row(row)
        if day is None or day >= NUM_DAYS:
            continue

        time_text = _strip_cell(ws.cell(row, TERM1_LABEL_COLS[0]).value)
        if not time_text:
            time_text = _strip_cell(ws.cell(row, TERM2_LABEL_COLS[0]).value)
        slots = _parse_time_range(time_text)
        if slots is None:
            continue
        start_slot, end_slot = slots
        span = end_slot - start_slot

        lec_raw = _strip_cell(ws.cell(row, TERM1_LABEL_COLS[1]).value)
        if _lecturer_placeholder(lec_raw):
            lec_raw = _strip_cell(ws.cell(row, TERM2_LABEL_COLS[1]).value)
        primary, co_name, co_t1, co_t2 = parse_import_lecturer_label(lec_raw)
        primary_norm = _normalize_lecturer_name(primary or "")
        if not primary_norm or _lecturer_placeholder(primary or ""):
            continue

        room_raw = _strip_cell(ws.cell(row, TERM1_LABEL_COLS[2]).value)
        if not room_raw:
            room_raw = _strip_cell(ws.cell(row, TERM2_LABEL_COLS[2]).value)

        text_by_col = _merged_row_text_by_col(ws, row)
        if not text_by_col:
            continue

        unique_labels: dict[str, str] = {}
        placecard_id: str | None = None
        for raw in text_by_col.values():
            cls = _class_text_from_band(raw)
            if not cls:
                continue
            norm = normalize_class_label_for_parse(cls)
            remainder, pid = _extract_placecard_id(norm)
            if pid and placecard_id is None:
                placecard_id = pid
            key = remainder.casefold()
            unique_labels.setdefault(key, remainder)
        joined_class = _join_unit_parts(sorted(unique_labels.values()))
        if not joined_class:
            continue

        active_weeks = _active_semester_weeks(text_by_col)
        in_term_1 = any(1 <= w <= 10 for w in active_weeks)
        in_term_2 = any(11 <= w <= 20 for w in active_weeks)
        if not in_term_1 and not in_term_2:
            continue

        display, ext, it1, it2 = _parse_external_and_terms(
            joined_class, placecard_id=placecard_id
        )

        uid = get_unit(display, span)
        rid = get_room(room_raw)
        co_norm = _normalize_lecturer_name(co_name or "") if co_name else None
        ensure_course_unit(uid)
        if uid is not None:
            course_unit_sets[course_code].add(int(uid))

        booking = Booking(
            week_id=week_id,
            course_id=course.id,
            unit_id=uid,
            room_id=rid,
            day=day,
            start_slot=start_slot,
            end_slot=end_slot,
            external_id=ext,
            in_term_1=1 if in_term_1 else 0,
            in_term_2=1 if in_term_2 else 0,
        )
        apply_parsed_lecturers_to_booking(
            booking,
            primary_name=primary_norm,
            co_teacher_name=co_norm if co_norm and not _lecturer_placeholder(co_norm) else None,
            co_in_term_1=co_t1,
            co_in_term_2=co_t2,
            resolve_staff_id=resolve_staff_id,
        )
        if active_weeks:
            booking.session_weeks = serialize_session_weeks(active_weeks)
        else:
            booking.in_term_1 = it1
            booking.in_term_2 = it2
            initialize_session_weeks(booking)
        session.add(booking)
        report.bookings_created += 1


def import_admin_visual(
    session: Session,
    xlsx_path: str,
    *,
    timetable_session_id: int | None = None,
) -> AdminVisualImportReport:
    """Parse admin-export course tabs and replace all bookings for the first week."""
    if not is_admin_visual_workbook(xlsx_path):
        raise ValueError(
            "Workbook does not look like an admin export (expected course tabs with "
            "TIME/Lecturer/Room rows and week bands)."
        )

    ctx = VisualImportContext(timetable_session_id)
    wb = load_workbook(xlsx_path, data_only=True, keep_vba=False)
    try:
        sheet_names = [n for n in wb.sheetnames if n != BACKUP_SHEET_NAME]
        if not sheet_names:
            raise ValueError("Workbook has no course sheets.")

        week = ctx.first_week(session)

        report = AdminVisualImportReport()
        report.courses_touched = len(sheet_names)

        session.query(Booking).filter(Booking.week_id == week.id).delete()

        staff_cache: dict[str, int] = {}
        room_cache: dict[str, int] = {}
        unit_cache: dict[str, int] = {}
        course_unit_sets: dict[str, set[int]] = defaultdict(set)

        for name in sheet_names:
            ws = wb[name]
            course_code = _course_code_from_sheet(ws, name)
            try:
                _parse_course_tab(
                    session,
                    ws,
                    ctx=ctx,
                    course_code=course_code,
                    week_id=week.id,
                    report=report,
                    staff_cache=staff_cache,
                    room_cache=room_cache,
                    unit_cache=unit_cache,
                    course_unit_sets=course_unit_sets,
                )
            except Exception as exc:
                report.warnings.append(f"{name}: {exc}")

        sig_to_courses: dict[frozenset[int], list[str]] = defaultdict(list)
        for cc, units in course_unit_sets.items():
            if units:
                sig_to_courses[frozenset(units)].append(cc)

        used_qual_names: set[str] = set()
        linked_unit_quals: set[tuple[int, int]] = set()
        for sig, course_codes in sig_to_courses.items():
            base_name = _base_qualification_name(course_codes[0])
            qual_name = base_name
            n = 1
            while qual_name in used_qual_names:
                n += 1
                qual_name = f"{base_name} ({n})"
            used_qual_names.add(qual_name)

            q = ctx.qualification_by_name(session, qual_name)
            if q is None:
                q = ctx.new_qualification(
                    name=qual_name,
                    num_groups=len(course_codes),
                    schedule_period=SCHEDULE_PERIOD_DAY,
                )
                session.add(q)
                session.flush()
                replace_qualification_time_windows(session, q)
                report.qualifications_created += 1
            else:
                q.num_groups = len(course_codes)

            for cc in course_codes:
                c = ctx.course_by_code(session, cc)
                if c is not None:
                    c.qualification_id = q.id
            for uid in sig:
                uq_key = (uid, q.id)
                if uq_key in linked_unit_quals:
                    continue
                exists = (
                    session.query(UnitQualification)
                    .filter_by(unit_id=uid, qualification_id=q.id)
                    .first()
                )
                if not exists:
                    session.add(UnitQualification(unit_id=uid, qualification_id=q.id))
                linked_unit_quals.add(uq_key)

        apply_unit_bracket_fields_from_names(
            session, timetable_session_id=timetable_session_id
        )
        session.commit()
        return report
    finally:
        wb.close()
