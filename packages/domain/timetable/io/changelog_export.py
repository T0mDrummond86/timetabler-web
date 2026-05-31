"""Write timetabling change log to a plain Excel workbook."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .export_headers import CHANGE_LOG_EXPORT_HEADERS

# Desktop-only display row type; web export uses dict rows until shared helpers land.
ChangeLogDisplayRow = dict  # type: ignore[misc,assignment]


def write_change_log_xlsx(
    path: str | Path,
    rows: list[tuple[str, str, dict[str, str]]] | list[ChangeLogDisplayRow],
    *,
    sheet_title: str = "Change log",
) -> str:
    """Write one sheet with headers + data. ``rows`` are ``(when, action, row_dict)``."""
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Change log"

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(start_color="FF374151", end_color="FF374151", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = [*CHANGE_LOG_EXPORT_HEADERS, "Notes"]
    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    body_align = Alignment(vertical="top", wrap_text=True)
    for r, src in enumerate(rows, start=2):
        when_s = ""
        action = ""
        row: dict[str, str] = {}
        note = ""
        if isinstance(src, ChangeLogDisplayRow):
            when_s, action, row, note = src.when, src.action, src.row, src.note
        else:
            when_s, action, row = src
        ws.cell(row=r, column=1, value=row.get("id", "")).alignment = body_align
        ws.cell(row=r, column=2, value=row.get("group", "")).alignment = body_align
        ws.cell(row=r, column=3, value=row.get("class", "")).alignment = body_align
        ws.cell(row=r, column=4, value=row.get("lecturer_change", "")).alignment = body_align
        ws.cell(row=r, column=5, value=row.get("time_change", "")).alignment = body_align
        ws.cell(row=r, column=6, value=row.get("day_change", "")).alignment = body_align
        ws.cell(row=r, column=7, value=row.get("room_change", "")).alignment = body_align
        ws.cell(row=r, column=8, value=row.get("delete", "")).alignment = body_align
        ws.cell(row=r, column=9, value=when_s).alignment = body_align
        ws.cell(row=r, column=10, value=action).alignment = body_align
        ws.cell(row=r, column=11, value=note).alignment = body_align

    widths = (14, 22, 36, 24, 22, 14, 12, 8, 18, 10, 30)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(path)
    return str(path)
