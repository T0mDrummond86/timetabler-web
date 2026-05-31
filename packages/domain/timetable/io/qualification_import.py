"""Bulk-import qualifications + classes from the QInputTemplate.xlsx format.

Template structure
==================
Each sheet (Q1..Q8) holds one qualification:
  - Cell A1: qualification name (default placeholder = '[qualification name and stage]')
  - Then 8 class blocks of 6 rows each, starting at row 2:
      row N+0 (col A):  'Class N'                              (header label)
      row N+1 (col A):  'class name:'                          (col B = value)
      row N+2 (col A):  'units:'                               (col B = value)
      row N+3 (col A):  'running time:'                        (col B = value)
      row N+4 (col A):  'acceptable rooms (blank for N/A):'    (col B = value)
      row N+5 (col A):  'acceptable lecturers (blank for N/A):'(col B = value)

Behaviour
=========
- Sheets where A1 is blank or still the placeholder are skipped.
- Class blocks where 'class name:' is blank are skipped.
- A class is uniquely identified by its name across the whole import: if the
  same class name appears under multiple qualifications, we create one Unit
  and link it to each qualification that mentions it.
- 'running time:' is parsed as hours (accepts '2', '2h', '2 hours', '1.5');
  stored as `length_slots` (each slot = 30 minutes).
- 'acceptable rooms' is comma- or newline-separated room codes; matched
  against existing Room rows by code (case-insensitive). Unknown codes are
  warned about, not auto-created.
- 'acceptable lecturers' likewise matched against Staff names. Unknowns
  warned about.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..core.models import (
    Course,
    Qualification,
    Room,
    Staff,
    StaffCompetency,
    Unit,
    UnitAllowedRoom,
    UnitQualification,
)
from ..core.qualification_schedule import (
    SCHEDULE_PERIOD_DAY,
    replace_qualification_time_windows,
)
from ..core.unit_brackets import apply_unit_bracket_fields_from_names, split_class_title_and_unit_codes


PLACEHOLDER_QNAME = "[qualification name and stage]"


@dataclass
class QualImportReport:
    qualifications_created: int = 0
    qualifications_linked: int = 0
    classes_created: int = 0
    classes_updated: int = 0
    class_qual_links_added: int = 0
    room_links_added: int = 0
    lecturer_links_added: int = 0
    courses_created: int = 0
    warnings: list[str] = field(default_factory=list)


def _text(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _parse_running_time(v) -> int | None:
    """Parse a hours-like value into number of half-hour slots."""
    s = _text(v)
    if s is None:
        return None
    s = s.lower()
    for suffix in ("hours", "hour", "hrs", "hr", "h"):
        if s.endswith(suffix):
            s = s[: -len(suffix)].strip()
            break
    try:
        hours = float(s)
    except ValueError:
        return None
    slots = int(round(hours * 2))
    return slots if slots > 0 else None


def _parse_csv(v) -> list[str]:
    s = _text(v)
    if s is None or s.lower() in ("n/a", "na", "none", "-"):
        return []
    parts = [p.strip() for p in s.replace("\n", ",").split(",")]
    return [p for p in parts if p]


def import_qualifications_from_template(session: Session, path: str | Path) -> QualImportReport:
    rep = QualImportReport()
    wb = load_workbook(path, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        qual_name = _text(ws.cell(row=1, column=1).value)
        if qual_name is None or qual_name.lower() == PLACEHOLDER_QNAME.lower():
            continue

        qual = session.query(Qualification).filter_by(name=qual_name).first()
        if qual is None:
            qual = Qualification(
                name=qual_name, num_groups=1, schedule_period=SCHEDULE_PERIOD_DAY
            )
            session.add(qual)
            session.flush()
            replace_qualification_time_windows(session, qual)
            rep.qualifications_created += 1
            # Spawn the default group-Course (Group A) so the qualification
            # appears on the timetable straight away.
            default_course_code = f"{qual.name} GrpA"
            existing_course = session.query(Course).filter_by(code=default_course_code).first()
            if existing_course is None:
                session.add(Course(code=default_course_code, qualification_id=qual.id))
                rep.courses_created += 1
            elif existing_course.qualification_id != qual.id:
                existing_course.qualification_id = qual.id
        else:
            rep.qualifications_linked += 1

        # 8 class blocks at rows 2 + i*6
        for i in range(8):
            base = 2 + i * 6  # 'Class N' header row
            class_name = _text(ws.cell(row=base + 1, column=2).value)
            if class_name is None:
                continue

            raw_cn = class_name.strip()
            title, bracket_codes = split_class_title_and_unit_codes(raw_cn)
            storage_name = title if bracket_codes is not None else raw_cn

            existing_unit = session.query(Unit).filter_by(name=storage_name).first()
            if existing_unit is None and bracket_codes is not None:
                legacy = session.query(Unit).filter_by(name=raw_cn).first()
                if legacy is not None:
                    legacy.name = storage_name
                    if bracket_codes and not (legacy.component_codes or "").strip():
                        legacy.component_codes = bracket_codes
                    existing_unit = legacy

            if existing_unit is None:
                unit = Unit(name=storage_name, component_codes=bracket_codes)
                session.add(unit)
                session.flush()
                rep.classes_created += 1
            else:
                unit = existing_unit
                rep.classes_updated += 1

            # Update fields. We only fill blanks — never overwrite manual data.
            comp = _text(ws.cell(row=base + 2, column=2).value)
            if comp and not unit.component_codes:
                unit.component_codes = comp
            length = _parse_running_time(ws.cell(row=base + 3, column=2).value)
            if length and not unit.length_slots:
                unit.length_slots = length

            # Link unit to qualification (idempotent).
            link = (
                session.query(UnitQualification)
                .filter_by(unit_id=unit.id, qualification_id=qual.id)
                .first()
            )
            if link is None:
                session.add(UnitQualification(unit_id=unit.id, qualification_id=qual.id))
                rep.class_qual_links_added += 1

            # Allowed rooms.
            for code in _parse_csv(ws.cell(row=base + 4, column=2).value):
                room = (
                    session.query(Room)
                    .filter(Room.code.ilike(code))
                    .first()
                )
                if room is None:
                    rep.warnings.append(
                        f"{sheet_name}: room {code!r} (for class {class_name!r}) not found in DB"
                    )
                    continue
                if not (
                    session.query(UnitAllowedRoom)
                    .filter_by(unit_id=unit.id, room_id=room.id)
                    .first()
                ):
                    session.add(UnitAllowedRoom(unit_id=unit.id, room_id=room.id))
                    rep.room_links_added += 1

            # Allowed lecturers.
            for name in _parse_csv(ws.cell(row=base + 5, column=2).value):
                staff = (
                    session.query(Staff)
                    .filter(Staff.name.ilike(name))
                    .first()
                )
                if staff is None:
                    rep.warnings.append(
                        f"{sheet_name}: lecturer {name!r} (for class {class_name!r}) not found in DB"
                    )
                    continue
                if not (
                    session.query(StaffCompetency)
                    .filter_by(staff_id=staff.id, unit_id=unit.id)
                    .first()
                ):
                    session.add(StaffCompetency(staff_id=staff.id, unit_id=unit.id))
                    rep.lecturer_links_added += 1

    apply_unit_bracket_fields_from_names(session)
    session.commit()
    return rep
