"""Import staff, rooms, qualifications, classes, and bookings from an aSc export (.xlsx).

Expected workbook layout (standard aSc export sheets)
=====================================================
Teachers (row 1 headers in col B+):
  Name, Short, …, Contract (weekly hours, optional)

Classrooms:
  Name, Short (room code), Type, Home classroom, …

Classes:
  Class (qualification / cohort name), Short (course code hint), …

Subjects:
  Subject (title), Short (unit code), Classrooms (optional default rooms)

Lessons (primary class + constraint source):
  Teacher, Class, Group, Subject, Length, Lessons/week, Available classrooms, Cycle

Contracts Classes (optional, room assignments):
  Class header rows then Teacher, Subject, …, Classrooms (col I)

Per-subject sheets (one tab per unit/subject):
  Row ~3: Day | Lesson headers (cols B–C)
  Data rows: day name, period number (1–18), class name when scheduled

Behaviour
=========
- Creates Staff, Room, Qualification (+ default course), and Unit rows from master sheets.
- Reads the per-subject schedule tabs and creates Week 0 bookings (one block per
  contiguous run of periods with the same class).
- aSc period *n* maps to timetable slot ``n - 1`` (30-minute slots from 08:00).
- Teacher and room come from the Lessons / Contracts Classes sheets.
- Replaces all bookings on the session's first week when schedule data is present.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import re

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..constants import DAYS, NUM_SLOTS
from ..core.models import (
    Booking,
    Course,
    CourseUnit,
    Qualification,
    Room,
    Staff,
    StaffCompetency,
    Unit,
    UnitAllowedRoom,
    UnitQualification,
)
from ..core.qualification_schedule import (
    SCHEDULE_PERIOD_DAY,
    replace_qualification_time_windows,
)
from ..core.unit_brackets import apply_unit_bracket_fields_from_names
from .visual_import_session import VisualImportContext

_REQUIRED_SHEETS = ("Teachers", "Classrooms", "Classes", "Lessons")
_LESSONS_HEADER = ("Teacher", "Class", "Group", "Subject", "Length")
_DAY_LESSON_HEADER = ("Day", "Lesson")
_CORE_SHEETS = frozenset(
    {
        "Teachers",
        "Classrooms",
        "Classes",
        "Subjects",
        "Lessons",
        "Contract",
        "Contracts Classes",
        "Available teachers",
        "Available teachers 2",
        "Work Placement.",
    }
)

# aSc meta tokens in room lists — not physical room codes.
_SKIP_ROOM_LABELS = frozenset(
    {
        "home classroom",
        "subject's classrooms",
        "subjects classrooms",
        "subject classrooms",
        "virtual classroom",
        "computer lab",
        "online",
        "virtual",
    }
)


@dataclass
class AscImportReport:
    staff_created: int = 0
    rooms_created: int = 0
    qualifications_created: int = 0
    qualifications_linked: int = 0
    classes_created: int = 0
    classes_updated: int = 0
    courses_created: int = 0
    class_qual_links_added: int = 0
    room_links_added: int = 0
    lecturer_links_added: int = 0
    bookings_created: int = 0
    warnings: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class _LessonAssignment:
    teacher: str | None
    room_token: str | None


@dataclass
class _ScheduleBlock:
    subject_code: str
    class_name: str
    day: str
    start_period: int
    end_period: int
    sheet_name: str


def _text(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _norm(s: str) -> str:
    return " ".join(s.split()).strip().lower()


def _model_has_session_id(model) -> bool:
    return "timetable_session_id" in model.__table__.columns


def _scoped_query(session: Session, model, timetable_session_id: int | None):
    q = session.query(model)
    if timetable_session_id is not None and _model_has_session_id(model):
        q = q.filter(model.timetable_session_id == timetable_session_id)
    return q


def _scoped_filter_by(session: Session, model, timetable_session_id: int | None, **filters):
    q = session.query(model).filter_by(**filters)
    if timetable_session_id is not None and _model_has_session_id(model):
        q = q.filter(model.timetable_session_id == timetable_session_id)
    return q


def _create_kwargs(model, timetable_session_id: int | None, **fields):
    if timetable_session_id is not None and _model_has_session_id(model):
        fields["timetable_session_id"] = timetable_session_id
    return fields


def _sheet_header_row(ws, expected: tuple[str, ...], *, max_scan: int = 5) -> int | None:
    for row_idx in range(1, max_scan + 1):
        vals = [_text(ws.cell(row=row_idx, column=col).value) for col in range(1, len(expected) + 1)]
        if vals and all(v and _norm(v) == _norm(exp) for v, exp in zip(vals, expected)):
            return row_idx
        vals_b = [_text(ws.cell(row=row_idx, column=col + 1).value) for col in range(1, len(expected) + 1)]
        if vals_b and all(v and _norm(v) == _norm(exp) for v, exp in zip(vals_b, expected)):
            return row_idx
    return None


def is_asc_export_workbook(path: str | Path) -> bool:
    """Return True when the workbook looks like an aSc Timetables export."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    try:
        names = set(wb.sheetnames)
        if not all(s in names for s in _REQUIRED_SHEETS):
            return False
        lessons_hdr = _sheet_header_row(wb["Lessons"], _LESSONS_HEADER)
        teachers_hdr = _sheet_header_row(wb["Teachers"], ("Name", "Short"))
        classes_hdr = _sheet_header_row(wb["Classes"], ("Class", "Short"))
        return lessons_hdr is not None and teachers_hdr is not None and classes_hdr is not None
    finally:
        wb.close()


def _parse_length(v) -> int | None:
    s = _text(v)
    if s is None:
        return None
    low = s.lower()
    if low == "triple":
        return 6
    if low == "double":
        return 4
    if low == "single":
        return 2
    try:
        n = int(float(s))
    except ValueError:
        return None
    return n if n > 0 else None


def _parse_room_tokens(raw: str | None) -> list[str]:
    if not raw:
        return []
    parts = re.split(r"[,;]", raw)
    out: list[str] = []
    for part in parts:
        token = part.strip()
        if not token:
            continue
        if _norm(token) in _SKIP_ROOM_LABELS:
            continue
        out.append(token)
    return out


def _read_subject_codes(ws) -> list[tuple[str, str]]:
    codes: list[tuple[str, str]] = []
    hdr = _sheet_header_row(ws, ("Subject", "Short"))
    start = (hdr or 1) + 1
    for row in ws.iter_rows(min_row=start, values_only=True):
        if not row:
            continue
        title = _text(row[1] if len(row) > 1 else row[0] if row else None)
        code = _text(row[2] if len(row) > 2 else row[1] if len(row) > 1 else None)
        if not code:
            continue
        codes.append((code.strip(), (title or code).strip()))
    return codes


def _read_subjects(ws) -> dict[str, str]:
    titles: dict[str, str] = {}
    for code, title in _read_subject_codes(ws):
        key = _norm(code)
        if title and key != _norm(title):
            titles.setdefault(key, title)
        elif key not in titles:
            titles[key] = code
    return titles


def _normalize_sheet_name(name: str) -> str:
    base = name.rstrip(".").strip()
    return re.sub(r"\s+\d+$", "", base).strip()


def _map_sheet_to_subject_code(sheet_name: str, subject_codes: list[tuple[str, str]]) -> str | None:
    base_n = _norm(_normalize_sheet_name(sheet_name))
    for code, _title in sorted(subject_codes, key=lambda pair: len(pair[0]), reverse=True):
        cn = _norm(code)
        if base_n == cn or base_n.startswith(f"{cn} "):
            return code
    for code, title in subject_codes:
        cn = _norm(code)
        if cn in base_n or base_n in cn:
            return code
    for code, title in subject_codes:
        tn = _norm(title)
        if base_n in tn or tn.startswith(base_n[: min(len(base_n), 18)]):
            return code
    return None


def _read_classes(ws) -> tuple[dict[str, str], dict[str, str]]:
    names: dict[str, str] = {}
    shorts: dict[str, str] = {}
    hdr = _sheet_header_row(ws, ("Class", "Short"))
    start = (hdr or 1) + 1
    for row in ws.iter_rows(min_row=start, values_only=True):
        if not row:
            continue
        cls = _text(row[0])
        short = _text(row[1]) if len(row) > 1 else None
        if not cls:
            continue
        key = _norm(cls)
        names[key] = cls.strip()
        if short:
            shorts[key] = short.strip()
    return names, shorts


def _resolve_class(class_raw: str, class_names: dict[str, str]) -> str | None:
    key = _norm(class_raw)
    if key in class_names:
        return class_names[key]
    for k, canonical in class_names.items():
        if key in k or k in key:
            return canonical
    return class_raw.strip() or None


def _find_room(session: Session, timetable_session_id: int | None, token: str) -> Room | None:
    token = token.strip()
    if not token:
        return None
    q = _scoped_query(session, Room, timetable_session_id)
    hit = q.filter(Room.code.ilike(token)).first()
    if hit:
        return hit
    hit = q.filter(Room.name.ilike(token)).first()
    if hit:
        return hit
    if "-" in token:
        tail = token.rsplit("-", 1)[-1].strip()
        if tail:
            hit = q.filter(Room.code.ilike(tail)).first()
            if hit:
                return hit
    return None


def _ensure_staff(
    session: Session,
    timetable_session_id: int | None,
    name: str,
    rep: AscImportReport,
    cache: dict[str, Staff],
) -> Staff | None:
    key = _norm(name)
    if key in cache:
        return cache[key]
    row = _scoped_query(session, Staff, timetable_session_id).filter(Staff.name.ilike(name.strip())).first()
    if row is None:
        row = Staff(**_create_kwargs(Staff, timetable_session_id, name=name.strip()))
        session.add(row)
        session.flush()
        rep.staff_created += 1
    cache[key] = row
    return row


def _ensure_room(
    session: Session,
    timetable_session_id: int | None,
    *,
    code: str,
    name: str | None,
    rep: AscImportReport,
    cache: dict[str, Room],
) -> Room:
    key = _norm(code)
    if key in cache:
        return cache[key]
    row = _scoped_query(session, Room, timetable_session_id).filter(Room.code.ilike(code)).first()
    if row is None:
        row = Room(
            **_create_kwargs(
                Room,
                timetable_session_id,
                code=code,
                name=name or code,
            )
        )
        session.add(row)
        session.flush()
        rep.rooms_created += 1
    cache[key] = row
    return row


def _ensure_qualification(
    session: Session,
    timetable_session_id: int | None,
    *,
    name: str,
    course_code_hint: str | None,
    rep: AscImportReport,
    qual_cache: dict[str, Qualification],
    course_cache: dict[str, Course],
) -> Qualification:
    key = _norm(name)
    if key in qual_cache:
        return qual_cache[key]
    qual = _scoped_filter_by(session, Qualification, timetable_session_id, name=name).first()
    if qual is None:
        qual = Qualification(
            **_create_kwargs(
                Qualification,
                timetable_session_id,
                name=name,
                num_groups=1,
                schedule_period=SCHEDULE_PERIOD_DAY,
            )
        )
        session.add(qual)
        session.flush()
        replace_qualification_time_windows(session, qual)
        rep.qualifications_created += 1

        course_code = (course_code_hint or f"{name} GrpA").strip()
        existing_course = _scoped_filter_by(
            session, Course, timetable_session_id, code=course_code
        ).first()
        if existing_course is None:
            course = Course(
                **_create_kwargs(
                    Course,
                    timetable_session_id,
                    code=course_code,
                    name=name,
                    qualification_id=qual.id,
                )
            )
            session.add(course)
            rep.courses_created += 1
        else:
            course = existing_course
            if course.qualification_id != qual.id:
                course.qualification_id = qual.id
        session.flush()
        course_cache[key] = course
    else:
        rep.qualifications_linked += 1
        if key not in course_cache:
            course = (
                _scoped_query(session, Course, timetable_session_id)
                .filter(Course.qualification_id == qual.id)
                .order_by(Course.id)
                .first()
            )
            if course is not None:
                course_cache[key] = course
    qual_cache[key] = qual
    return qual


def _unit_storage_name(
    subject_code: str,
    subject_title: str | None,
    class_name: str,
    *,
    class_short: str | None = None,
) -> str:
    title = (subject_title or subject_code).strip()
    cohort = (class_short or class_name).strip()
    return f"{title} — {cohort}"


def _subject_keys_equivalent(a: str, b: str) -> bool:
    na, nb = _norm(a), _norm(b)
    return na == nb or na.startswith(nb) or nb.startswith(na)


def _read_lesson_assignments(wb) -> dict[tuple[str, str], _LessonAssignment]:
    """Map (normalised class, normalised subject code) -> teacher + room."""
    out: dict[tuple[str, str], _LessonAssignment] = {}

    ws = wb["Lessons"]
    hdr = _sheet_header_row(ws, _LESSONS_HEADER) or 1
    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        teacher = _text(row[0] if len(row) > 0 else None)
        class_raw = _text(row[1] if len(row) > 1 else None)
        subject = _text(row[3] if len(row) > 3 else None)
        rooms_raw = _text(row[6] if len(row) > 6 else None)
        if not class_raw or not subject:
            continue
        room_tokens = _parse_room_tokens(rooms_raw)
        out[(_norm(class_raw), _norm(subject))] = _LessonAssignment(
            teacher=teacher,
            room_token=room_tokens[0] if room_tokens else None,
        )

    if "Contracts Classes" in wb.sheetnames:
        ws_cc = wb["Contracts Classes"]
        current_class: str | None = None
        for row in ws_cc.iter_rows(min_row=2, values_only=True):
            cls_a = _text(row[0] if len(row) > 0 else None)
            teacher = _text(row[1] if len(row) > 1 else None)
            subject = _text(row[3] if len(row) > 3 else None)
            room = _text(row[8] if len(row) > 8 else None)
            if cls_a and not teacher and not subject:
                current_class = cls_a
                continue
            if not teacher or not subject or not current_class:
                continue
            key = (_norm(current_class), _norm(subject))
            prev = out.get(key)
            out[key] = _LessonAssignment(
                teacher=teacher,
                room_token=room or (prev.room_token if prev else None),
            )

    return out


def _lookup_lesson_assignment(
    assignments: dict[tuple[str, str], _LessonAssignment],
    class_name: str,
    subject_code: str,
) -> _LessonAssignment | None:
    ck = _norm(class_name)
    sk = _norm(subject_code)
    hit = assignments.get((ck, sk))
    if hit:
        return hit
    for (c, s), val in assignments.items():
        if c == ck and _subject_keys_equivalent(s, sk):
            return val
    return None


def _is_subject_schedule_sheet(ws) -> bool:
    return _sheet_header_row(ws, _DAY_LESSON_HEADER, max_scan=12) is not None


def _extract_blocks_from_subject_sheet(ws, *, subject_code: str, sheet_name: str) -> list[_ScheduleBlock]:
    hdr = _sheet_header_row(ws, _DAY_LESSON_HEADER, max_scan=12) or 3
    blocks: list[_ScheduleBlock] = []
    current: _ScheduleBlock | None = None

    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        day = _text(row[1] if len(row) > 1 else None)
        lesson_raw = row[2] if len(row) > 2 else None
        cls = _text(row[3] if len(row) > 3 else None)
        if not day or lesson_raw is None or not cls:
            if current is not None:
                blocks.append(current)
                current = None
            continue
        try:
            period = int(float(str(lesson_raw).strip()))
        except ValueError:
            continue
        if period <= 0:
            continue

        if (
            current is not None
            and current.day == day
            and _norm(current.class_name) == _norm(cls)
            and period == current.end_period + 1
        ):
            current.end_period = period
            continue

        if current is not None:
            blocks.append(current)
        current = _ScheduleBlock(
            subject_code=subject_code,
            class_name=cls.strip(),
            day=day,
            start_period=period,
            end_period=period,
            sheet_name=sheet_name,
        )

    if current is not None:
        blocks.append(current)
    return blocks


def _extract_all_schedule_blocks(
    wb,
    subject_codes: list[tuple[str, str]],
) -> list[_ScheduleBlock]:
    blocks: list[_ScheduleBlock] = []
    for sheet_name in wb.sheetnames:
        if sheet_name in _CORE_SHEETS:
            continue
        code = _map_sheet_to_subject_code(sheet_name, subject_codes)
        if not code:
            continue
        ws = wb[sheet_name]
        if not _is_subject_schedule_sheet(ws):
            continue
        blocks.extend(
            _extract_blocks_from_subject_sheet(ws, subject_code=code, sheet_name=sheet_name)
        )
    return blocks


def _asc_periods_to_slots(start_period: int, end_period: int) -> tuple[int, int] | None:
    start_slot = start_period - 1
    end_slot = end_period
    if start_slot < 0 or start_slot >= NUM_SLOTS:
        return None
    end_slot = min(end_slot, NUM_SLOTS)
    if end_slot <= start_slot:
        return None
    return start_slot, end_slot


def _day_index(day_name: str) -> int | None:
    key = _norm(day_name)
    for i, d in enumerate(DAYS):
        if _norm(d) == key:
            return i
    return None


def _ensure_course_unit(session: Session, course_id: int, unit_id: int, linked: set[tuple[int, int]]) -> None:
    key = (course_id, unit_id)
    if key in linked:
        return
    exists = session.query(CourseUnit).filter_by(course_id=course_id, unit_id=unit_id).first()
    if not exists:
        session.add(CourseUnit(course_id=course_id, unit_id=unit_id))
    linked.add(key)


def _import_bookings(
    session: Session,
    *,
    timetable_session_id: int | None,
    blocks: list[_ScheduleBlock],
    class_names: dict[str, str],
    class_shorts: dict[str, str],
    subject_titles: dict[str, str],
    assignments: dict[tuple[str, str], _LessonAssignment],
    qual_cache: dict[str, Qualification],
    course_cache: dict[str, Course],
    unit_cache: dict[tuple[str, str], Unit],
    staff_cache: dict[str, Staff],
    room_cache: dict[str, Room],
    rep: AscImportReport,
) -> None:
    if not blocks:
        return

    ctx = VisualImportContext(timetable_session_id)
    week = ctx.first_week(session)
    session.query(Booking).filter(Booking.week_id == week.id).delete()

    course_unit_linked: set[tuple[int, int]] = set()

    for block in blocks:
        class_name = _resolve_class(block.class_name, class_names)
        if not class_name:
            rep.warnings.append(
                f"{block.sheet_name}: unknown class {block.class_name!r} on {block.day} "
                f"period {block.start_period}"
            )
            continue

        qual_key = _norm(class_name)
        qual = qual_cache.get(qual_key)
        course = course_cache.get(qual_key)
        if qual is None or course is None:
            rep.warnings.append(
                f"{block.sheet_name}: no course for class {class_name!r} ({block.subject_code})"
            )
            continue

        subj_key = _norm(block.subject_code)
        unit_key = (qual_key, subj_key)
        unit = unit_cache.get(unit_key)
        if unit is None:
            storage_name = _unit_storage_name(
                block.subject_code,
                subject_titles.get(subj_key),
                class_name,
                class_short=class_shorts.get(qual_key),
            )
            unit = _scoped_filter_by(session, Unit, timetable_session_id, name=storage_name).first()
        if unit is None:
            rep.warnings.append(
                f"{block.sheet_name}: no class/unit for {block.subject_code!r} / {class_name!r}"
            )
            continue

        slot_range = _asc_periods_to_slots(block.start_period, block.end_period)
        day_idx = _day_index(block.day)
        if slot_range is None or day_idx is None:
            rep.warnings.append(
                f"{block.sheet_name}: invalid day/slot for {class_name!r} on {block.day} "
                f"periods {block.start_period}-{block.end_period}"
            )
            continue
        start_slot, end_slot = slot_range

        assign = _lookup_lesson_assignment(assignments, class_name, block.subject_code)
        staff_id = None
        room_id = None
        if assign:
            if assign.teacher:
                staff = _ensure_staff(session, timetable_session_id, assign.teacher, rep, staff_cache)
                staff_id = staff.id if staff else None
            if assign.room_token:
                room = _find_room(session, timetable_session_id, assign.room_token)
                if room is None:
                    rep.warnings.append(
                        f"{block.sheet_name}: room {assign.room_token!r} not found for "
                        f"{block.subject_code!r}/{class_name!r}"
                    )
                else:
                    room_id = room.id

        _ensure_course_unit(session, course.id, unit.id, course_unit_linked)
        session.add(
            Booking(
                week_id=week.id,
                course_id=course.id,
                unit_id=unit.id,
                staff_id=staff_id,
                room_id=room_id,
                day=day_idx,
                start_slot=start_slot,
                end_slot=end_slot,
            )
        )
        rep.bookings_created += 1


def is_asc_export_file(path: str | Path) -> bool:
    """Return True when the file looks like an aSc export (.xlsx or 2012 XML)."""
    from .asc_xml_import import is_asc_export_xml

    return is_asc_export_xml(path) or is_asc_export_workbook(path)


def import_asc_export(
    session: Session,
    path: str | Path,
    *,
    timetable_session_id: int | None = None,
) -> AscImportReport:
    from .asc_xml_import import import_asc_xml_export, is_asc_export_xml

    if is_asc_export_xml(path):
        return import_asc_xml_export(session, path, timetable_session_id=timetable_session_id)

    if not is_asc_export_workbook(path):
        raise ValueError(
            "File does not look like an aSc Timetables export "
            "(expected .xlsx with Teachers/Classrooms/Classes/Lessons sheets, "
            "or aSc 2012 XML with teachers, lessons, and cards)."
        )

    rep = AscImportReport()
    wb = load_workbook(path, data_only=True)

    staff_cache: dict[str, Staff] = {}
    room_cache: dict[str, Room] = {}
    qual_cache: dict[str, Qualification] = {}
    course_cache: dict[str, Course] = {}
    unit_cache: dict[tuple[str, str], Unit] = {}

    ws_teachers = wb["Teachers"]
    t_hdr = _sheet_header_row(ws_teachers, ("Name", "Short")) or 1
    for row in ws_teachers.iter_rows(min_row=t_hdr + 1, values_only=True):
        name = _text(row[1] if len(row) > 1 else row[0] if row else None)
        if not name:
            continue
        _ensure_staff(session, timetable_session_id, name, rep, staff_cache)

    ws_rooms = wb["Classrooms"]
    r_hdr = _sheet_header_row(ws_rooms, ("Name", "Short")) or 1
    for row in ws_rooms.iter_rows(min_row=r_hdr + 1, values_only=True):
        full_name = _text(row[1] if len(row) > 1 else None)
        code = _text(row[2] if len(row) > 2 else None) or full_name
        if not code:
            continue
        _ensure_room(
            session,
            timetable_session_id,
            code=code,
            name=full_name,
            rep=rep,
            cache=room_cache,
        )

    class_names, class_shorts = _read_classes(wb["Classes"])
    subject_codes = _read_subject_codes(wb["Subjects"]) if "Subjects" in wb.sheetnames else []
    subject_titles = _read_subjects(wb["Subjects"]) if "Subjects" in wb.sheetnames else {}

    for key, canonical in class_names.items():
        _ensure_qualification(
            session,
            timetable_session_id,
            name=canonical,
            course_code_hint=class_shorts.get(key),
            rep=rep,
            qual_cache=qual_cache,
            course_cache=course_cache,
        )

    ws_lessons = wb["Lessons"]
    l_hdr = _sheet_header_row(ws_lessons, _LESSONS_HEADER) or 1
    for row in ws_lessons.iter_rows(min_row=l_hdr + 1, values_only=True):
        teacher = _text(row[0] if len(row) > 0 else None)
        class_raw = _text(row[1] if len(row) > 1 else None)
        subject_code = _text(row[3] if len(row) > 3 else None)
        length_raw = row[4] if len(row) > 4 else None
        rooms_raw = _text(row[6] if len(row) > 6 else None)

        if not class_raw or not subject_code:
            continue

        class_name = _resolve_class(class_raw, class_names)
        if class_name is None:
            rep.warnings.append(f"Lessons: unknown class {class_raw!r}")
            continue

        qual_key = _norm(class_name)
        qual = qual_cache.get(qual_key)
        if qual is None:
            qual = _ensure_qualification(
                session,
                timetable_session_id,
                name=class_name,
                course_code_hint=class_shorts.get(qual_key),
                rep=rep,
                qual_cache=qual_cache,
                course_cache=course_cache,
            )

        subj_key = _norm(subject_code)
        unit_key = (qual_key, subj_key)
        storage_name = _unit_storage_name(
            subject_code,
            subject_titles.get(subj_key),
            class_name,
            class_short=class_shorts.get(qual_key),
        )

        unit = unit_cache.get(unit_key)
        if unit is None:
            unit = _scoped_filter_by(session, Unit, timetable_session_id, name=storage_name).first()
            if unit is None:
                unit = Unit(
                    **_create_kwargs(
                        Unit,
                        timetable_session_id,
                        name=storage_name,
                        component_codes=subject_code,
                    )
                )
                session.add(unit)
                session.flush()
                rep.classes_created += 1
            else:
                rep.classes_updated += 1
            unit_cache[unit_key] = unit
        else:
            rep.classes_updated += 1

        if not unit.component_codes:
            unit.component_codes = subject_code
        length = _parse_length(length_raw)
        if length and not unit.length_slots:
            unit.length_slots = length

        if not (
            session.query(UnitQualification)
            .filter_by(unit_id=unit.id, qualification_id=qual.id)
            .first()
        ):
            session.add(UnitQualification(unit_id=unit.id, qualification_id=qual.id))
            rep.class_qual_links_added += 1

        if teacher:
            staff = _ensure_staff(session, timetable_session_id, teacher, rep, staff_cache)
            if staff and not (
                session.query(StaffCompetency)
                .filter_by(staff_id=staff.id, unit_id=unit.id)
                .first()
            ):
                session.add(StaffCompetency(staff_id=staff.id, unit_id=unit.id))
                rep.lecturer_links_added += 1

        for token in _parse_room_tokens(rooms_raw):
            room = _find_room(session, timetable_session_id, token)
            if room is None:
                rep.warnings.append(
                    f"Lessons: room {token!r} (class {class_name!r}, subject {subject_code!r}) not found"
                )
                continue
            if not (
                session.query(UnitAllowedRoom)
                .filter_by(unit_id=unit.id, room_id=room.id)
                .first()
            ):
                session.add(UnitAllowedRoom(unit_id=unit.id, room_id=room.id))
                rep.room_links_added += 1

    assignments = _read_lesson_assignments(wb)
    schedule_blocks = _extract_all_schedule_blocks(wb, subject_codes)
    _import_bookings(
        session,
        timetable_session_id=timetable_session_id,
        blocks=schedule_blocks,
        class_names=class_names,
        class_shorts=class_shorts,
        subject_titles=subject_titles,
        assignments=assignments,
        qual_cache=qual_cache,
        course_cache=course_cache,
        unit_cache=unit_cache,
        staff_cache=staff_cache,
        room_cache=room_cache,
        rep=rep,
    )

    apply_unit_bracket_fields_from_names(session, timetable_session_id=timetable_session_id)
    session.commit()
    return rep
