"""Export the database back to an Overall-style .xlsm.

`write_into_template(template, out)` is the primary path: it clones an existing
input workbook (preserving its design, formatting, and macros), strips out any
macro-generated sheets, then repaints the Overall sheet's booking body from
the database without disturbing headers, time grid, day banners, fonts,
borders, or merged cells.

`write_fresh(out)` is a no-template fallback that emits a plain .xlsx with the
same logical layout but no styling.
"""
from __future__ import annotations

import colorsys
import hashlib
import shutil
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from ..bundle_paths import resource_path
from ..constants import NUM_DAYS, NUM_SLOTS, slot_to_time
from ..core.booking_staff import sfs_co_teacher_term_labels, staff_active_in_term
from ..core.class_colour import booking_colour_key
from ..core.models import Booking, Course, Room, Staff, Week
from ..core.room_types import room_is_physical
from .backup_payload import BACKUP_SHEET_NAME, PAYLOAD_VERSION, write_backup_sheet
# Kept in sync with import/grid layout constants.
DAY_BLOCK_STRIDE = 29
FIRST_DATA_ROW = 3
SECTION_FACILITATION = "Facilitation"
SECTION_STAFF = "Staff!"
SECTION_ROOMS = "Rooms!"


KEPT_SHEETS = {"Overall", "Course Template", "Staff Template", "Room Template"}

def _write_backup_sheet(
    wb, session: Session, *, timetable_session_id: int | None = None
) -> None:
    """Embed a hidden sheet carrying the full structured session payload."""
    write_backup_sheet(wb, session, timetable_session_id=timetable_session_id)


def _course_fill_hex(key: str) -> str:
    """Stable fill hex for a course/class key (matches the in-app palette)."""
    from ..core.screen_colours import screen_fill_xlsx

    return screen_fill_xlsx(key)


def _course_border_hex(key: str) -> str:
    """Border hex for a course/class key (matches the in-app palette)."""
    from ..core.screen_colours import screen_border_xlsx

    return screen_border_xlsx(key)


@dataclass
class ExportReport:
    out_path: Path
    courses_written: int = 0
    courses_skipped: list[str] = field(default_factory=list)
    sheets_removed: list[str] = field(default_factory=list)
    bookings: int = 0
    course_tabs: int = 0
    staff_tabs: int = 0
    room_tabs: int = 0


# ---- Per-entity tab layout constants (mirror the source spreadsheet) ----
ENTITY_FIRST_SLOT_ROW = 7      # row of slot 0 (08:00) on each per-entity tab
ENTITY_DAY_FIRST_COL = 3       # col of Monday's first sub-column (C)
# Per day: unit + room + lecturer + narrow gap column (matches Overall 3-subcol block).
ENTITY_DAY_STRIDE = 4          # stride between Mondays (cols 3–6, next day 7, …)
ENTITY_TITLE_CELL = (1, 5)     # E1 — entity name

# Classic 3-column (Subject | Room | Lecturer) entity tab layout for export.
CLASS_LECTURER_ENTITY_TEMPLATE = resource_path("class_lecturer_term_template.xlsx")


def _copy_worksheet_layout(src, dest) -> None:
    """Copy grid structure and styles from one worksheet to another (cross-workbook)."""
    for row in src.iter_rows():
        for cell in row:
            dest_cell = dest.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dest_cell.font = copy(cell.font)
                dest_cell.border = copy(cell.border)
                dest_cell.fill = copy(cell.fill)
                dest_cell.number_format = copy(cell.number_format)
                dest_cell.protection = copy(cell.protection)
                dest_cell.alignment = copy(cell.alignment)
    for merged in src.merged_cells.ranges:
        dest.merge_cells(str(merged))
    dest.freeze_panes = src.freeze_panes
    for col_letter, dim in src.column_dimensions.items():
        if dim.width is not None:
            dest.column_dimensions[col_letter].width = dim.width
    for row_idx, dim in src.row_dimensions.items():
        if dim.height is not None:
            dest.row_dimensions[row_idx].height = dim.height


def _ensure_class_lecturer_entity_templates(wb) -> None:
    """Replace Course/Staff template sheets with the repo 3-column layout."""
    path = CLASS_LECTURER_ENTITY_TEMPLATE
    if not path.is_file():
        return
    src_wb = load_workbook(path, read_only=False)
    try:
        for sheet_name in ("Course Template", "Staff Template"):
            if sheet_name not in src_wb.sheetnames:
                continue
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            dest = wb.create_sheet(sheet_name)
            _copy_worksheet_layout(src_wb[sheet_name], dest)
    finally:
        src_wb.close()


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Excel sheet names: ≤31 chars, no [\\:/?*[]] characters, must be unique."""
    bad = set("[]:*?/\\")
    cleaned = "".join("_" if ch in bad else ch for ch in str(name)).strip()
    cleaned = cleaned[:31] or "Sheet"
    base = cleaned
    n = 1
    while cleaned in used:
        n += 1
        suffix = f" ({n})"
        cleaned = (base[: 31 - len(suffix)] + suffix).strip()
    used.add(cleaned)
    return cleaned


def _terms_of(b: Booking) -> tuple[bool, bool]:
    return bool(getattr(b, "in_term_1", 1)), bool(getattr(b, "in_term_2", 1))


def _week_bookings(session: Session, week_id: int) -> list[Booking]:
    """Bookings for export with relationships loaded."""
    return (
        session.query(Booking)
        .options(
            joinedload(Booking.unit),
            joinedload(Booking.course),
            joinedload(Booking.staff),
            joinedload(Booking.sfs_co_teacher),
            joinedload(Booking.room),
        )
        .filter(Booking.week_id == week_id)
        .all()
    )


def _class_title_line(b: Booking) -> str:
    """Class display title (name, booking id, double-session part) without units-of-study codes."""
    name = (b.unit.name if b.unit else "") or ""
    ext = (b.external_id or "").strip()
    if ext:
        name = f"[{ext}] {name}".strip()
    part = getattr(b, "session_part", 1) or 1
    if part > 1:
        name = f"{name} ({part}/2)".strip()
    elif b.unit and getattr(b.unit, "double_session", 0):
        name = f"{name} (1/2)".strip()
    return name


def _component_codes_line(b: Booking) -> str:
    """Comma-separated units-of-study codes for the class, if any."""
    if b.unit is None:
        return ""
    from ..core.unit_brackets import normalize_component_codes_commas

    raw = (b.unit.component_codes or "").strip()
    if not raw:
        return ""
    return (normalize_component_codes_commas(raw) or raw).strip()


def placecard_subject_block(b: Booking, *, prefix: str = "", term: str | None = None) -> str:
    """Placecard class block: title, codes, notes, and optional partial-week label."""
    from ..core.block_delivery import block_booking_label, is_block_booking
    from ..core.booking_sessions import format_session_weeks_label

    parts: list[str] = []
    title = _class_title_line(b)
    if prefix:
        title = f"{prefix}{title}"
    if title:
        parts.append(title)
    codes = _component_codes_line(b)
    if codes:
        parts.append(codes)
    notes = (b.notes or "").strip()
    if notes:
        parts.append(notes)
    if is_block_booking(b):
        blk = block_booking_label(b)
        if blk:
            parts.append(blk)
    else:
        week_label = format_session_weeks_label(b, term=term)
        if week_label:
            parts.append(week_label)
    return "\n".join(parts)


def _unit_label(b: Booking, *, term: str | None = None) -> str:
    """Class line for v1 workbook cards (title, notes, optional session weeks)."""
    from ..core.block_delivery import block_booking_label, is_block_booking
    from ..core.booking_sessions import format_session_weeks_label

    parts: list[str] = []
    name = _class_title_line(b)
    if name:
        parts.append(name)
    notes = (b.notes or "").strip()
    if notes:
        parts.append(notes)
    if is_block_booking(b):
        blk = block_booking_label(b)
        if blk:
            parts.append(blk)
    else:
        week_label = format_session_weeks_label(b, term=term)
        if week_label:
            parts.append(week_label)
    return "\n".join(parts)


def _lecturer_label(b: Booking) -> str:
    """Primary and SFS co-teacher names for export placecards."""
    from ..core.booking_staff import export_lecturer_label

    return export_lecturer_label(b)


def _bookings_for_staff_tab(bookings: list[Booking], staff_id: int) -> list[Booking]:
    """Bookings shown on a staff tab (primary role or term-scoped SFS co-teach)."""
    out: list[Booking] = []
    for b in bookings:
        if b.staff_id == staff_id:
            out.append(b)
        elif getattr(b, "sfs_co_teacher_staff_id", None) == staff_id:
            if staff_active_in_term(b, staff_id, "t1") or staff_active_in_term(
                b, staff_id, "t2"
            ):
                out.append(b)
    return out


def _split_pairs(bookings: list[Booking]) -> tuple[list[Booking], list[tuple[Booking, Booking]]]:
    """Split a list of bookings into:
      - solo: bookings that don't share their (day, start, end) with another
      - pairs: (T1-only, T2-only) pairs sharing the same (day, start, end)
    Pairs render side by side at full slot height (T1 left column, T2 right column).
    Both-term bookings use all three columns; single-term bookings use one outer column.
    """
    by_slot: dict[tuple[int, int, int], list[Booking]] = {}
    for b in bookings:
        by_slot.setdefault((b.day, b.start_slot, b.end_slot), []).append(b)
    solo: list[Booking] = []
    pairs: list[tuple[Booking, Booking]] = []
    for slot_key, group in by_slot.items():
        if len(group) == 2:
            t1_only = [b for b in group if _terms_of(b) == (True, False)]
            t2_only = [b for b in group if _terms_of(b) == (False, True)]
            if len(t1_only) == 1 and len(t2_only) == 1:
                pairs.append((t1_only[0], t2_only[0]))
                continue
        solo.extend(group)
    return solo, pairs


def _bookings_by_slot(bookings: list[Booking]) -> list[list[Booking]]:
    """Group bookings that share the same day and slot span."""
    from collections import defaultdict

    by_slot: dict[tuple[int, int, int], list[Booking]] = defaultdict(list)
    for b in bookings:
        by_slot[(b.day, b.start_slot, b.end_slot)].append(b)
    return [
        sorted(g, key=lambda b: ((b.course.code if b.course else ""), b.id))
        for g in by_slot.values()
    ]


def _staff_lane_subgroups(
    group: list[Booking],
    view_staff_id: int,
) -> list[tuple[tuple[bool, bool], list[Booking]]]:
    """Split a same-slot staff group into T1-only, T2-only, and semester subgroups.

    Uses the viewed lecturer's role (primary class terms vs SFS co-teach terms).
    """
    from ..core.booking_staff import lane_terms_for_staff_view

    t1_only: list[Booking] = []
    t2_only: list[Booking] = []
    both: list[Booking] = []
    for b in group:
        t1, t2 = lane_terms_for_staff_view(b, view_staff_id)
        if not t1 and not t2:
            continue
        if t1 and t2:
            both.append(b)
        elif t1:
            t1_only.append(b)
        elif t2:
            t2_only.append(b)
    lanes: list[tuple[tuple[bool, bool], list[Booking]]] = []
    if t1_only:
        lanes.append(((True, False), t1_only))
    if t2_only:
        lanes.append(((False, True), t2_only))
    if both:
        lanes.append(((True, True), both))
    return lanes


def _staff_placecard_lines(
    group: list[Booking],
    *,
    view_staff_id: int | None = None,
    colour_by_class: bool = True,
    term: str | None = None,
    subject_block=placecard_subject_block,
) -> tuple[tuple[str, str, str], str]:
    """Build staff placecard text; list every cohort (group) in the same timeslot."""

    def _staff_subject(b: Booking) -> str:
        block = subject_block(b, term=term)
        if (
            view_staff_id is not None
            and getattr(b, "sfs_co_teacher_staff_id", None) == view_staff_id
            and b.staff_id != view_staff_id
        ):
            terms = sfs_co_teacher_term_labels(b)
            badge = f"[SFS co-teach {terms}] " if terms else "[SFS co-teach] "
            return subject_block(b, prefix=badge, term=term)
        return block

    if len(group) == 1:
        b = group[0]
        from ..core.booking_staff import export_lecturer_label, has_sfs_co_teacher

        third = (b.course.code if b.course else "") or ""
        if has_sfs_co_teacher(b):
            third = export_lecturer_label(b)
        return (
            _staff_subject(b),
            (b.room.code if b.room else "") or "",
            third,
        ), booking_colour_key(b, by_class=colour_by_class)

    units: list[str] = []
    rooms: list[str] = []
    groups: list[str] = []
    seen_unit: set[str] = set()
    seen_room: set[str] = set()
    seen_group: set[str] = set()
    for b in group:
        subj = _staff_subject(b)
        v1 = (b.room.code if b.room else "") or ""
        v2 = (b.course.code if b.course else "") or ""
        if subj and subj not in seen_unit:
            seen_unit.add(subj)
            units.append(subj)
        if v1 and v1 not in seen_room:
            seen_room.add(v1)
            rooms.append(v1)
        if v2 and v2 not in seen_group:
            seen_group.add(v2)
            groups.append(v2)
    tint = booking_colour_key(group[0], by_class=colour_by_class)
    return (
        "\n".join(units),
        ", ".join(rooms),
        ", ".join(groups),
    ), tint


def _entity_values(
    b: Booking,
    kind: str,
    *,
    view_staff_id: int | None = None,
    colour_by_class: bool = True,
) -> tuple[str, str, str, str]:
    """Returns (v0, v1, v2, tint_key): subject line(s), secondary cells, tint key."""
    tint_key = booking_colour_key(b, by_class=colour_by_class)
    unit = _unit_label(b)
    if kind == "course":
        return (
            unit,
            (b.room.code if b.room else "") or "",
            _lecturer_label(b),
            tint_key,
        )
    if kind == "staff":
        if (
            view_staff_id is not None
            and getattr(b, "sfs_co_teacher_staff_id", None) == view_staff_id
            and b.staff_id != view_staff_id
        ):
            terms = sfs_co_teacher_term_labels(b)
            badge = f"[SFS co-teach {terms}] " if terms else "[SFS co-teach] "
            unit = badge + unit
        return (
            unit,
            (b.room.code if b.room else "") or "",
            (b.course.code if b.course else "") or "",
            tint_key,
        )
    # room
    return (
        unit,
        _lecturer_label(b),
        (b.course.code if b.course else "") or "",
        tint_key,
    )


def _render_entity_tab(
    ws,
    title: str,
    bookings: list[Booking],
    kind: str,
    *,
    view_staff_id: int | None = None,
    colour_by_class: bool = True,
) -> None:
    """Populate a per-entity sheet (course/staff/room) with card-styled bookings.

    Staff tabs use T1/T2 lane columns for term-specific classes; semester classes
    use the full width. Concurrent same-slot cohorts are merged per lane.
    """
    title_cell = ws.cell(row=ENTITY_TITLE_CELL[0], column=ENTITY_TITLE_CELL[1], value=title)
    title_cell.font = Font(name="Tahoma", size=20, bold=True)

    slot_groups = _bookings_by_slot(bookings) if kind == "staff" else [[b] for b in bookings]
    for group in slot_groups:
        b0 = group[0]
        day_col = ENTITY_DAY_FIRST_COL + b0.day * ENTITY_DAY_STRIDE
        top_row = ENTITY_FIRST_SLOT_ROW + b0.start_slot
        bot_row = ENTITY_FIRST_SLOT_ROW + b0.end_slot - 1
        if kind == "staff":
            assert view_staff_id is not None
            for term_flags, lane_bookings in _staff_lane_subgroups(group, view_staff_id):
                t1, t2 = term_flags
                lane_term = "t1" if t1 and not t2 else ("t2" if t2 and not t1 else None)
                (v0, v1, v2), course_code = _staff_placecard_lines(
                    lane_bookings,
                    view_staff_id=view_staff_id,
                    colour_by_class=colour_by_class,
                    term=lane_term,
                )
                if t1 and t2:
                    _draw_card(ws, top_row, bot_row, day_col, v0, v1, v2, course_code)
                elif t1:
                    _draw_single_term_card(
                        ws, top_row, bot_row, day_col, v0, v1, v2, course_code, term="T1"
                    )
                elif t2:
                    _draw_single_term_card(
                        ws, top_row, bot_row, day_col, v0, v1, v2, course_code, term="T2"
                    )
        else:
            v0, v1, v2, course_code = _entity_values(
                b0, kind, view_staff_id=view_staff_id, colour_by_class=colour_by_class
            )
            _draw_card(ws, top_row, bot_row, day_col, v0, v1, v2, course_code)


def _day_first_row(day: int) -> int:
    return day * DAY_BLOCK_STRIDE + FIRST_DATA_ROW


def _resolve_week(session: Session, week_id: int | None) -> Week:
    if week_id:
        w = session.get(Week, week_id)
        if w is None:
            raise ValueError(f"week id {week_id} not found")
        return w
    w = session.query(Week).order_by(Week.semester_id, Week.week_number).first()
    if w is None:
        raise ValueError("no weeks in database")
    return w


def _cell_text(c: Cell) -> str | None:
    v = c.value
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _scan_row1_markers(ws) -> dict[str, int]:
    found: dict[str, int] = {}
    for cell in ws[1]:
        v = _cell_text(cell)
        if v in (SECTION_FACILITATION, SECTION_STAFF, SECTION_ROOMS):
            found[v] = cell.column
    missing = {SECTION_FACILITATION, SECTION_STAFF, SECTION_ROOMS} - found.keys()
    if missing:
        raise ValueError(f"Overall sheet missing section markers: {sorted(missing)}")
    return found


def _course_columns_in_template(ws, markers: dict[str, int]) -> dict[str, int]:
    """Map course code → 1-based unit column from the template's row 1."""
    out: dict[str, int] = {}
    col = markers[SECTION_FACILITATION]
    end = markers[SECTION_STAFF]
    while col < end:
        name = _cell_text(ws.cell(row=1, column=col))
        if name and name not in (SECTION_FACILITATION, SECTION_STAFF, SECTION_ROOMS):
            out[name] = col
        # The Facilitation marker IS the first course's name in the source
        # convention, so it gets recorded above too. Step in 3-col groups.
        col += 3
    # Special-case: Facilitation marker AND course share row 1 col E. Make
    # sure 'Facilitation' lands in the map even if the literal scan missed it.
    fac_col = markers[SECTION_FACILITATION]
    if SECTION_FACILITATION not in out:
        out[SECTION_FACILITATION] = fac_col
    return out


def _reset_course_body_cell(cell) -> None:
    """Remove booking text and card styling from one Overall body cell."""
    cell.value = None
    cell.fill = PatternFill()
    cell.border = Border()


def _clear_course_body(ws, course_col: int) -> None:
    """Clear unit/room/lecturer cells in a course's 3-col block across all day blocks.

    Wipes values and card fill/borders so removed or moved bookings do not leave
    empty coloured boxes in the export.
    """
    for day in range(NUM_DAYS):
        first = _day_first_row(day)
        rows = list(range(first, first + NUM_SLOTS))
        if first > FIRST_DATA_ROW:
            rows.insert(0, first - 1)
        for row in rows:
            for off in (0, 1, 2):
                _reset_course_body_cell(ws.cell(row=row, column=course_col + off))


def _draw_card(
    ws,
    top_row: int,
    bot_row: int,
    base_col: int,
    v0: str,
    v1: str,
    v2: str,
    course_code: str,
    *,
    term_badge: str | None = None,
    col_span: int = 3,
    col_offset: int = 0,
) -> None:
    """Draw one coloured card from `top_row` to `bot_row`.

    Default ``col_span=3`` uses unit / room / lecturer columns. Single-term
    bookings use ``col_span=1`` in the left (T1) or right (T2) column so the
    card spans the full timeslot height (horizontal split across the block).
    """
    start_col = base_col + col_offset
    fill = PatternFill(
        start_color=_course_fill_hex(course_code),
        end_color=_course_fill_hex(course_code),
        fill_type="solid",
    )
    border_side = Side(border_style="medium", color=_course_border_hex(course_code))
    border = Border(top=border_side, bottom=border_side, left=border_side, right=border_side)
    align_top = Alignment(horizontal="center", vertical="top", wrap_text=True)
    align_mid = Alignment(horizontal="center", vertical="center", wrap_text=True)
    unit_font = Font(name="Calibri", size=10, bold=True, color="FF1A1A1A")
    sub_font = Font(name="Calibri", size=10, color="FF1A1A1A")
    if term_badge:
        v0 = f"[{term_badge}] {v0}".strip()

    if any(
        isinstance(ws.cell(row=r, column=start_col + off), MergedCell)
        for r in range(top_row, bot_row + 1)
        for off in range(col_span)
    ):
        for r in range(top_row, FIRST_DATA_ROW - 1, -1):
            anchor = ws.cell(row=r, column=start_col)
            if not isinstance(anchor, MergedCell):
                suffix = f" ⚠ + {v0 or '?'}"
                cur = anchor.value or ""
                if suffix.strip() not in (cur or ""):
                    anchor.value = (cur + suffix).strip()
                break
        return

    if col_span == 1:
        lines = [x for x in (v0, v1, v2) if x]
        text = "\n".join(lines)
        c0 = ws.cell(row=top_row, column=start_col, value=text)
        c0.fill = fill
        c0.font = unit_font
        c0.border = border
        c0.alignment = align_top
        for r in range(top_row + 1, bot_row + 1):
            ws.cell(row=r, column=start_col).fill = fill
        if bot_row > top_row:
            ws.merge_cells(
                start_row=top_row,
                end_row=bot_row,
                start_column=start_col,
                end_column=start_col,
            )
        return

    c0 = ws.cell(row=top_row, column=start_col, value=v0)
    c1 = ws.cell(row=top_row, column=start_col + 1, value=v1)
    c2 = ws.cell(row=top_row, column=start_col + 2, value=v2)
    c0.fill = fill
    c0.font = unit_font
    c0.border = border
    c0.alignment = align_top
    c1.fill = fill
    c1.font = sub_font
    c1.border = border
    c1.alignment = align_mid
    c2.fill = fill
    c2.font = sub_font
    c2.border = border
    c2.alignment = align_mid
    for r in range(top_row + 1, bot_row + 1):
        for off in range(3):
            ws.cell(row=r, column=start_col + off).fill = fill
    if bot_row > top_row:
        for off in range(3):
            ws.merge_cells(
                start_row=top_row,
                end_row=bot_row,
                start_column=start_col + off,
                end_column=start_col + off,
            )


def _draw_single_term_card(
    ws,
    top_row: int,
    bot_row: int,
    base_col: int,
    v0: str,
    v1: str,
    v2: str,
    course_code: str,
    *,
    term: str,
) -> None:
    """Single-term booking at full slot height in the outer third of the 3-col block."""
    col_offset = 0 if term == "T1" else 2
    _draw_card(
        ws,
        top_row,
        bot_row,
        base_col,
        v0,
        v1,
        v2,
        course_code,
        term_badge=term,
        col_span=1,
        col_offset=col_offset,
    )


def _draw_term_split_or_full(
    ws, b: Booking, base_col: int, course_code: str,
    v0: str, v1: str, v2: str,
) -> None:
    """For a single (un-paired) booking: full 3-col width or left/right single-term column."""
    first = _day_first_row(b.day)
    top_row = first + b.start_slot
    bot_row = first + b.end_slot - 1
    t1, t2 = _terms_of(b)
    if t1 and t2:
        _draw_card(ws, top_row, bot_row, base_col, v0, v1, v2, course_code)
    elif t1:
        _draw_single_term_card(
            ws, top_row, bot_row, base_col, v0, v1, v2, course_code, term="T1"
        )
    elif t2:
        _draw_single_term_card(
            ws, top_row, bot_row, base_col, v0, v1, v2, course_code, term="T2"
        )


def _write_course_body(ws, course_col: int, bookings: list[Booking], course_code: str) -> None:
    """Render bookings into the course's 3-col block (subject / room / lecturer).

    Term-scoped and semester classes use the same full-width card layout.
    """
    for b in bookings:
        first = _day_first_row(b.day)
        top_row = first + b.start_slot
        bot_row = first + b.end_slot - 1
        cc = booking_colour_key(b, by_class=True)
        _draw_card(
            ws,
            top_row,
            bot_row,
            course_col,
            _unit_label(b),
            (b.room.code if b.room else "") or "",
            _lecturer_label(b),
            cc,
        )


def _unmerge_course_body(ws, markers: dict[str, int]) -> None:
    """Unmerge any merged ranges that sit inside the course body grid.

    The source file occasionally merges a course's unit cells visually for
    pretty multi-slot blocks (e.g. H4:H7). Those merges block per-cell writes,
    so flatten them before repainting.
    """
    fac_col = markers[SECTION_FACILITATION]
    staff_col = markers[SECTION_STAFF]
    for r in list(ws.merged_cells.ranges):
        if r.min_row >= FIRST_DATA_ROW and fac_col <= r.min_col < staff_col:
            ws.unmerge_cells(str(r))


def _rewrite_overall(
    ws,
    session: Session,
    week_id: int,
    report: ExportReport,
    *,
    timetable_session_id: int | None = None,
) -> None:
    markers = _scan_row1_markers(ws)
    _unmerge_course_body(ws, markers)
    template_courses = _course_columns_in_template(ws, markers)
    course_q = session.query(Course).order_by(Course.id)
    if timetable_session_id is not None:
        course_q = course_q.filter(Course.timetable_session_id == timetable_session_id)
    db_courses = {c.code: c for c in course_q.all()}

    # Clear body for every template course column, then repaint from DB.
    for code, col in template_courses.items():
        _clear_course_body(ws, col)

    bookings = _week_bookings(session, week_id)
    by_course: dict[int, list[Booking]] = {}
    for b in bookings:
        by_course.setdefault(b.course_id, []).append(b)

    for code, course in db_courses.items():
        col = template_courses.get(code)
        if col is None:
            report.courses_skipped.append(code)
            continue
        _write_course_body(ws, col, by_course.get(course.id, []), code)
        report.courses_written += 1
        report.bookings += len(by_course.get(course.id, []))


def write_into_template(
    session: Session,
    template_path: str | Path,
    out_path: str | Path,
    week_id: int | None = None,
    *,
    colour_by_class: bool = True,
    timetable_session_id: int | None = None,
) -> ExportReport:
    """Clone template, strip macro-generated sheets, repaint Overall body."""
    template_path = Path(template_path)
    out_path = Path(out_path)
    # Allow overwrite-in-place when user picks the same file for template and
    # output; copyfile would raise SameFileError in that case.
    if template_path.resolve() != out_path.resolve():
        shutil.copyfile(template_path, out_path)
    wb = load_workbook(out_path, keep_vba=True)
    if "Overall" not in wb.sheetnames:
        raise ValueError("Template has no 'Overall' sheet")

    report = ExportReport(out_path=out_path)

    # Drop everything that isn't Overall + the three Templates.
    for name in list(wb.sheetnames):
        if name not in KEPT_SHEETS:
            del wb[name]
            report.sheets_removed.append(name)

    _ensure_class_lecturer_entity_templates(wb)

    week = _resolve_week(session, week_id)
    _rewrite_overall(
        wb["Overall"],
        session,
        week.id,
        report,
        timetable_session_id=timetable_session_id,
    )

    # Per-course / per-staff / per-room tabs cloned from the templates.
    _generate_entity_tabs(
        wb,
        session,
        week.id,
        report,
        colour_by_class=colour_by_class,
        timetable_session_id=timetable_session_id,
    )

    # Embed the round-trip backup payload (hidden sheet).
    _write_backup_sheet(wb, session, timetable_session_id=timetable_session_id)

    wb.save(out_path)
    return report


def _generate_entity_tabs(
    wb,
    session: Session,
    week_id: int,
    report: ExportReport,
    *,
    colour_by_class: bool = True,
    timetable_session_id: int | None = None,
) -> None:
    """Clone each of the 3 templates once per entity, populated with bookings."""
    from ..core.sidebar_order import ordered_courses, ordered_staff

    bookings = _week_bookings(session, week_id)
    used_names = set(wb.sheetnames)

    course_t = wb["Course Template"] if "Course Template" in wb.sheetnames else None
    staff_t = wb["Staff Template"] if "Staff Template" in wb.sheetnames else None
    room_t = wb["Room Template"] if "Room Template" in wb.sheetnames else None

    if course_t is not None:
        for c in ordered_courses(
            session, include_block_cohorts=True, timetable_session_id=timetable_session_id
        ):
            ws = wb.copy_worksheet(course_t)
            ws.title = _safe_sheet_name(c.code, used_names)
            _render_entity_tab(
                ws,
                c.code,
                [b for b in bookings if b.course_id == c.id],
                "course",
                colour_by_class=colour_by_class,
            )
            report.course_tabs += 1

    if staff_t is not None:
        for s in ordered_staff(session, timetable_session_id=timetable_session_id):
            ws = wb.copy_worksheet(staff_t)
            ws.title = _safe_sheet_name(s.name, used_names)
            staff_bookings = _bookings_for_staff_tab(bookings, s.id)
            _render_entity_tab(
                ws,
                s.name,
                staff_bookings,
                "staff",
                view_staff_id=s.id,
                colour_by_class=colour_by_class,
            )
            report.staff_tabs += 1

    if room_t is not None:
        room_q = session.query(Room).order_by(Room.code)
        if timetable_session_id is not None:
            room_q = room_q.filter(Room.timetable_session_id == timetable_session_id)
        rooms = [r for r in room_q.all() if room_is_physical(r)]
        for r in rooms:
            ws = wb.copy_worksheet(room_t)
            ws.title = _safe_sheet_name(r.code, used_names)
            _render_entity_tab(
                ws,
                r.code,
                [b for b in bookings if b.room_id == r.id],
                "room",
                colour_by_class=colour_by_class,
            )
            report.room_tabs += 1

    # The template sheets were only needed as a source for cloning the
    # per-entity tabs — drop them from the final export.
    for tpl_name in ("Course Template", "Staff Template", "Room Template"):
        if tpl_name in wb.sheetnames:
            del wb[tpl_name]


# ---------------- Plain-xlsx fallback (no template) ----------------

def write_fresh(
    session: Session,
    out_path: str | Path,
    week_id: int | None = None,
    *,
    timetable_session_id: int | None = None,
) -> ExportReport:
    """Emit a styled .xlsx (no macros) with Overall + per-entity tabs."""
    out_path = Path(out_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Overall"
    week = _resolve_week(session, week_id)
    _populate_overall_fresh(
        ws, session, week.id, timetable_session_id=timetable_session_id
    )

    report = ExportReport(out_path=out_path)
    bookings = _week_bookings(session, week.id)
    used = set(wb.sheetnames)

    from ..core.sidebar_order import ordered_courses, ordered_staff

    for c in ordered_courses(
        session, include_block_cohorts=True, timetable_session_id=timetable_session_id
    ):
        sheet = wb.create_sheet(_safe_sheet_name(c.code, used))
        _scaffold_entity_tab(sheet)
        _render_entity_tab(sheet, c.code, [b for b in bookings if b.course_id == c.id], "course")
        report.course_tabs += 1
    for s in ordered_staff(session, timetable_session_id=timetable_session_id):
        sheet = wb.create_sheet(_safe_sheet_name(s.name, used))
        _scaffold_entity_tab(sheet)
        staff_bookings = _bookings_for_staff_tab(bookings, s.id)
        _render_entity_tab(
            sheet, s.name, staff_bookings, "staff", view_staff_id=s.id
        )
        report.staff_tabs += 1
    room_q = session.query(Room).order_by(Room.code)
    if timetable_session_id is not None:
        room_q = room_q.filter(Room.timetable_session_id == timetable_session_id)
    for r in [r for r in room_q.all() if room_is_physical(r)]:
        sheet = wb.create_sheet(_safe_sheet_name(r.code, used))
        _scaffold_entity_tab(sheet)
        _render_entity_tab(sheet, r.code, [b for b in bookings if b.room_id == r.id], "room")
        report.room_tabs += 1

    _write_backup_sheet(wb, session, timetable_session_id=timetable_session_id)
    wb.save(out_path)
    return report


def _scaffold_entity_tab(ws) -> None:
    """Build the time grid + day headers on a per-entity tab from scratch."""
    from openpyxl.utils import get_column_letter as col_letter
    header_font = Font(name="Calibri", size=10, bold=True, color="FF1F2937")
    header_fill = PatternFill(start_color="FFE9ECEF", end_color="FFE9ECEF", fill_type="solid")
    time_font = Font(name="Calibri", size=9, color="FF555555")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.cell(row=5, column=1, value="Start\nTime").font = header_font
    ws.cell(row=5, column=1).alignment = align_center
    for d, name in enumerate(DAYS_FOR_BANNER):
        col = ENTITY_DAY_FIRST_COL + d * ENTITY_DAY_STRIDE
        cell = ws.cell(row=6, column=col, value=name.title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        ws.merge_cells(start_row=6, end_row=6, start_column=col, end_column=col + 2)
    for slot in range(NUM_SLOTS):
        row = ENTITY_FIRST_SLOT_ROW + slot
        cell = ws.cell(row=row, column=1, value=slot_to_time(slot).strftime("%H:%M"))
        cell.font = time_font
        cell.alignment = align_center
        ws.row_dimensions[row].height = 18

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 2  # day-block separator gap
    for d in range(NUM_DAYS):
        col = ENTITY_DAY_FIRST_COL + d * ENTITY_DAY_STRIDE
        ws.column_dimensions[col_letter(col)].width = 12      # unit
        ws.column_dimensions[col_letter(col + 1)].width = 8   # room
        ws.column_dimensions[col_letter(col + 2)].width = 14  # lecturer / course
        ws.column_dimensions[col_letter(col + 3)].width = 1   # gap to next day
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[6].height = 22
    ws.freeze_panes = "C7"


def _populate_overall_fresh(
    ws,
    session: Session,
    week_id: int,
    *,
    timetable_session_id: int | None = None,
) -> None:
    """Build a styled Overall sheet from scratch with card-merged bookings."""
    from openpyxl.utils import get_column_letter as col_letter

    course_q = session.query(Course).order_by(Course.id)
    staff_q = session.query(Staff).order_by(Staff.name)
    room_q = session.query(Room).order_by(Room.code)
    if timetable_session_id is not None:
        course_q = course_q.filter(Course.timetable_session_id == timetable_session_id)
        staff_q = staff_q.filter(Staff.timetable_session_id == timetable_session_id)
        room_q = room_q.filter(Room.timetable_session_id == timetable_session_id)
    courses = course_q.all()
    staff_all = staff_q.all()
    rooms = room_q.all()

    header_fill = PatternFill(start_color="FFE9ECEF", end_color="FFE9ECEF", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FF1F2937")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    time_font = Font(name="Calibri", size=9, color="FF555555")
    time_align = Alignment(horizontal="center", vertical="center")
    day_banner_fill = PatternFill(start_color="FFF1F3F5", end_color="FFF1F3F5", fill_type="solid")
    day_banner_font = Font(name="Calibri", size=11, bold=True, color="FF111827")
    grid_side = Side(border_style="thin", color="FFD0D7DE")
    grid_border = Border(top=grid_side, bottom=grid_side, left=grid_side, right=grid_side)

    # Row 1: section markers + per-course headers (neutral; booking cells are class-tinted)
    ws.cell(row=1, column=2, value="Start\nTime").font = header_font
    ws.cell(row=1, column=4, value="End Time").font = header_font
    next_col = 5
    # The 'Facilitation' marker doubles as the first course's header in the
    # source convention. If our first course isn't named that, reserve a
    # 3-col placeholder block holding only the marker so importers can find it.
    if not courses or courses[0].code != SECTION_FACILITATION:
        ws.cell(row=1, column=next_col, value=SECTION_FACILITATION).font = header_font
        ws.merge_cells(start_row=1, end_row=1, start_column=next_col, end_column=next_col + 2)
        for off, label in enumerate(("Subject", "Room", "Lecturer")):
            cell = ws.cell(row=2, column=next_col + off, value=label)
            cell.fill = header_fill
            cell.font = time_font
            cell.alignment = header_align
            cell.border = grid_border
        next_col += 3
    course_col: dict[int, int] = {}
    for c in courses:
        cell = ws.cell(row=1, column=next_col, value=c.code)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = grid_border
        ws.merge_cells(start_row=1, end_row=1, start_column=next_col, end_column=next_col + 2)
        course_col[c.id] = next_col
        next_col += 3
    staff_marker_col = next_col
    ws.cell(row=1, column=staff_marker_col, value=SECTION_STAFF).font = header_font
    ws.cell(row=1, column=staff_marker_col).fill = header_fill
    next_col += 1
    for s in staff_all:
        cell = ws.cell(row=2, column=next_col, value=s.name)
        cell.alignment = header_align
        cell.font = time_font
        next_col += 1
    rooms_marker_col = next_col
    ws.cell(row=1, column=rooms_marker_col, value=SECTION_ROOMS).font = header_font
    ws.cell(row=1, column=rooms_marker_col).fill = header_fill
    next_col += 1
    for r in rooms:
        cell = ws.cell(row=2, column=next_col, value=r.code)
        cell.alignment = header_align
        cell.font = time_font
        next_col += 1
    last_col = next_col - 1

    # Row 2: per-course sub-headers (Subject / Room / Lecturer)
    for c in courses:
        col = course_col[c.id]
        for off, label in enumerate(("Subject", "Room", "Lecturer")):
            cell = ws.cell(row=2, column=col + off, value=label)
            cell.fill = header_fill
            cell.font = time_font
            cell.alignment = header_align
            cell.border = grid_border

    # Time grid (col B start, col C "-", col D end)
    for day in range(NUM_DAYS):
        first = _day_first_row(day)
        banner_row = first - 1
        if banner_row >= 1:
            cell = ws.cell(row=banner_row, column=1, value=DAYS_FOR_BANNER[day])
            cell.fill = day_banner_fill
            cell.font = day_banner_font
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.merge_cells(start_row=banner_row, end_row=banner_row, start_column=1, end_column=last_col)
            ws.row_dimensions[banner_row].height = 22

        for slot in range(NUM_SLOTS):
            row = first + slot
            ws.row_dimensions[row].height = 18
            for col, val in (
                (2, slot_to_time(slot).strftime("%H:%M")),
                (3, "–"),
                (4, slot_to_time(slot + 1).strftime("%H:%M") if slot + 1 < NUM_SLOTS else "22:00"),
            ):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = time_font
                cell.alignment = time_align

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 7
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 7
    for c in courses:
        col = course_col[c.id]
        ws.column_dimensions[col_letter(col)].width = 15
        ws.column_dimensions[col_letter(col + 1)].width = 7
        ws.column_dimensions[col_letter(col + 2)].width = 14

    bookings = _week_bookings(session, week_id)
    by_course: dict[int, list[Booking]] = {}
    for b in bookings:
        by_course.setdefault(b.course_id, []).append(b)
    for c in courses:
        col = course_col[c.id]
        _write_course_body(ws, col, by_course.get(c.id, []), c.code)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "E3"


# Title-cased day banners (constants.DAYS already provides these).
DAYS_FOR_BANNER = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY"]
