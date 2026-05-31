"""Import completed lecturer-preferences spreadsheets into the session."""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import re

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..constants import DAYS, NUM_DAYS, NUM_SLOTS
from ..core.models import Staff, StaffAvailability, StaffPreference, Unit


@dataclass
class LecturerPrefsImportReport:
    staff_updated: int = 0
    preferences_imported: int = 0
    avail_windows_written: int = 0
    warnings: list[str] = field(default_factory=list)


_TIME_RANGE_RE = re.compile(r"^\s*(\d{2}):(\d{2})[–-](\d{2}):(\d{2})\s*$")


def _title_staff_name(v: object) -> str | None:
    text = (str(v or "")).strip()
    if not text.lower().startswith("preferences"):
        return None
    if "—" in text:
        return text.split("—", 1)[1].strip() or None
    if "-" in text:
        return text.split("-", 1)[1].strip() or None
    return None


def _priority_num(v: object) -> int | None:
    t = (str(v or "")).strip().lower()
    if t == "first":
        return 1
    if t == "second":
        return 2
    if t == "third":
        return 3
    return None


def _slot_from_label(v: object) -> int | None:
    m = _TIME_RANGE_RE.match(str(v or ""))
    if not m:
        return None
    hh, mm = int(m.group(1)), int(m.group(2))
    return (hh - 8) * 2 + (1 if mm >= 30 else 0)


def _windows_from_blocked(blocked: set[tuple[int, int]]) -> list[tuple[int, int, int]]:
    """Convert blocked slots to availability windows (day, start, end)."""
    out: list[tuple[int, int, int]] = []
    for day in range(NUM_DAYS):
        s = 0
        while s < NUM_SLOTS:
            while s < NUM_SLOTS and (day, s) in blocked:
                s += 1
            if s >= NUM_SLOTS:
                break
            e = s
            while e < NUM_SLOTS and (day, e) not in blocked:
                e += 1
            out.append((day, s, e))
            s = e
    return out


def import_lecturer_preferences(session: Session, xlsx_path: str | Path) -> LecturerPrefsImportReport:
    rep = LecturerPrefsImportReport()
    wb = load_workbook(xlsx_path, data_only=True)

    staff_by_name = {s.name.strip().lower(): s for s in session.query(Staff).all() if (s.name or "").strip()}
    unit_by_name = {u.name.strip().lower(): u for u in session.query(Unit).all() if (u.name or "").strip()}

    for ws in wb.worksheets:
        if ws.title.startswith("_"):
            continue
        staff_name = _title_staff_name(ws["A1"].value) or ws.title.strip()
        staff = staff_by_name.get(staff_name.lower())
        if staff is None:
            rep.warnings.append(f"Sheet {ws.title!r}: lecturer {staff_name!r} not found in session")
            continue

        # Preferences rows are 6..11 in current template.
        session.query(StaffPreference).filter(StaffPreference.staff_id == staff.id).delete()
        slot_counts: dict[int, int] = {1: 0, 2: 0, 3: 0}
        for row in range(6, 12):
            pnum = _priority_num(ws.cell(row=row, column=1).value)
            if pnum is None:
                continue
            qname = (str(ws.cell(row=row, column=2).value or "")).strip() or None
            cname = (str(ws.cell(row=row, column=3).value or "")).strip() or None
            if not qname and not cname:
                continue
            slot_counts[pnum] += 1
            unit = unit_by_name.get((cname or "").lower()) if cname else None
            session.add(
                StaffPreference(
                    staff_id=staff.id,
                    priority=pnum,
                    slot_number=min(2, max(1, slot_counts[pnum])),
                    qualification_name=qname,
                    class_name=cname,
                    unit_id=unit.id if unit else None,
                )
            )
            rep.preferences_imported += 1

        # Non-teaching day lives in B14.
        non_teaching = (str(ws.cell(row=14, column=2).value or "")).strip()
        if non_teaching:
            try:
                staff.non_teaching_day = DAYS.index(non_teaching)
            except ValueError:
                rep.warnings.append(
                    f"Sheet {ws.title!r}: non-teaching day {non_teaching!r} is not valid"
                )
                staff.non_teaching_day = None
        else:
            staff.non_teaching_day = None

        blocked: set[tuple[int, int]] = set()
        # Grid headers start around row 21; we scan by detecting time labels in col A.
        max_row = min(ws.max_row, 200)
        for row in range(1, max_row + 1):
            slot = _slot_from_label(ws.cell(row=row, column=1).value)
            if slot is None or slot < 0 or slot >= NUM_SLOTS:
                continue
            for day in range(NUM_DAYS):
                raw = ws.cell(row=row, column=2 + day).value
                if str(raw or "").strip().upper() == "X":
                    blocked.add((day, slot))

        session.query(StaffAvailability).filter(StaffAvailability.staff_id == staff.id).delete()
        # Blocked grid only — non-teaching day is stored on staff.non_teaching_day.
        if blocked:
            for day, start, end in _windows_from_blocked(blocked):
                if end > start:
                    session.add(
                        StaffAvailability(
                            staff_id=staff.id, day=day, start_slot=start, end_slot=end
                        )
                    )
                    rep.avail_windows_written += 1
        rep.staff_updated += 1

    session.commit()
    return rep
