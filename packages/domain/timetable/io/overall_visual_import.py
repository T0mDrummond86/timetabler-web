"""Import timetable data from a legacy Joondalup-style Overall workbook.

Reads the visual grid on sheet ``Overall``: row~1 merged course headers (3-column
blocks: class / room / lecturer), ``Staff!`` / ``Rooms!`` markers, and
29-row day blocks (28 half-hour slots + day banner) matching ``xlsm_export``.

This path does **not** use embedded ``__timetable_data__`` JSON; it parses cell
values only and replaces all bookings for the current week.
"""
from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass, field

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..constants import NUM_DAYS, NUM_SLOTS
from ..core.booking_staff import apply_parsed_lecturers_to_booking, parse_import_lecturer_label
from ..core.unit_brackets import (
    apply_unit_bracket_fields_from_names,
    normalize_class_label_for_parse,
    normalize_component_codes_commas,
    split_class_title_and_unit_codes,
)
from ..core.qualification_schedule import (
    SCHEDULE_PERIOD_DAY,
    replace_qualification_time_windows,
)
from ..core.models import (
    Booking,
    Course,
    CourseUnit,
    Qualification,
    Room,
    Staff,
    Unit,
    UnitQualification,
    Week,
)
from .xlsm_export import (
    DAY_BLOCK_STRIDE,
    FIRST_DATA_ROW,
    _course_columns_in_template,
    _scan_row1_markers,
)
from .visual_import_session import VisualImportContext


def _normalized_unit_codes_equivalent(a: str, b: str) -> bool:
    """True when ``a`` and ``b`` are the same code list after comma normalization."""
    a = (a or "").strip()
    b = (b or "").strip()
    if not a or not b:
        return False
    na = normalize_component_codes_commas(a) or a
    nb = normalize_component_codes_commas(b) or b
    return na == nb


def _day_first_row(day: int) -> int:
    return day * DAY_BLOCK_STRIDE + FIRST_DATA_ROW


_DURATION_RE = re.compile(r"^\d+\s*hrs?$", re.I)
_EXTERNAL_TAG_RE = re.compile(r"^\s*\[([^\]]+)\]\s*(.*)$", re.S)
_PLACE_CARD_ID_RE = re.compile(r"\bID:?\s*(\d{7})\b", re.IGNORECASE)
_PLACEHOLDER_LECTURER_NAMES = frozenset({"—", "-", "n/a", "tbc", "tba"})
_GROUP_SUFFIX_RE = re.compile(r"\s+Grp[0-9A-Za-z]+$", re.I)


def _base_qualification_name(course_code: str) -> str:
    """Best-effort: strip trailing group suffix (e.g. '... GrpA' -> '...')."""
    cc = (course_code or "").strip() or "Untitled"
    base = _GROUP_SUFFIX_RE.sub("", cc).strip()
    return base or cc


def _strip_cell(val: object | None) -> str:
    if val is None:
        return ""
    if hasattr(val, "strftime"):
        return ""
    s = str(val).strip()
    return s


def _lecturer_placeholder(name: str) -> bool:
    return name.strip().lower() in _PLACEHOLDER_LECTURER_NAMES


def _normalize_lecturer_name(name: str) -> str | None:
    n = name.strip()
    if not n:
        return None
    if n.lower() == "all":
        return "All"
    return n


def _lecturers_match(left: str | None, right: str | None) -> bool:
    return _normalize_lecturer_name(left or "") == _normalize_lecturer_name(right or "")


def _is_duration_only(text: str) -> bool:
    return bool(text and _DURATION_RE.match(text.strip()))


def _noise_unit(text: str) -> bool:
    t = text.strip().lower()
    return t in {"section", "professional", "development"}


def _extract_placecard_id(text: str) -> tuple[str, str | None]:
    """Pull ``ID: 1234567`` from placecard unit text; return (remainder, id or None)."""
    s = text.strip()
    if not s:
        return "", None
    m = _PLACE_CARD_ID_RE.search(s)
    if not m:
        return s, None
    placecard_id = m.group(1)
    remainder = _PLACE_CARD_ID_RE.sub(" ", s)
    remainder = re.sub(r"\s+", " ", remainder).strip()
    return remainder, placecard_id


def _clean_unit_piece(text: str) -> str | None:
    s = text.strip()
    if not s or _is_duration_only(s):
        return None
    remainder, _placecard_id = _extract_placecard_id(s)
    return remainder or None


def _join_unit_parts(parts: list[str]) -> str:
    out: list[str] = []
    for p in parts:
        c = p.strip()
        if not c:
            continue
        if out and c.lower() == out[-1].lower():
            continue
        out.append(c)
    return " ".join(out)


def _resolve_unit_storage_name(
    session: Session,
    title: str,
    suffix: str | None,
    *,
    timetable_session_id: int | None = None,
) -> tuple[str, str | None]:
    """Return ``(Unit.name, component_codes)`` for create/lookup.

    When ``suffix`` is None, ``title`` is the full display string. Otherwise
    ``title`` is the class title and ``suffix`` populates ``Unit.component_codes``
    (DB column ``units``). If a unit already exists under ``title`` with a
    different non-empty code, use a disambiguated ``name`` so uniqueness holds.
    """
    if suffix is None:
        return title.strip(), None
    t = title.strip()
    c_raw = suffix.strip()
    c = (normalize_component_codes_commas(c_raw) or c_raw) if c_raw else ""
    if not c:
        return t, None
    q = session.query(Unit).filter(Unit.name == t)
    if timetable_session_id is not None:
        q = q.filter(Unit.timetable_session_id == timetable_session_id)
    row = q.one_or_none()
    if row is None:
        return t, c
    existing = (row.component_codes or "").strip()
    if existing == "" or existing == c or _normalized_unit_codes_equivalent(existing, c):
        return t, c
    return f"{t} ({c})", c


def _parse_external_and_terms(
    raw_name: str, *, placecard_id: str | None = None
) -> tuple[str, str | None, int, int]:
    """Return (display_name, external_id or None, in_term_1, in_term_2)."""
    name = raw_name.strip()
    ext: str | None = placecard_id
    name, embedded_id = _extract_placecard_id(name)
    if embedded_id:
        ext = embedded_id
    m = _EXTERNAL_TAG_RE.match(name)
    if m:
        if ext is None:
            ext = m.group(1).strip() or None
        name = (m.group(2) or "").strip()
    t1, t2 = 1, 1
    upper = name.upper()
    if "[T1]" in upper and "[T2]" not in upper:
        t1, t2 = 1, 0
        name = re.sub(r"\[T1\]", "", name, flags=re.I).strip()
    elif "[T2]" in upper and "[T1]" not in upper:
        t1, t2 = 0, 1
        name = re.sub(r"\[T2\]", "", name, flags=re.I).strip()
    elif "[T1]" in upper or "[T2]" in upper:
        name = re.sub(r"\[T[12]\]", "", name, flags=re.I).strip()
    return name, ext, t1, t2


def _parse_day_course_column(ws, day: int, base_col: int) -> list[dict]:
    """Build booking segments for one course column on one day."""
    first = _day_first_row(day)
    segments: list[dict] = []
    cur: dict | None = None

    def flush() -> None:
        nonlocal cur
        if cur is None:
            return
        if cur["end_slot"] > cur["start_slot"] and _join_unit_parts(cur["unit_parts"]).strip():
            segments.append(cur)
        cur = None

    for s in range(NUM_SLOTS):
        r = first + s
        u_raw = _strip_cell(ws.cell(row=r, column=base_col).value)
        rm_raw = _strip_cell(ws.cell(row=r, column=base_col + 1).value)
        lec_raw = _strip_cell(ws.cell(row=r, column=base_col + 2).value)
        rm_key = rm_raw.strip()
        lec_key = lec_raw.strip()
        primary, co_name, co_t1, co_t2 = parse_import_lecturer_label(lec_key)
        lec_primary = _normalize_lecturer_name(primary or "")

        # Row whose only unit text is like "3hrs" — duration tag on the last slot
        # of the block (same room/lecturer). Include this slot, then close.
        if u_raw and _is_duration_only(u_raw):
            if (
                cur is not None
                and lec_key == cur["lecturer_key"]
                and rm_key == cur["room"]
            ):
                cur["end_slot"] = s + 1
            flush()
            continue

        if _lecturer_placeholder(lec_key) or _lecturer_placeholder(primary or ""):
            flush()
            continue

        if not _clean_unit_piece(u_raw) and not rm_key and lec_primary is None:
            flush()
            continue

        if _noise_unit(u_raw) and not rm_key:
            flush()
            continue

        piece = _clean_unit_piece(u_raw)
        _remainder, row_id = _extract_placecard_id(u_raw)
        if row_id:
            if cur is None:
                if not piece and not rm_key and lec_primary is None:
                    continue
                cur = {
                    "start_slot": s,
                    "end_slot": s + 1,
                    "unit_parts": [piece] if piece else [],
                    "room": rm_key,
                    "lecturer_key": lec_key,
                    "primary_name": lec_primary,
                    "co_name": _normalize_lecturer_name(co_name or "") if co_name else None,
                    "co_t1": co_t1,
                    "co_t2": co_t2,
                    "placecard_id": row_id,
                }
                continue
            cur["placecard_id"] = row_id

        if cur is None:
            if not piece and not rm_key:
                continue
            cur = {
                "start_slot": s,
                "end_slot": s + 1,
                "unit_parts": [piece] if piece else [],
                "room": rm_key,
                "lecturer_key": lec_key,
                "primary_name": lec_primary,
                "co_name": _normalize_lecturer_name(co_name or "") if co_name else None,
                "co_t1": co_t1,
                "co_t2": co_t2,
            }
            continue

        same = lec_key == cur["lecturer_key"] and rm_key == cur["room"]
        if same:
            cur["end_slot"] = s + 1
            if piece:
                cur["unit_parts"].append(piece)
        else:
            flush()
            cur = {
                "start_slot": s,
                "end_slot": s + 1,
                "unit_parts": [piece] if piece else [],
                "room": rm_key,
                "lecturer_key": lec_key,
                "primary_name": lec_primary,
                "co_name": _normalize_lecturer_name(co_name or "") if co_name else None,
                "co_t1": co_t1,
                "co_t2": co_t2,
            }
            if row_id:
                cur["placecard_id"] = row_id

    flush()
    return segments


@dataclass
class OverallVisualImportReport:
    courses_touched: int = 0
    bookings_created: int = 0
    qualifications_created: int = 0
    units_created: int = 0
    staff_created: int = 0
    rooms_created: int = 0
    warnings: list[str] = field(default_factory=list)


def import_overall_visual(
    session: Session,
    xlsm_path: str,
    *,
    timetable_session_id: int | None = None,
) -> OverallVisualImportReport:
    """Parse ``Overall`` and replace all bookings for the first week."""
    ctx = VisualImportContext(timetable_session_id)
    wb = load_workbook(xlsm_path, data_only=True, keep_vba=False)
    if "Overall" not in wb.sheetnames:
        raise ValueError("Workbook has no 'Overall' sheet.")
    ws = wb["Overall"]

    markers = _scan_row1_markers(ws)
    course_cols = _course_columns_in_template(ws, markers)
    if not course_cols:
        raise ValueError("No course columns found between Facilitation and Staff!.")

    week = ctx.first_week(session)

    report = OverallVisualImportReport()
    report.courses_touched = len(course_cols)

    session.query(Booking).filter(Booking.week_id == week.id).delete()

    staff_cache: dict[str, int] = {}
    room_cache: dict[str, int] = {}
    unit_cache: dict[str, int] = {}
    course_unit_sets: dict[str, set[int]] = defaultdict(set)

    def get_staff(name: str) -> int:
        key = name.strip()
        if key in staff_cache:
            return staff_cache[key]
        row = ctx.staff_by_name(session, key)
        if row is None:
            row = ctx.new_staff(key)
            session.add(row)
            session.flush()
            report.staff_created += 1
        staff_cache[key] = row.id
        return row.id

    def resolve_staff_id(name: str | None) -> int | None:
        if name is None:
            return None
        return get_staff(name)

    def get_room(code: str) -> int | None:
        if not code.strip():
            return None
        key = code.strip()
        if key in room_cache:
            return room_cache[key]
        row = ctx.room_by_code(session, key)
        if row is None:
            from ..core.room_types import ROOM_TYPE_ON_CAMPUS

            row = ctx.new_room(key, room_type=ROOM_TYPE_ON_CAMPUS)
            session.add(row)
            session.flush()
            report.rooms_created += 1
        room_cache[key] = row.id
        return row.id

    def get_unit(display_name: str, span_slots: int) -> int | None:
        display_name = display_name.strip()
        if not display_name:
            return None

        split_title, suffix = split_class_title_and_unit_codes(display_name)
        name_key, comp_codes = _resolve_unit_storage_name(
            session,
            split_title,
            suffix,
            timetable_session_id=ctx.timetable_session_id,
        )

        for key in (display_name, name_key):
            if key in unit_cache:
                uid = unit_cache[key]
                u = session.get(Unit, uid)
                if u is not None:
                    need = max(span_slots, 1)
                    if u.length_slots is None or u.length_slots < need:
                        u.length_slots = need
                    if comp_codes is not None:
                        ex = (u.component_codes or "").strip()
                        if not ex or ex == comp_codes or _normalized_unit_codes_equivalent(
                            ex, comp_codes
                        ):
                            u.component_codes = comp_codes
                unit_cache[display_name] = uid
                unit_cache[name_key] = uid
                return uid

        # Rows created before bracket-splitting used the full cell text as ``Unit.name``.
        if suffix is not None:
            leg = None
            for cand in {display_name, normalize_class_label_for_parse(display_name)}:
                if not cand:
                    continue
                leg = ctx.unit_by_name(session, cand)
                if leg is not None:
                    break
            if leg is not None:
                other = ctx.unit_by_name(session, name_key)
                if other is None or other.id == leg.id:
                    leg.name = name_key
                    if comp_codes:
                        ex = (leg.component_codes or "").strip()
                        if not ex or ex == comp_codes or _normalized_unit_codes_equivalent(
                            ex, comp_codes
                        ):
                            leg.component_codes = comp_codes
                elif comp_codes and not (leg.component_codes or "").strip():
                    leg.component_codes = comp_codes
                need = max(span_slots, 1)
                if leg.length_slots is None or leg.length_slots < need:
                    leg.length_slots = need
                session.flush()
                unit_cache[leg.name] = leg.id
                unit_cache[display_name] = leg.id
                unit_cache[name_key] = leg.id
                return leg.id

        row = ctx.unit_by_name(session, name_key)
        if row is None:
            row = ctx.new_unit(
                name_key,
                length_slots=max(span_slots, 1),
                component_codes=comp_codes,
            )
            session.add(row)
            session.flush()
            report.units_created += 1
        else:
            need = max(span_slots, 1)
            if row.length_slots is None or row.length_slots < need:
                row.length_slots = need
            if comp_codes is not None:
                ex = (row.component_codes or "").strip()
                if not ex or ex == comp_codes or _normalized_unit_codes_equivalent(
                    ex, comp_codes
                ):
                    row.component_codes = comp_codes
        unit_cache[name_key] = row.id
        unit_cache[display_name] = row.id
        return row.id

    linked_course_units: set[tuple[int, int]] = set()

    def ensure_course_unit(course_id: int, unit_id: int | None) -> None:
        if unit_id is None:
            return
        key = (course_id, unit_id)
        if key in linked_course_units:
            return
        exists = (
            session.query(CourseUnit)
            .filter_by(course_id=course_id, unit_id=unit_id)
            .first()
        )
        if not exists:
            session.add(CourseUnit(course_id=course_id, unit_id=unit_id))
        linked_course_units.add(key)

    for course_code, base_col in course_cols.items():
        course_code = (course_code or "").strip() or "Untitled"
        course = ctx.course_by_code(session, course_code)
        if course is None:
            course = ctx.new_course(course_code)
            session.add(course)
            session.flush()

        for day in range(NUM_DAYS):
            if base_col + 2 > ws.max_column:
                report.warnings.append(
                    f"{course_code}: day {day + 1} — column block exceeds sheet width"
                )
                break
            segments = _parse_day_course_column(ws, day, base_col)
            for seg in segments:
                span = seg["end_slot"] - seg["start_slot"]
                if span <= 0:
                    continue
                joined = _join_unit_parts(seg["unit_parts"])
                display, ext, it1, it2 = _parse_external_and_terms(
                    normalize_class_label_for_parse(joined),
                    placecard_id=seg.get("placecard_id"),
                )
                uid = get_unit(display, span)
                rid = get_room(seg["room"])
                ensure_course_unit(course.id, uid)
                if uid is not None:
                    course_unit_sets[course_code].add(int(uid))
                booking = Booking(
                    week_id=week.id,
                    course_id=course.id,
                    unit_id=uid,
                    room_id=rid,
                    day=day,
                    start_slot=seg["start_slot"],
                    end_slot=seg["end_slot"],
                    external_id=ext,
                    in_term_1=it1,
                    in_term_2=it2,
                )
                co_norm = seg.get("co_name")
                apply_parsed_lecturers_to_booking(
                    booking,
                    primary_name=seg.get("primary_name"),
                    co_teacher_name=(
                        co_norm if co_norm and not _lecturer_placeholder(co_norm) else None
                    ),
                    co_in_term_1=seg.get("co_t1"),
                    co_in_term_2=seg.get("co_t2"),
                    resolve_staff_id=resolve_staff_id,
                )
                session.add(booking)
                report.bookings_created += 1

    # Dedupe courses into qualifications by identical class sets.
    sig_to_courses: dict[frozenset[int], list[str]] = defaultdict(list)
    for cc, units in course_unit_sets.items():
        sig_to_courses[frozenset(units)].append(cc)

    used_qual_names: set[str] = set()
    linked_unit_quals: set[tuple[int, int]] = set()
    for sig, course_codes in sig_to_courses.items():
        if not sig:
            # No classes parsed for these courses; leave unqualified.
            continue
        base_name = _base_qualification_name(course_codes[0])
        qual_name = base_name
        n = 1
        while qual_name in used_qual_names:
            n += 1
            qual_name = f"{base_name} ({n})"
        used_qual_names.add(qual_name)

        q = ctx.qualification_by_name(session, qual_name)
        if q is None:
            q = ctx.new_qualification(
                name=qual_name,
                num_groups=len(course_codes),
                schedule_period=SCHEDULE_PERIOD_DAY,
            )
            session.add(q)
            session.flush()
            replace_qualification_time_windows(session, q)
            report.qualifications_created += 1
        else:
            q.num_groups = len(course_codes)

        for cc in course_codes:
            c = ctx.course_by_code(session, cc)
            if c is not None:
                c.qualification_id = q.id
        for uid in sig:
            uq_key = (uid, q.id)
            if uq_key in linked_unit_quals:
                continue
            exists = (
                session.query(UnitQualification)
                .filter_by(unit_id=uid, qualification_id=q.id)
                .first()
            )
            if not exists:
                session.add(UnitQualification(unit_id=uid, qualification_id=q.id))
            linked_unit_quals.add(uq_key)

    apply_unit_bracket_fields_from_names(
        session, timetable_session_id=timetable_session_id
    )
    session.commit()
    return report
