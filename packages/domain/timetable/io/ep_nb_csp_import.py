"""Import qualifications + classes from NMTAFE CSP planning workbooks (.xlsx).

These workbooks are laid out by hand and come in more than one shape, so the
parser reads the header row to find out where things are rather than assuming
fixed columns. Two shapes are known:

  * a **BB Shell** layout — column A names the class, and the skill-set text in
    column F is a fallback when column A is a placeholder such as ``???``
  * a **lecturer** layout — column A names the lecturer, and the class name
    comes from the skill-set description instead

Common to both: the class's weekly hours, a TPN unit code per row, and the unit
of competency title. A class spanning several units is one block — the first
row carries the name and hours, and continuation rows carry only their TPN.

Bands (``Semester 1``, ``Part 2 • …``) group the rows visually. They are read
when present but are not required: the importer produces one qualification
regardless, because the bands describe the curriculum's shape rather than the
timetable's.
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .csp_qualification_import import CspClass, CspStage, flatten_stages
from .qualification_import import (
    QualImportReport,
    _create_kwargs,
    _parse_running_time,
    _scoped_filter_by,
    _text,
)
from ..core.models import Course, Qualification, Unit, UnitQualification
from ..core.qualification_schedule import SCHEDULE_PERIOD_DAY, replace_qualification_time_windows
from ..core.unit_brackets import apply_unit_bracket_fields_from_names, normalize_component_codes_commas

#: Fallback positions, used only when a header row does not name the column.
_COL_BB_SHELL = 0
_COL_HOURS = 1
_COL_SKILL_SET = 5
_COL_TPN = 7
_COL_UOC_TITLE = 8
_COL_SEMESTER = 1

#: Bands that group rows. "Part 1 • Thursday night" is as common as "Semester 1",
#: and neither is required — see the module docstring. Deliberately excludes
#: "Term N": real unit descriptions begin with it, and treating those as bands
#: chopped a semester's classes in half.
_BAND_RE = re.compile(r"^(semester|part|stage)\s+(\d+)", re.IGNORECASE)
#: A band is a heading, so its row is nearly empty. A populated row that merely
#: mentions a band is a class row.
_MAX_CELLS_ON_A_BAND_ROW = 3
#: Column A sometimes holds a stand-in rather than a class name.
_PLACEHOLDER_SHELLS = frozenset({"???", "??", "?", "-", "—", "n/a", "na"})
#: Given to a class whose name is blank in the workbook, so the units still
#: import and the gap is obvious enough to be filled in afterwards.
_UNNAMED_PREFIX = "Unnamed class"


class _Layout:
    """Where the columns are, read from the workbook's own header row.

    These workbooks are maintained by people, and the columns move. Reading the
    header is the difference between supporting one variant and supporting the
    ones that have not been written yet.
    """

    def __init__(self, row: tuple) -> None:
        self.label: int | None = None
        self.hours: int = _COL_HOURS
        self.skill: int = _COL_SKILL_SET
        self.tpn: int = _COL_TPN
        self.uoc: int = _COL_UOC_TITLE
        for index, value in enumerate(row):
            text = (_clean_cell(value) or "").lower()
            if not text:
                continue
            if text.startswith("bb shell"):
                self.label = index
            elif text.startswith("hrs"):
                self.hours = index
            elif text.startswith("skill set"):
                self.skill = index
            elif text == "tpn":
                self.tpn = index
            elif text.startswith("uoc"):
                self.uoc = index

    @property
    def names_the_class_in_column_a(self) -> bool:
        """True for the BB Shell layout, false where column A is the lecturer."""
        return self.label is not None


def _clean_cell(v) -> str | None:
    if v is None:
        return None
    s = str(v).replace("\xa0", " ").strip()
    return s or None


def _parse_hours(v) -> float | None:
    if v is None:
        return None
    try:
        n = float(v)
    except (TypeError, ValueError):
        return None
    return n if n > 0 else None


def _is_header_row(row: tuple) -> bool:
    """A row that names its columns.

    Keyed on TPN plus one other recognisable heading rather than on "BB Shell",
    which only one of the layouts has.
    """
    labels = {(_clean_cell(v) or "").lower() for v in row}
    if "tpn" not in labels:
        return False
    others = {"bb shell", "sin", "lecturer(s)", "lecturers", "core / elect"}
    if labels & others:
        return True
    return any(
        text.startswith(("hrs", "skill set", "uoc")) for text in labels if text
    )


def _band_label(row: tuple) -> str | None:
    """A grouping band such as "Semester 1" or "Part 2 • Monday night".

    Scanned across the row, because the band is a hand-written heading and its
    column is not a promise — but only on rows sparse enough to *be* a heading.
    A full class row that happens to start with a band-like word is a class row.
    """
    populated = [t for t in (_clean_cell(v) for v in row) if t]
    if not populated or len(populated) > _MAX_CELLS_ON_A_BAND_ROW:
        return None
    for text in populated:
        m = _BAND_RE.match(text)
        if m:
            return f"{m.group(1).capitalize()} {m.group(2)}"
    return None


def _qualification_title_from_sheet(ws) -> str:
    """The first real text on row 1 — the title is not always in column A."""
    for column in range(1, min(ws.max_column or 1, 12) + 1):
        title = _text(ws.cell(row=1, column=column).value)
        if title:
            return title
    return "Imported qualification"


def _cell(row: tuple, index: int | None) -> str | None:
    if index is None or len(row) <= index:
        return None
    return _clean_cell(row[index])


def _class_name_from_row(row: tuple, layout: "_Layout") -> str | None:
    """The class's name, or None when the workbook does not give one.

    The skill-set description is the name in the lecturer layout and the
    fallback in the BB Shell one, so it is consulted in both. None here is not
    a failure — the caller supplies a placeholder, because a class with no name
    is still a class with units worth importing.
    """
    if layout.names_the_class_in_column_a:
        label = _cell(row, layout.label)
        if label and label.lower() not in _PLACEHOLDER_SHELLS:
            return label

    skill = _cell(row, layout.skill)
    if skill and skill.lower() not in _PLACEHOLDER_SHELLS:
        return skill.replace("\n", " ").strip()
    return None


def _is_subtotal_row(row: tuple, layout: "_Layout") -> bool:
    first = _cell(row, 0)
    if first and first.lower() in {"course total", "total"}:
        return True
    if _cell(row, layout.tpn):
        return False
    hrs = _parse_hours(row[layout.hours] if len(row) > layout.hours else None)
    # Band summary rows (e.g. 19 hrs / 310 actual) carry hours but no TPN.
    return hrs is not None and hrs >= 10


def _qualification_title_from_sheet_has_text(ws) -> bool:
    return _qualification_title_from_sheet(ws) != "Imported qualification"


def is_ep_nb_csp_workbook(path: str | Path) -> bool:
    """Return True when the workbook matches the EP-NB CSP Excel layout."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    try:
        ws = wb.active
        if not _qualification_title_from_sheet_has_text(ws):
            return False
        # A header row is the only structural requirement. Bands are not: some
        # of these workbooks have none, and rejecting those was the whole bug.
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row or 0, 120), values_only=True):
            if _is_header_row(row):
                return True
        return False
    finally:
        wb.close()


def extract_ep_nb_csp_stages(path: str | Path) -> list[CspStage]:
    """Parse EP-NB CSP spreadsheet into per-semester qualification payloads."""
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb.active
        base_title = _qualification_title_from_sheet(ws)
        stages: list[CspStage] = []
        pending_band: str | None = None
        layout: _Layout | None = None
        current: CspClass | None = None
        current_classes: list[CspClass] = []
        unnamed_count = 0

        def flush_stage() -> None:
            nonlocal current, current_classes, pending_band, layout
            if current_classes:
                # A workbook with no bands still yields a stage — everything is
                # flattened into one qualification downstream either way.
                label = pending_band
                name = f"{base_title} – {label}" if label else base_title
                stages.append(
                    CspStage(
                        qualification_name=name,
                        stage_label=label,
                        classes=current_classes,
                    )
                )
            current = None
            current_classes = []
            layout = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            band = _band_label(row)
            if band:
                flush_stage()
                pending_band = band
                continue

            if _is_header_row(row):
                # A repeated header inside the same band continues it; the
                # layout is re-read in case the columns moved.
                classes_so_far = list(current_classes)
                flush_stage()
                current_classes = classes_so_far
                layout = _Layout(row)
                continue

            if layout is None:
                continue
            if _is_subtotal_row(row, layout):
                continue

            tpn = _cell(row, layout.tpn)
            if not tpn or tpn.lower() == "tpn":
                continue

            row_hours = _parse_hours(row[layout.hours] if len(row) > layout.hours else None)
            name = _class_name_from_row(row, layout)

            if layout.names_the_class_in_column_a:
                starts_new_class = bool(_cell(row, layout.label)) or current is None
            else:
                # No label column, so hours mark the first row of a block and
                # continuation rows carry only their unit code.
                starts_new_class = row_hours is not None or current is None

            if starts_new_class:
                if name is None:
                    unnamed_count += 1
                    name = f"{_UNNAMED_PREFIX} {unnamed_count}"
                if current is None or (current.name, current.hours) != (name, row_hours):
                    current = CspClass(name=name, hours=row_hours, unit_codes=[])
                    current_classes.append(current)
            elif row_hours is not None and current.hours is None:
                current.hours = row_hours

            current.unit_codes.append(tpn)

        flush_stage()
        return stages
    finally:
        wb.close()


def import_qualifications_from_ep_nb_csp(
    session: Session,
    path: str | Path,
    *,
    timetable_session_id: int | None = None,
) -> QualImportReport:
    """Create/update qualifications and classes from an EP-NB CSP .xlsx file."""
    rep = QualImportReport()
    path = Path(path)
    if not is_ep_nb_csp_workbook(path):
        raise ValueError(
            "Workbook does not look like a CSP planning spreadsheet "
            "(expected a title on row 1 and a header row naming a TPN column)."
        )

    stages = extract_ep_nb_csp_stages(path)
    if not stages:
        rep.warnings.append(f"No class rows found in {path.name}")
        return rep

    # Say which classes came in without a name, so they can be renamed rather
    # than quietly living as "Unnamed class 3" for a term.
    unnamed = [
        c.name
        for stage in stages
        for c in stage.classes
        if c.name.startswith(_UNNAMED_PREFIX)
    ]
    if unnamed:
        rep.warnings.append(
            f"{len(unnamed)} class(es) had no skill set/description in the workbook "
            f"and were imported with placeholder names ({', '.join(unnamed)}). "
            "Rename them on the Classes tab."
        )

    # Same reasoning as the .docx importer: the workbook's Semester bands are
    # the curriculum's shape, not the timetable's. One qualification in.
    stages = flatten_stages(stages)

    for stage in stages:
        qual_name = stage.qualification_name
        qual = _scoped_filter_by(
            session, Qualification, timetable_session_id, name=qual_name
        ).first()
        if qual is None:
            qual = Qualification(
                **_create_kwargs(
                    Qualification,
                    timetable_session_id,
                    name=qual_name,
                    num_groups=1,
                    schedule_period=SCHEDULE_PERIOD_DAY,
                )
            )
            session.add(qual)
            session.flush()
            replace_qualification_time_windows(session, qual)
            rep.qualifications_created += 1
            default_course_code = f"{qual.name} GrpA"
            existing_course = _scoped_filter_by(
                session, Course, timetable_session_id, code=default_course_code
            ).first()
            if existing_course is None:
                session.add(
                    Course(
                        **_create_kwargs(
                            Course,
                            timetable_session_id,
                            code=default_course_code,
                            qualification_id=qual.id,
                        )
                    )
                )
                rep.courses_created += 1
            elif existing_course.qualification_id != qual.id:
                existing_course.qualification_id = qual.id
        else:
            rep.qualifications_linked += 1

        for cls in stage.classes:
            storage_name = cls.name.strip()
            if not storage_name:
                continue

            component_codes = normalize_component_codes_commas(", ".join(cls.unit_codes))
            existing_unit = _scoped_filter_by(
                session, Unit, timetable_session_id, name=storage_name
            ).first()
            if existing_unit is None:
                unit = Unit(
                    **_create_kwargs(
                        Unit,
                        timetable_session_id,
                        name=storage_name,
                        component_codes=component_codes,
                    )
                )
                session.add(unit)
                session.flush()
                rep.classes_created += 1
            else:
                unit = existing_unit
                rep.classes_updated += 1

            if component_codes and not (unit.component_codes or "").strip():
                unit.component_codes = component_codes
            elif component_codes:
                merged = normalize_component_codes_commas(
                    f"{unit.component_codes}, {component_codes}"
                )
                if merged and merged != (unit.component_codes or "").strip():
                    unit.component_codes = merged

            if cls.hours is not None and not unit.length_slots:
                slots = _parse_running_time(str(cls.hours))
                if slots:
                    unit.length_slots = slots

            link = (
                session.query(UnitQualification)
                .filter_by(unit_id=unit.id, qualification_id=qual.id)
                .first()
            )
            if link is None:
                session.add(UnitQualification(unit_id=unit.id, qualification_id=qual.id))
                rep.class_qual_links_added += 1

    apply_unit_bracket_fields_from_names(session, timetable_session_id=timetable_session_id)
    session.commit()
    return rep
