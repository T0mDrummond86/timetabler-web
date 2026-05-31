"""Bundle v2 export templates from the user-maintained styleguide workbook.

The styleguide file supplies all grid styling (borders, fills, headers). This script
only copies that layout into the three named template tabs and sets placeholder titles.
"""
from __future__ import annotations

import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from .xlsm_export_v2 import (
    V2_LAST_COL,
    V2_SUMMARY_HOURS_COL,
    V2_TEMPLATE_PATH,
    V2_TEMPLATE_SHEETS,
)

STYLEGUIDE_FALLBACK = Path("/Users/tomdrummond/Documents/styleguide2.xlsx")


def _styleguide_path() -> Path:
    for p in (
        STYLEGUIDE_FALLBACK,
        Path(__file__).resolve().parents[2] / "styleguide2.xlsx",
    ):
        if p.is_file():
            return p
    raise FileNotFoundError("styleguide2.xlsx not found")


def _copy_sheet(src, dst, title: str) -> None:
    """Copy layout and styles from the styleguide sheet; set title text only."""
    max_row = max(src.max_row, 34)
    max_col = max(src.max_column, V2_LAST_COL + 2, V2_SUMMARY_HOURS_COL)
    for row in src.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            dst_cell = dst.cell(row=cell.row, column=cell.column)
            if cell.row == 1 and cell.column == 1:
                dst_cell.value = title
            else:
                dst_cell.value = cell.value
            if cell.has_style:
                dst_cell.font = copy(cell.font)
                dst_cell.border = copy(cell.border)
                dst_cell.fill = copy(cell.fill)
                dst_cell.number_format = copy(cell.number_format)
                dst_cell.protection = copy(cell.protection)
                dst_cell.alignment = copy(cell.alignment)
    for merged in src.merged_cells.ranges:
        dst.merge_cells(str(merged))
    dst.freeze_panes = src.freeze_panes
    dst.sheet_format.defaultRowHeight = src.sheet_format.defaultRowHeight
    dst.sheet_format.defaultColWidth = src.sheet_format.defaultColWidth
    for col, dim in src.column_dimensions.items():
        if dim.width is not None:
            dst.column_dimensions[col].width = dim.width
    for row, dim in src.row_dimensions.items():
        if dim.height is not None:
            dst.row_dimensions[row].height = dim.height


def build_v2_templates(
    styleguide: Path | None = None,
    out_path: Path | None = None,
) -> Path:
    styleguide = styleguide or _styleguide_path()
    out_path = out_path or V2_TEMPLATE_PATH
    out_path.parent.mkdir(parents=True, exist_ok=True)

    src_wb = load_workbook(styleguide)
    src = src_wb.active

    if out_path.exists():
        out_path.unlink()
    shutil.copy(styleguide, out_path)
    wb = load_workbook(out_path)
    for name in list(wb.sheetnames):
        del wb[name]
    titles = {
        "Course Template v2": "Course Cluster Name",
        "Staff Template v2": "Lecturer Name",
        "Room Template v2": "Room Name",
    }
    for sheet_name, title in titles.items():
        ws = wb.create_sheet(sheet_name)
        _copy_sheet(src, ws, title)
    wb.save(out_path)
    src_wb.close()
    wb.close()
    return out_path


if __name__ == "__main__":
    path = build_v2_templates()
    print("Wrote", path)
