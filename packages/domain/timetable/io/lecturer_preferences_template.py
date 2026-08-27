"""Build an Excel template for capturing lecturer preferences.

One tab per lecturer in `Staff`, each carrying:
  - six class preferences split into:
      2x first preference, 2x second preference, 2x third preference
  - each preference row has two dropdowns:
      Qualification, then Class (filtered by Qualification)
  - a single non-teaching day picked from a dropdown
  - the number of delivery hours the lecturer is asking for, pre-filled with
    the 21 hours one FTE carries, so most people only have to change it
  - a free-text box for anything the fixed sections cannot say
  - a blocked-times grid (Mon–Sat × half-hour slots), ending at 21:30

A hidden `_classes` sheet holds validation data.

Everything is scoped to one timetable session. On the desktop that is implicit
-- one `*.db` file holds one session -- but the web app keeps every session in
one database, so an unscoped query put every lecturer, qualification and class
in the whole organisation into the workbook. Pass `timetable_session_id` there;
leaving it None keeps the desktop's whole-file behaviour.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from sqlalchemy.orm import Session

from ..constants import DAYS, NUM_DAYS, NUM_SLOTS, slot_to_time
from ..core.models import Qualification, Staff, Unit, UnitQualification


CLASS_LIST_SHEET = "_classes"

#: Hours one FTE delivers, the same figure `Staff.fte` is multiplied by. Almost
#: every lecturer is asking for exactly this, so the cell arrives filled in and
#: the few who want more or less are the only ones who have to type.
DEFAULT_REQUESTED_HOURS = 21

# ---------------------------------------------------------------------------
# Palette
#
# This is a form that goes out to people who do not work in the app, so the
# styling has one job: make it obvious what to read and where to type. Sheet
# gridlines are switched off and the boxes are drawn instead, so the only ruled
# areas on the page are the ones asking for an answer.
# ---------------------------------------------------------------------------
INK = "FF1F3A5F"          # headings and the title banner
RULE = "FF2E5C8A"          # the line under a section heading
HEADER_FILL = "FFE8EDF3"   # table header rows
INPUT_FILL = "FFF4F8FC"    # every cell the lecturer fills in
INPUT_EDGE = "FF8FAAC6"    # and its outline, darker than the hairlines
BAND_FILL = "FFEDF2F8"     # alternating bands in the blocked-times grid
HAIRLINE = "FFD6DEE7"
MUTED = "FF5A6B7C"
#: First preference reads as the strongest of the three, third as the faintest.
PRIORITY_FILLS = {"First": "FFDCE9F7", "Second": "FFE8EFF7", "Third": "FFF2F5F9"}

_HINT_FONT = Font(name="Calibri", size=9, italic=True, color=MUTED)
_HINT_ALIGN = Alignment(horizontal="left", vertical="center")
_CENTRE = Alignment(horizontal="center", vertical="center")


def _fill(colour: str) -> PatternFill:
    return PatternFill(start_color=colour, end_color=colour, fill_type="solid")


def _box(colour: str = HAIRLINE, style: str = "thin") -> Border:
    side = Side(border_style=style, color=colour)
    return Border(top=side, bottom=side, left=side, right=side)


def _input_cell(cell) -> None:
    """A cell someone is being asked to type in."""
    cell.fill = _fill(INPUT_FILL)
    cell.border = _box(INPUT_EDGE)
    cell.alignment = _CENTRE


def _section_heading(ws, row: int, text: str, last_col: int) -> None:
    """A heading with the row to itself and a rule under it.

    Alone on the row on purpose: column A is only wide enough for the grid's
    "08:00-08:30" labels, so a heading with a value beside it is clipped to
    half a word.
    """
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Calibri", size=11, bold=True, color=INK)
    cell.alignment = Alignment(horizontal="left", vertical="bottom")
    ws.row_dimensions[row].height = 22
    underline = Border(bottom=Side(border_style="medium", color=RULE))
    for col in range(1, last_col + 1):
        ws.cell(row=row, column=col).border = underline


def _hint(ws, row: int, col: int, text: str):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = _HINT_FONT
    cell.alignment = _HINT_ALIGN
    return cell


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Excel sheet names: ≤31 chars, no [\\:/?*[]] characters, must be unique."""
    bad = set("[]:*?/\\")
    cleaned = "".join("_" if ch in bad else ch for ch in str(name)).strip()
    cleaned = cleaned[:31] or "Sheet"
    base = cleaned
    n = 1
    while cleaned in used:
        n += 1
        suffix = f" ({n})"
        cleaned = (base[: 31 - len(suffix)] + suffix).strip()
    used.add(cleaned)
    return cleaned


def _populate_class_list_sheet(
    ws,
    qualifications: list[Qualification],
    classes: list[Unit],
    unit_to_quals: dict[int, list[str]],
) -> tuple[int, int]:
    """Hidden sheet backing Qualification/Class dependent dropdowns.

    Returns:
      (qualifications_row_count, qual_class_map_row_count)
    """
    ws["A1"] = "Qualification list"
    ws["A1"].font = Font(bold=True)
    ws["C1"] = "Qualification"
    ws["D1"] = "Class"
    ws["C1"].font = Font(bold=True)
    ws["D1"].font = Font(bold=True)

    qual_names = [q.name for q in qualifications]
    for i, qname in enumerate(qual_names, start=2):
        ws.cell(row=i, column=1, value=qname)

    map_row = 2
    for qname in qual_names:
        q_classes = [u.name for u in classes if qname in unit_to_quals.get(u.id, [])]
        for cname in q_classes:
            ws.cell(row=map_row, column=3, value=qname)
            ws.cell(row=map_row, column=4, value=cname)
            map_row += 1

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 36
    ws.sheet_state = "hidden"
    return len(qual_names), max(0, map_row - 2)


def _populate_lecturer_sheet(ws, staff: Staff, n_quals: int, n_map_rows: int) -> None:
    """Lay out and style one lecturer's tab.

    Three row positions are load-bearing and must not move: the title in A1,
    the six preference rows at 6..11 and the non-teaching day in B14, all of
    which the importer reads by position. Everything below that is found by its
    content -- the blocked-times grid by its time labels -- so the lower half of
    the sheet can be laid out freely.
    """
    last_col = 1 + NUM_DAYS  # A..F; the grid is the widest thing on the sheet
    last_letter = get_column_letter(last_col)

    # Gridlines off: the ruled boxes below are then the only ruled things on
    # the page, which is what makes the answer cells findable at a glance.
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = INK[2:]

    # ---- Title banner ----
    ws.merge_cells(f"A1:{last_letter}1")
    title = ws.cell(row=1, column=1, value=f"Preferences — {staff.name}")
    title.font = Font(name="Tahoma", size=16, bold=True, color="FFFFFFFF")
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(1, last_col + 1):
        ws.cell(row=1, column=col).fill = _fill(INK)
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f"A2:{last_letter}2")
    strap = ws.cell(
        row=2,
        column=1,
        value="Fill in the shaded boxes and return this workbook — every section is optional.",
    )
    strap.font = _HINT_FONT
    strap.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    # ---- Class preferences ----
    _section_heading(ws, 3, "Class preferences", last_col)
    _hint(ws, 4, 1, "Two firsts, two seconds, two thirds. Pick the qualification, then the class.")
    ws.row_dimensions[4].height = 14

    header_row = 5
    for col, label in enumerate(("Priority", "Qualification", "Class"), start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = Font(name="Calibri", size=10, bold=True, color=INK)
        cell.fill = _fill(HEADER_FILL)
        cell.alignment = _CENTRE
        cell.border = _box()
    ws.row_dimensions[header_row].height = 20

    qual_dv = None
    if n_quals > 0:
        qual_formula = f"={CLASS_LIST_SHEET}!$A$2:$A${n_quals + 1}"
        qual_dv = DataValidation(type="list", formula1=qual_formula, allow_blank=True)
        qual_dv.error = "Please pick a qualification from the list."
        qual_dv.errorTitle = "Unknown qualification"
        ws.add_data_validation(qual_dv)

    class_dv = None
    if n_map_rows > 0:
        # Filter class list by selected qualification in column B.
        class_formula = (
            f"=OFFSET({CLASS_LIST_SHEET}!$D$2,"
            f"MATCH($B6,{CLASS_LIST_SHEET}!$C$2:$C${n_map_rows + 1},0)-1,"
            f"0,"
            f"COUNTIF({CLASS_LIST_SHEET}!$C$2:$C${n_map_rows + 1},$B6),1)"
        )
        class_dv = DataValidation(type="list", formula1=class_formula, allow_blank=True)
        class_dv.error = "Please pick a class from the list."
        class_dv.errorTitle = "Unknown class"
        ws.add_data_validation(class_dv)

    priorities = ("First", "First", "Second", "Second", "Third", "Third")
    for idx, priority in enumerate(priorities, start=1):
        row = header_row + idx
        # Repeated rather than merged across the pair: the importer reads the
        # priority from every one of rows 6..11, and the second cell of a merge
        # reads back empty, which would silently drop that preference.
        rank_cell = ws.cell(row=row, column=1, value=priority)
        rank_cell.font = Font(name="Calibri", size=10, bold=True, color=INK)
        rank_cell.alignment = _CENTRE
        rank_cell.fill = _fill(PRIORITY_FILLS[priority])
        rank_cell.border = _box()
        for col, dv in ((2, qual_dv), (3, class_dv)):
            cell = ws.cell(row=row, column=col)
            _input_cell(cell)
            # Names run past the column width, and the grid below fixes that
            # width, so they wrap instead of disappearing under the next cell.
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            if dv is not None:
                dv.add(cell)
        ws.row_dimensions[row].height = 28

    # ---- Non-teaching day ----
    _section_heading(ws, 13, "Non-teaching day", last_col)
    nt_row = 14
    day_dv = DataValidation(
        type="list",
        formula1='"' + ",".join(DAYS) + '"',
        allow_blank=True,
    )
    day_dv.error = "Pick a day of the week."
    day_dv.errorTitle = "Unknown day"
    ws.add_data_validation(day_dv)
    day_cell = ws.cell(row=nt_row, column=2)
    _input_cell(day_cell)
    day_dv.add(day_cell)
    _hint(ws, nt_row, 3, "One weekday, or leave blank.")
    ws.row_dimensions[nt_row].height = 22

    # ---- Requested delivery hours ----
    hours_row = nt_row + 2
    _section_heading(ws, hours_row, "Delivery hours requested", last_col)
    hours_cell = ws.cell(row=hours_row + 1, column=2, value=DEFAULT_REQUESTED_HOURS)
    _input_cell(hours_cell)
    hours_cell.font = Font(name="Calibri", size=11, bold=True, color=INK)
    # General, not "0.#": a literal format code always shows its decimal
    # separator, so 21 rendered as "21." — General gives 21 and 18.5 alike.
    hours_cell.number_format = "General"
    hours_dv = DataValidation(
        type="decimal", operator="between", formula1=0, formula2=40, allow_blank=True
    )
    hours_dv.error = "Enter the number of hours you are asking to deliver each week."
    hours_dv.errorTitle = "Hours out of range"
    ws.add_data_validation(hours_dv)
    hours_dv.add(hours_cell)
    _hint(
        ws,
        hours_row + 1,
        3,
        f"Hours per week. {DEFAULT_REQUESTED_HOURS} is a full load — change it if yours differs.",
    )
    ws.row_dimensions[hours_row + 1].height = 22

    # ---- Additional notes ----
    notes_row = hours_row + 3
    _section_heading(ws, notes_row, "Additional notes", last_col)
    _hint(
        ws,
        notes_row + 1,
        1,
        "Anything the sections above cannot say — job-share, travel between campuses, study leave.",
    )
    ws.row_dimensions[notes_row + 1].height = 14
    # One merged box rather than ruled lines: a lecturer with two sentences and
    # a lecturer with ten both get somewhere obvious to put them.
    notes_top = notes_row + 2
    notes_bottom = notes_top + 5
    ws.merge_cells(f"A{notes_top}:{last_letter}{notes_bottom}")
    for r in range(notes_top, notes_bottom + 1):
        ws.row_dimensions[r].height = 18
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = _fill(INPUT_FILL)
            cell.border = _box(INPUT_EDGE)
    ws.cell(row=notes_top, column=1).alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True, indent=1
    )

    # ---- Blocked times grid ----
    grid_top = notes_bottom + 2
    _section_heading(ws, grid_top, "Blocked times", last_col)
    _hint(ws, grid_top + 1, 1, "Write X in every half-hour you cannot teach. Leave the rest empty.")
    ws.row_dimensions[grid_top + 1].height = 14

    day_header_row = grid_top + 2
    for col, label in enumerate(("Time",) + tuple(DAYS), start=1):
        cell = ws.cell(row=day_header_row, column=col, value=label)
        cell.font = Font(name="Calibri", size=10, bold=True, color=INK)
        cell.fill = _fill(HEADER_FILL)
        cell.alignment = _CENTRE
        cell.border = _box()
    ws.row_dimensions[day_header_row].height = 20

    hour_edge = Side(border_style="thin", color=INPUT_EDGE)
    hairline = Side(border_style="thin", color=HAIRLINE)
    for s in range(max(0, NUM_SLOTS - 1)):
        row = day_header_row + 1 + s
        # Kept exactly as "08:00–08:30": this label is how the importer finds
        # the grid, wherever on the sheet it has ended up.
        time_label = (
            f"{slot_to_time(s).strftime('%H:%M')}–"
            f"{slot_to_time(s + 1).strftime('%H:%M')}"
        )
        starts_hour = s % 2 == 0
        # Banded by the hour, not the half-hour, so a row is easy to follow
        # across five columns without counting lines.
        band = _fill(BAND_FILL) if (s // 2) % 2 else _fill("FFFFFFFF")
        label_cell = ws.cell(row=row, column=1, value=time_label)
        label_cell.alignment = _CENTRE
        label_cell.font = Font(
            name="Calibri", size=9, bold=starts_hour, color=INK if starts_hour else MUTED
        )
        label_cell.fill = _fill(HEADER_FILL) if starts_hour else _fill(BAND_FILL)
        label_cell.border = Border(
            top=hour_edge if starts_hour else hairline,
            bottom=hairline,
            left=hairline,
            right=hour_edge,
        )
        for d in range(NUM_DAYS):
            cell = ws.cell(row=row, column=2 + d)
            cell.fill = band
            cell.alignment = _CENTRE
            cell.font = Font(name="Calibri", size=10, bold=True, color=INK)
            cell.border = Border(
                top=hour_edge if starts_hour else hairline,
                bottom=hairline,
                left=hairline,
                right=hairline,
            )
        ws.row_dimensions[row].height = 15

    # ---- Column widths ----
    # Every column past A is the same width, because the grid and the
    # preference table share them: the grid wants five even day columns, and
    # the preference table's long names wrap to suit rather than forcing B and
    # C wide enough to leave the grid lopsided.
    ws.column_dimensions["A"].width = 15
    for col in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # No frozen panes: the sheet is a form to read top to bottom, and a split
    # under the title only got in the way of that.
    ws.freeze_panes = None

    # ---- Print setup ----
    # It gets printed and marked up on paper as often as it gets typed into.
    ws.print_area = f"A1:{last_letter}{day_header_row + NUM_SLOTS - 1}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def write_lecturer_preferences_template(
    session: Session,
    out_path: str | Path,
    *,
    timetable_session_id: int | None = None,
) -> Path:
    """Build the workbook. Returns the saved path.

    ``timetable_session_id`` restricts the lecturers, qualifications and classes
    to one session. None means the whole database, which is right for a desktop
    file and wrong for anything sharing a database.
    """
    out_path = Path(out_path)
    wb = Workbook()
    # Use the default sheet for the class-list, then add lecturer sheets.
    classes_ws = wb.active
    classes_ws.title = CLASS_LIST_SHEET

    def scoped(query, model):
        if timetable_session_id is None:
            return query
        return query.filter(model.timetable_session_id == timetable_session_id)

    classes = scoped(session.query(Unit), Unit).order_by(Unit.name).all()
    qualifications = (
        scoped(session.query(Qualification), Qualification).order_by(Qualification.name).all()
    )
    # Joined through Qualification rather than filtered on the link table, which
    # carries no session of its own.
    unit_to_quals: dict[int, list[str]] = {}
    for unit_id, qname in (
        scoped(
            session.query(UnitQualification.unit_id, Qualification.name).join(
                Qualification, Qualification.id == UnitQualification.qualification_id
            ),
            Qualification,
        ).all()
    ):
        unit_to_quals.setdefault(unit_id, []).append(qname)
    n_quals, n_map_rows = _populate_class_list_sheet(
        classes_ws, qualifications, classes, unit_to_quals
    )

    used = {CLASS_LIST_SHEET}
    staff_rows = scoped(session.query(Staff), Staff).order_by(Staff.name).all()
    for s in staff_rows:
        sheet_name = _safe_sheet_name(s.name, used)
        ws = wb.create_sheet(sheet_name)
        _populate_lecturer_sheet(ws, s, n_quals, n_map_rows)

    if not staff_rows:
        # Need at least one visible sheet.
        placeholder = wb.create_sheet("(no staff)")
        placeholder["A1"] = "No staff in this session — add lecturers in the app first."

    wb.save(out_path)
    return out_path
