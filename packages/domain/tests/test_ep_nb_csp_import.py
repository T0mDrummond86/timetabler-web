"""Tests for EP-NB CSP (.xlsx) qualification import."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy.orm import sessionmaker

import timetable.core.tenancy_models  # noqa: F401
from timetable.core.models import Base, Qualification, Semester, Unit, UnitQualification, Week
from timetable.core.storage import make_engine
from timetable.core.tenancy_models import Organization, TimetableSession
from timetable.io.ep_nb_csp_import import (
    extract_ep_nb_csp_stages,
    import_qualifications_from_ep_nb_csp,
    is_ep_nb_csp_workbook,
)

_EP_NB_SAMPLE = Path(
    "/Users/tomdrummond/Downloads/!2026 ICT40120 AC10 EP CSP - C4 Net v1.0.xlsx"
)


def _write_minimal_ep_nb(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "ICT40120 - Test Qualification 2026"
    ws["B23"] = "Semester 1"
    ws["A24"] = "BB Shell"
    ws["B24"] = "Hrs in class"
    ws["F24"] = "Skill set/ description"
    ws["G24"] = "SIN"
    ws["H24"] = "TPN"
    ws["I24"] = "UoC(s) being assessed"
    ws["A25"] = "Cluster Alpha"
    ws["B25"] = 3
    ws["H25"] = "VU11111"
    ws["H26"] = "VU22222"
    ws["B38"] = "Semester 2"
    ws["A39"] = "BB Shell"
    ws["B39"] = "Hrs in class"
    ws["F39"] = "Skill set/ description"
    ws["G39"] = "SIN"
    ws["H39"] = "TPN"
    ws["I39"] = "UoC(s) being assessed"
    ws["A40"] = "???"
    ws["B40"] = 2
    ws["F40"] = "Beta Class"
    ws["H40"] = "ICT99999"
    wb.save(path)


@pytest.fixture
def session(tmp_path):
    eng = make_engine(tmp_path / "epnb.db")
    Base.metadata.create_all(eng)
    S = sessionmaker(bind=eng, expire_on_commit=False)
    with S() as s:
        org = Organization(name="Test Org", slug=f"test-{tmp_path.name}")
        s.add(org)
        s.flush()
        ts = TimetableSession(organization_id=org.id, name="Test session")
        s.add(ts)
        s.flush()
        sem = Semester(
            timetable_session_id=ts.id,
            name="Semester 2, 2026",
            num_weeks=18,
            repeating=1,
        )
        s.add(sem)
        s.flush()
        s.add(Week(semester_id=sem.id, week_number=0, label="Repeating week"))
        s.commit()
        s.timetable_session_id = ts.id  # type: ignore[attr-defined]
        yield s


def test_is_ep_nb_csp_workbook_detects_minimal(tmp_path):
    path = tmp_path / "epnb.xlsx"
    _write_minimal_ep_nb(path)
    assert is_ep_nb_csp_workbook(path) is True


def test_extract_minimal_ep_nb_stages(tmp_path):
    path = tmp_path / "epnb.xlsx"
    _write_minimal_ep_nb(path)
    stages = extract_ep_nb_csp_stages(path)
    assert len(stages) == 2
    assert stages[0].stage_label == "Semester 1"
    assert len(stages[0].classes) == 1
    assert stages[0].classes[0].unit_codes == ["VU11111", "VU22222"]
    assert stages[0].classes[0].hours == 3.0
    assert stages[1].classes[0].name == "Beta Class"


def test_import_minimal_ep_nb(session, tmp_path):
    path = tmp_path / "epnb.xlsx"
    _write_minimal_ep_nb(path)
    rep = import_qualifications_from_ep_nb_csp(
        session, path, timetable_session_id=session.timetable_session_id
    )
    # One qualification, not one per band: the bands describe the curriculum's
    # shape, and splitting into stages is a later, manual decision.
    assert rep.qualifications_created == 1
    assert rep.classes_created == 2
    assert session.query(Qualification).count() == 1
    unit = session.query(Unit).filter_by(name="Cluster Alpha").one()
    assert "VU11111" in (unit.component_codes or "")
    assert unit.length_slots == 6


def test_import_ep_nb_sample_when_available(session):
    if not _EP_NB_SAMPLE.is_file():
        pytest.skip("EP-NB CSP sample not available")

    stages = extract_ep_nb_csp_stages(_EP_NB_SAMPLE)
    assert len(stages) == 2
    s1_units = sum(len(c.unit_codes) for c in stages[0].classes)
    s2_units = sum(len(c.unit_codes) for c in stages[1].classes)
    assert s1_units >= 10
    assert s2_units >= 8

    cluster = next(c for c in stages[0].classes if "Introduction to Networks" in c.name)
    assert len(cluster.unit_codes) == 3
    assert cluster.hours == 5.0

    rep = import_qualifications_from_ep_nb_csp(
        session, _EP_NB_SAMPLE, timetable_session_id=session.timetable_session_id
    )
    # The two semester bands are read (asserted above) but land in a single
    # qualification; stage splitting is a decision made later, by hand.
    assert rep.qualifications_created == 1
    # 11 classes carrying 20 unit codes — verified against this workbook, and
    # unchanged by the layout work (the previous >= 15 never matched it).
    assert rep.classes_created == 11

    win = session.query(Unit).filter(Unit.name.ilike("%Windows Desktop%")).one()
    assert "ICTNWK422" in (win.component_codes or "")


# ---------------------------------------------------------------------------
# The "lecturer" layout — no BB Shell column, bands written as "Part N", and
# the class name coming from the skill-set description (blank in real files).
# ---------------------------------------------------------------------------


def _write_lecturer_layout(path: Path, *, skill_set: str | None = None) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["B1"] = "AE780 Transition to Cyber Security"   # title in B, not A
    ws["B3"] = "Part 1 •\tUnassigned – Thursday night"
    for col, label in zip("ABCDEFGHIJ", [
        "Lecturer(s)", "Hrs in class", "Actual", "Effic.", "Total NH",
        "Skill set/ description", "SIN", "TPN", "UoC(s) being assessed", "Core / Elect",
    ]):
        ws[f"{col}4"] = label
    ws["A5"] = "??"
    ws["B5"] = 3
    if skill_set:
        ws["F5"] = skill_set
    ws["G5"] = "AZ343"
    ws["H5"] = "VU23214"
    ws["H6"] = "VU23217"          # continuation row: unit code only
    ws["B7"] = 3                   # part subtotal, no TPN
    ws["B8"] = "Part 2 •\tJohn Robertson – Monday night"
    for col, label in zip("ABCDEFGHIJ", [
        "Lecturer(s)", "Hrs in class", "Actual", "Effic.", "Total NH",
        "Skill set/ description", "SIN", "TPN", "UoC(s) being assessed", "Core / Elect",
    ]):
        ws[f"{col}9"] = label
    ws["A10"] = "John Robertson"
    ws["B10"] = 2
    ws["H10"] = "ICTPRG435"
    ws["H11"] = "VU23216"
    wb.save(path)


def test_lecturer_layout_is_recognised(tmp_path):
    path = tmp_path / "lecturer.xlsx"
    _write_lecturer_layout(path)

    # Used to be rejected outright for having no Semester bands and no BB Shell.
    assert is_ep_nb_csp_workbook(path) is True


def test_lecturer_layout_reads_title_from_anywhere_on_row_one(tmp_path):
    path = tmp_path / "lecturer.xlsx"
    _write_lecturer_layout(path)

    stages = extract_ep_nb_csp_stages(path)

    assert all("AE780 Transition to Cyber Security" in s.qualification_name for s in stages)


def test_part_bands_are_read_like_semesters(tmp_path):
    path = tmp_path / "lecturer.xlsx"
    _write_lecturer_layout(path)

    stages = extract_ep_nb_csp_stages(path)

    assert [s.stage_label for s in stages] == ["Part 1", "Part 2"]


def test_class_name_comes_from_the_skill_set_description(tmp_path):
    path = tmp_path / "lecturer.xlsx"
    _write_lecturer_layout(path, skill_set="Cyber Fundamentals")

    stages = extract_ep_nb_csp_stages(path)

    # Column A holds the lecturer here, so it must not become the class name.
    assert stages[0].classes[0].name == "Cyber Fundamentals"


def test_a_blank_skill_set_gets_a_placeholder_rather_than_failing(tmp_path):
    path = tmp_path / "lecturer.xlsx"
    _write_lecturer_layout(path)  # no skill set at all, as in the real file

    stages = extract_ep_nb_csp_stages(path)
    names = [c.name for s in stages for c in s.classes]

    assert names == ["Unnamed class 1", "Unnamed class 2"]
    assert "??" not in names  # the lecturer column is not a fallback


def test_multi_unit_blocks_and_hours_survive_the_lecturer_layout(tmp_path):
    path = tmp_path / "lecturer.xlsx"
    _write_lecturer_layout(path)

    stages = extract_ep_nb_csp_stages(path)
    first, second = stages[0].classes[0], stages[1].classes[0]

    assert first.unit_codes == ["VU23214", "VU23217"]
    assert first.hours == 3
    assert second.unit_codes == ["ICTPRG435", "VU23216"]
    assert second.hours == 2


def test_placeholder_names_are_reported_as_a_warning(session, tmp_path):
    path = tmp_path / "lecturer.xlsx"
    _write_lecturer_layout(path)

    rep = import_qualifications_from_ep_nb_csp(
        session, path, timetable_session_id=session.timetable_session_id
    )

    assert rep.classes_created == 2
    assert any("placeholder names" in w for w in rep.warnings)
