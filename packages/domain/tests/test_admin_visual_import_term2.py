"""Admin visual import: term-2-only rows (labels in term-2 columns only)."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy.orm import sessionmaker

from timetable.core.models import Base, Booking, Course, Semester, Week
from timetable.core.tenancy_models import Organization, TimetableSession
from timetable.core.storage import make_engine
from timetable.io.admin_export import (
    TERM2_LABEL_COLS,
    TERM2_WEEK_COLS,
    _COURSE_TITLE_COL,
    _COURSE_TITLE_ROW,
)
from timetable.io.admin_visual_import import import_admin_visual


@pytest.fixture
def session(tmp_path):
    eng = make_engine(tmp_path / "admin_term2.db")
    Base.metadata.create_all(eng)
    S = sessionmaker(bind=eng, expire_on_commit=False)
    with S() as s:
        org = Organization(name="Test Org", slug="test-org")
        s.add(org)
        s.flush()
        ts = TimetableSession(name="S2", organization_id=org.id)
        s.add(ts)
        s.flush()
        sem = Semester(
            timetable_session_id=ts.id,
            name="Semester 2, 2026",
            num_weeks=18,
            repeating=1,
        )
        s.add(sem)
        s.flush()
        s.add(Week(semester_id=sem.id, week_number=0, label="Repeating week"))
        s.commit()
        s._timetable_session_id = ts.id  # type: ignore[attr-defined]
        yield s


def _term2_only_workbook(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "T2ONLY"
    ws.cell(_COURSE_TITLE_ROW, _COURSE_TITLE_COL, value="Course: T2ONLY")
    row = 10
    ws.cell(row, TERM2_LABEL_COLS[0], value="09:30-12:00")
    ws.cell(row, TERM2_LABEL_COLS[1], value="Jeff Hoyle")
    ws.cell(row, TERM2_LABEL_COLS[2], value="A114")
    label = "Cyber lab (VU23220) ID: 4401542"
    ws.merge_cells(
        start_row=row,
        end_row=row,
        start_column=TERM2_WEEK_COLS[0],
        end_column=TERM2_WEEK_COLS[-1],
    )
    ws.cell(row, TERM2_WEEK_COLS[0], value=label)
    path = tmp_path / "admin_term2_only.xlsx"
    wb.save(path)
    return path


def test_admin_visual_import_term2_only_row(session, tmp_path):
    path = _term2_only_workbook(tmp_path)
    report = import_admin_visual(
        session, str(path), timetable_session_id=session._timetable_session_id
    )
    assert report.bookings_created == 1

    course = session.query(Course).filter(Course.code == "T2ONLY").one()
    booking = session.query(Booking).filter(Booking.course_id == course.id).one()
    assert booking.in_term_1 == 0
    assert booking.in_term_2 == 1
    assert booking.external_id == "4401542"
    assert booking.staff.name == "Jeff Hoyle"
