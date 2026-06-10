"""Tests for aSc Timetables export import."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy.orm import sessionmaker

import timetable.core.tenancy_models  # noqa: F401 — register web tenancy tables
from timetable.core.models import Base, Booking, Course, Qualification, Room, Semester, Staff, Unit, Week
from timetable.core.storage import make_engine
from timetable.core.tenancy_models import Organization, TimetableSession
from timetable.io.asc_import import import_asc_export, is_asc_export_workbook

_ASC_SAMPLE = Path("/Users/tomdrummond/Downloads/asc export.xlsx")


def _write_minimal_asc(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    teachers = wb.create_sheet("Teachers")
    teachers.append([None, "Name", "Short"])
    teachers.append([None, "Ada Lovelace", "AL"])

    rooms = wb.create_sheet("Classrooms")
    rooms.append([None, "Name", "Short"])
    rooms.append([None, "Room 101", "R101"])

    classes = wb.create_sheet("Classes")
    classes.append(["Class", "Short"])
    classes.append(["Diploma Stage 1", "DIP1"])

    subjects = wb.create_sheet("Subjects")
    subjects.append([None, "Subject", "Short"])
    subjects.append([None, "Cataloguing", "BSBINS305"])

    lessons = wb.create_sheet("Lessons")
    lessons.append(["Teacher", "Class", "Group", "Subject", "Length", "Lessons/week", "Available classrooms"])
    lessons.append(["Ada Lovelace", "Diploma Stage 1", "Entire class", "BSBINS305", "6", "1", "R101"])

    schedule = wb.create_sheet("BSBINS305 sched")
    schedule.append([None, None, None])
    schedule.append([None, "Day", "Lesson"])
    schedule.append([None, "Monday", "1"])
    schedule.append([None, "Monday", "2", "Diploma Stage 1"])
    schedule.append([None, "Monday", "3", "Diploma Stage 1"])

    wb.save(path)


@pytest.fixture
def session(tmp_path):
    eng = make_engine(tmp_path / "asc.db")
    Base.metadata.create_all(eng)
    S = sessionmaker(bind=eng, expire_on_commit=False)
    with S() as s:
        org = Organization(name="Test Org", slug=f"test-{tmp_path.name}")
        s.add(org)
        s.flush()
        ts = TimetableSession(organization_id=org.id, name="Test session")
        s.add(ts)
        s.flush()
        sem = Semester(
            timetable_session_id=ts.id,
            name="Semester 2, 2026",
            num_weeks=18,
            repeating=1,
        )
        s.add(sem)
        s.flush()
        s.add(Week(semester_id=sem.id, week_number=0, label="Repeating week"))
        s.commit()
        s.timetable_session_id = ts.id  # type: ignore[attr-defined]
        yield s


def test_is_asc_export_workbook_detects_minimal(tmp_path):
    path = tmp_path / "asc.xlsx"
    _write_minimal_asc(path)
    assert is_asc_export_workbook(path) is True


def test_is_asc_export_workbook_rejects_empty(tmp_path):
    path = tmp_path / "empty.xlsx"
    Workbook().save(path)
    assert is_asc_export_workbook(path) is False


def test_import_minimal_asc(session, tmp_path):
    path = tmp_path / "asc.xlsx"
    _write_minimal_asc(path)
    rep = import_asc_export(session, path, timetable_session_id=session.timetable_session_id)

    assert rep.staff_created == 1
    assert rep.rooms_created == 1
    assert rep.qualifications_created == 1
    assert rep.classes_created == 1
    assert rep.lecturer_links_added == 1
    assert rep.room_links_added == 1
    assert rep.bookings_created == 1

    sid = session.timetable_session_id
    assert (
        session.query(Staff)
        .filter(Staff.timetable_session_id == sid, Staff.name == "Ada Lovelace")
        .one()
    )
    assert session.query(Room).filter(Room.timetable_session_id == sid, Room.code == "R101").one()
    assert (
        session.query(Qualification)
        .filter(Qualification.timetable_session_id == sid, Qualification.name == "Diploma Stage 1")
        .one()
    )
    unit = session.query(Unit).filter(Unit.timetable_session_id == sid).one()
    assert unit.component_codes == "BSBINS305"
    assert unit.length_slots == 6
    assert session.query(Course).filter(Course.timetable_session_id == sid, Course.code == "DIP1").one()
    booking = session.query(Booking).one()
    assert booking.day == 0
    assert booking.start_slot == 1
    assert booking.end_slot == 3


def test_import_asc_sample_when_available(session):
    if not _ASC_SAMPLE.is_file():
        pytest.skip("ASC sample not available")

    rep = import_asc_export(session, _ASC_SAMPLE, timetable_session_id=session.timetable_session_id)

    assert rep.staff_created >= 9
    assert rep.rooms_created >= 15
    assert rep.qualifications_created >= 6
    assert rep.classes_created >= 30
    assert rep.bookings_created >= 30

    sid = session.timetable_session_id
    units = (
        session.query(Unit)
        .filter(Unit.timetable_session_id == sid, Unit.component_codes.ilike("BSBTEC201"))
        .all()
    )
    assert len(units) >= 2
